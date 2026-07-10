import csv
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from groupcore.models import MemberProfile, SavingsAccount
from deposits.models import DepositSubmission
from loans.models import LoanGuarantorApproval, LoanRepayment, LoanRequest
from messaging.models import MessageRecipient


class LoanGuarantorWorkflowTests(TestCase):
    def setUp(self):
        self.applicant = MemberProfile.objects.create_user(
            username='applicant',
            password='pass12345',
            role='MEMBER',
        )
        self.account = SavingsAccount.objects.create(owner=self.applicant, label='A')
        self.guarantors = [
            MemberProfile.objects.create_user(
                username=f'guarantor{i}',
                password='pass12345',
                role='MEMBER',
            )
            for i in range(1, 4)
        ]
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.secretary = MemberProfile.objects.create_user(
            username='loansecretary',
            password='pass12345',
            role='SECRETARY',
        )
        self.chairman = MemberProfile.objects.create_user(
            username='chairman',
            password='pass12345',
            role='CHAIRMAN',
        )

    def submit_loan(self):
        self.client.login(username='applicant', password='pass12345')
        response = self.client.post(reverse('request_loan'), {
            'account': str(self.account.id),
            'principal': '100000',
            'duration_months': '6',
            'purpose': 'School fees',
            'guarantors': [str(guarantor.id) for guarantor in self.guarantors],
        }, follow=True)
        self.assertEqual(response.status_code, 200)
        return LoanRequest.objects.get(member=self.applicant)

    def test_request_loan_page_has_guarantor_search(self):
        self.client.login(username='applicant', password='pass12345')

        response = self.client.get(reverse('request_loan'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="guarantor-search"')
        self.assertContains(response, 'placeholder="Search members"')
        self.assertContains(response, 'data-guarantor-option')

    def decide_as_guarantor(self, guarantor, loan, decision, comments=''):
        self.client.logout()
        self.client.login(username=guarantor.username, password='pass12345')
        approval = LoanGuarantorApproval.objects.get(loan=loan, guarantor=guarantor)
        return self.client.post(reverse('guarantor_request_detail', args=[approval.id]), {
            'decision': decision,
            'comments': comments,
        }, follow=True)

    def test_new_loan_waits_for_three_guarantors_before_management(self):
        loan = self.submit_loan()

        self.assertEqual(loan.status, LoanRequest.STATUS_PENDING_GUARANTOR)
        self.assertEqual(loan.guarantor_approvals.count(), 3)
        self.assertEqual(
            MessageRecipient.objects.filter(
                recipient__in=self.guarantors,
                message__subject__contains='Guarantor approval needed',
            ).count(),
            3,
        )

        self.client.logout()
        self.client.login(username='treasurer', password='pass12345')
        queue_response = self.client.get(reverse('pending_loans'))
        self.assertContains(queue_response, "Awaiting Guarantors' Approval")
        blocked_response = self.client.post(reverse('approve_loan', args=[loan.id]))
        self.assertEqual(blocked_response.status_code, 404)

        self.decide_as_guarantor(self.guarantors[0], loan, 'approve', 'Looks fine')
        loan.refresh_from_db()
        self.assertEqual(loan.status, LoanRequest.STATUS_PENDING_GUARANTOR)

        self.decide_as_guarantor(self.guarantors[1], loan, 'approve')
        loan.refresh_from_db()
        self.assertEqual(loan.status, LoanRequest.STATUS_PENDING_GUARANTOR)

        self.decide_as_guarantor(self.guarantors[2], loan, 'approve')
        loan.refresh_from_db()
        self.assertEqual(loan.status, LoanRequest.STATUS_PENDING)
        self.assertEqual(loan.guarantor_approval_count, 3)
        self.assertTrue(
            MessageRecipient.objects.filter(
                recipient=self.treasurer,
                message__subject__contains='Loan ready for management review',
            ).exists()
        )
        self.assertTrue(
            MessageRecipient.objects.filter(
                recipient=self.secretary,
                message__subject__contains='Loan ready for management review',
            ).exists()
        )
        self.assertTrue(
            MessageRecipient.objects.filter(
                recipient=self.chairman,
                message__subject__contains='Loan ready for management review',
            ).exists()
        )

    def test_guarantor_rejection_stops_workflow(self):
        loan = self.submit_loan()

        self.decide_as_guarantor(self.guarantors[0], loan, 'reject', 'Savings are too low')
        loan.refresh_from_db()

        self.assertEqual(loan.status, LoanRequest.STATUS_REJECTED_GUARANTOR)
        self.assertIn('Savings are too low', loan.remarks)
        self.assertTrue(
            MessageRecipient.objects.filter(
                recipient=self.applicant,
                message__subject__contains='Loan rejected by guarantor',
            ).exists()
        )

        self.decide_as_guarantor(self.guarantors[1], loan, 'approve')
        loan.refresh_from_db()
        self.assertEqual(loan.status, LoanRequest.STATUS_REJECTED_GUARANTOR)


class HistoricalLoanImportCommandTests(TestCase):
    def _write_members_csv(self, directory):
        members_path = Path(directory) / 'members.csv'
        members_path.write_text(
            '\n'.join([
                'username,first_name,last_name,email,phone_number,role,account_labels,is_active',
                'member_one,Member,One,one@example.com,+256700000001,MEMBER,Member One,true',
                'member_two,Member,Two,two@example.com,+256700000002,MEMBER,Member Two,true',
            ]),
            encoding='utf-8',
        )
        return members_path

    def _write_loan_workbook(self, directory):
        workbook_path = Path(directory) / 'loan_schedule.xlsx'
        workbook = Workbook()
        summary = workbook.active
        summary.title = 'Summary'

        for title in ['January', 'February']:
            worksheet = workbook.create_sheet(title)
            worksheet.cell(row=1, column=3, value='Date of Loan Issue')
            worksheet.cell(row=1, column=4, value='Principal')
            worksheet.cell(row=1, column=5, value='First Instalment')
            worksheet.cell(row=1, column=7, value='Second Installment')
            worksheet.cell(row=2, column=1, value='S/No')
            worksheet.cell(row=2, column=2, value='Name')
            worksheet.cell(row=2, column=5, value='Amount Paid ')
            worksheet.cell(row=2, column=6, value='Amount Due ')
            worksheet.cell(row=2, column=7, value='Amount Paid ')
            worksheet.cell(row=2, column=8, value='Amount Due ')

        january = workbook['January']
        january.cell(row=3, column=1, value=1)
        january.cell(row=3, column=2, value='Member One')
        january.cell(row=3, column=3, value=datetime(2026, 9, 1))
        january.cell(row=3, column=4, value=400000)
        january.cell(row=3, column=5, value=408000)
        january.cell(row=3, column=6, value=0)
        january.cell(row=4, column=2, value='Total')
        january.cell(row=4, column=4, value=999999)
        january.cell(row=4, column=5, value=999999)

        february = workbook['February']
        february.cell(row=3, column=1, value=1)
        february.cell(row=3, column=2, value='Member Two')
        february.cell(row=3, column=3, value=datetime(2026, 3, 2))
        february.cell(row=3, column=4, value=2000000)
        february.cell(row=3, column=5, value=300000)
        february.cell(row=3, column=6, value=1740000)
        february.cell(row=3, column=7, value=1350000)
        february.cell(row=3, column=8, value=424800)

        workbook.save(workbook_path)
        return workbook_path

    def _prepare_files(self, directory):
        members_path = self._write_members_csv(directory)
        workbook_path = self._write_loan_workbook(directory)
        loans_path = Path(directory) / 'loans.csv'
        repayments_path = Path(directory) / 'repayments.csv'
        review_path = Path(directory) / 'review.csv'
        balance_path = Path(directory) / 'balance.csv'

        call_command(
            'prepare_loan_schedule_import',
            workbook=str(workbook_path),
            members=str(members_path),
            loans=str(loans_path),
            repayments=str(repayments_path),
            review=str(review_path),
            balance_compare=str(balance_path),
            year=2026,
        )
        return loans_path, repayments_path, review_path, balance_path

    def test_prepare_loan_schedule_import_writes_ready_rows_and_corrects_dates(self):
        with TemporaryDirectory() as directory:
            loans_path, repayments_path, review_path, _balance_path = self._prepare_files(directory)

            with loans_path.open(newline='', encoding='utf-8') as handle:
                loans = list(csv.DictReader(handle))
            with repayments_path.open(newline='', encoding='utf-8') as handle:
                repayments = list(csv.DictReader(handle))

            self.assertEqual(len(loans), 2)
            self.assertEqual(len(repayments), 3)
            self.assertEqual(loans[0]['issue_date'], '2026-01-09')
            self.assertEqual(loans[1]['issue_date'], '2026-02-03')
            self.assertEqual(loans[0]['duration_months'], '')
            self.assertIn('READY', review_path.read_text(encoding='utf-8'))
            self.assertNotIn('Total', loans_path.read_text(encoding='utf-8'))

    def test_import_historical_loans_creates_loans_repayments_and_deposit_rows(self):
        member_one = MemberProfile.objects.create_user(username='member_one', password='pass12345')
        member_two = MemberProfile.objects.create_user(username='member_two', password='pass12345')
        treasurer = MemberProfile.objects.create_user(username='treasurer', password='pass12345', role='TREASURER')
        SavingsAccount.objects.create(owner=member_one, label='Member One')
        SavingsAccount.objects.create(owner=member_two, label='Member Two')

        with TemporaryDirectory() as directory:
            loans_path, repayments_path, _review_path, _balance_path = self._prepare_files(directory)
            report_path = Path(directory) / 'import_report.csv'

            call_command(
                'import_historical_loans',
                loans=str(loans_path),
                repayments=str(repayments_path),
                report=str(report_path),
            )
            self.assertEqual(LoanRequest.objects.count(), 0)

            call_command(
                'import_historical_loans',
                loans=str(loans_path),
                repayments=str(repayments_path),
                report=str(report_path),
                submitted_by=treasurer.username,
                commit=True,
            )

            self.assertEqual(LoanRequest.objects.count(), 2)
            self.assertEqual(LoanRepayment.objects.count(), 3)
            self.assertEqual(DepositSubmission.objects.filter(loan_repayment_amount__gt=0).count(), 3)
            loan = LoanRequest.objects.get(account__label='Member One')
            self.assertEqual(loan.status, LoanRequest.STATUS_APPROVED)
            self.assertIsNone(loan.duration_months)
            self.assertEqual(loan.repayment_status, 'FULLY_PAID')

            call_command(
                'import_historical_loans',
                loans=str(loans_path),
                repayments=str(repayments_path),
                report=str(report_path),
                submitted_by=treasurer.username,
                commit=True,
            )
            self.assertEqual(LoanRequest.objects.count(), 2)
            self.assertEqual(LoanRepayment.objects.count(), 3)
            self.assertEqual(DepositSubmission.objects.filter(loan_repayment_amount__gt=0).count(), 3)
            self.assertIn('SKIPPED_DUPLICATE', report_path.read_text(encoding='utf-8'))
