import csv
import hashlib
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from openpyxl import load_workbook

from groupcore.models import SavingsAccount


MONTHS = {
    'january': 1,
    'february': 2,
    'march': 3,
    'april': 4,
    'may': 5,
    'june': 6,
    'july': 7,
    'august': 8,
    'september': 9,
    'october': 10,
    'november': 11,
    'december': 12,
}
DATE_FORMATS = ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d')
LOAN_FIELDNAMES = [
    'loan_reference',
    'username',
    'account_label',
    'issue_date',
    'approved_on',
    'principal',
    'monthly_interest_rate',
    'duration_months',
    'status',
    'purpose',
    'remarks',
    'source_sheet',
    'source_row',
    'workbook_final_due',
]
REPAYMENT_FIELDNAMES = [
    'repayment_reference',
    'deposit_reference',
    'loan_reference',
    'username',
    'account_label',
    'paid_on',
    'payment_week',
    'payment_time',
    'amount',
    'notes',
    'source_sheet',
    'source_row',
    'source_column',
    'workbook_due_after_payment',
]
REVIEW_FIELDNAMES = [
    'status',
    'issues',
    'sheet',
    'source_row',
    'workbook_account_label',
    'mapped_username',
    'mapped_account_label',
    'loan_reference',
    'issue_date',
    'principal',
    'repayment_count',
    'repayment_total',
    'date_note',
]
BALANCE_FIELDNAMES = [
    'loan_reference',
    'account_label',
    'issue_date',
    'principal',
    'repayment_total',
    'comparison_date',
    'system_balance',
    'workbook_due_after_last_payment',
    'difference',
    'workbook_final_due',
]


class Command(BaseCommand):
    help = 'Prepare import-ready loans and repayments CSVs from the historical loan schedule workbook.'

    def add_arguments(self, parser):
        parser.add_argument('--workbook', required=True, help='Loan schedule workbook, for example Loan Schedule 2026.xlsx.')
        parser.add_argument('--members', help='Members CSV containing username and account_labels. Uses database accounts if omitted.')
        parser.add_argument('--loans', default='imports/loans_from_schedule_2026.csv')
        parser.add_argument('--repayments', default='imports/loan_repayments_from_schedule_2026.csv')
        parser.add_argument('--review', default='imports/loan_schedule_2026_review.csv')
        parser.add_argument('--balance-compare', default='imports/loan_schedule_2026_balance_compare.csv')
        parser.add_argument('--year', type=int, default=2026)
        parser.add_argument('--monthly-rate', default='2.00')

    def handle(self, *args, **options):
        workbook_path = Path(options['workbook'])
        if not workbook_path.exists():
            raise CommandError(f'Workbook not found: {workbook_path}')

        rate = self._parse_amount(options['monthly_rate'])
        account_map = self._load_account_map(options.get('members'))
        wb = load_workbook(workbook_path, data_only=True, read_only=False)

        loans = []
        repayments = []
        review_rows = []
        balance_rows = []
        repayment_groups = defaultdict(list)

        for worksheet in wb.worksheets:
            sheet_month = MONTHS.get(worksheet.title.strip().lower())
            if not sheet_month:
                continue
            sheet_rows = self._prepare_sheet(
                worksheet=worksheet,
                sheet_month=sheet_month,
                year=options['year'],
                rate=rate,
                account_map=account_map,
            )
            loans.extend(sheet_rows['loans'])
            repayments.extend(sheet_rows['repayments'])
            review_rows.extend(sheet_rows['review_rows'])
            for repayment in sheet_rows['repayments']:
                repayment_groups[repayment['loan_reference']].append(repayment)

        for loan in loans:
            loan_repayments = repayment_groups.get(loan['loan_reference'], [])
            balance_rows.append(self._balance_row(loan, loan_repayments, rate))

        self._write_csv(Path(options['loans']), LOAN_FIELDNAMES, loans)
        self._write_csv(Path(options['repayments']), REPAYMENT_FIELDNAMES, repayments)
        self._write_csv(Path(options['review']), REVIEW_FIELDNAMES, review_rows)
        self._write_csv(Path(options['balance_compare']), BALANCE_FIELDNAMES, balance_rows)

        status_counts = defaultdict(int)
        for row in review_rows:
            status_counts[row['status']] += 1
        status_text = ', '.join(f'{status}: {count}' for status, count in sorted(status_counts.items()))
        self.stdout.write(
            self.style.SUCCESS(
                f"Prepared {len(loans)} loan row(s) and {len(repayments)} repayment row(s). "
                f"Review {len(review_rows)} workbook row(s). Status counts: {status_text}."
            )
        )

    def _prepare_sheet(self, worksheet, sheet_month, year, rate, account_map):
        columns = self._find_columns(worksheet)
        missing_columns = []
        if not columns['issue_col']:
            missing_columns.append('Date of Loan Issue')
        if not columns['principal_col']:
            missing_columns.append('Principal')
        if not columns['paid_cols']:
            missing_columns.append('Amount Paid')
        if missing_columns:
            raise CommandError(f"{worksheet.title}: missing required column(s): {', '.join(missing_columns)}")

        prepared_loans = []
        prepared_repayments = []
        review_rows = []
        sheet_month_end = date(year, sheet_month, monthrange(year, sheet_month)[1])
        sheet_payment_week = self._last_friday(year, sheet_month)

        for row_number in range(3, worksheet.max_row + 1):
            workbook_label = self._normalize_label(worksheet.cell(row_number, 2).value)
            if not workbook_label or workbook_label.lower() == 'total':
                continue

            principal = self._parse_amount(worksheet.cell(row_number, columns['principal_col']).value)
            paid_values = [
                (column, self._parse_amount(worksheet.cell(row_number, column).value))
                for column in columns['paid_cols']
            ]
            repayment_total = sum((amount for _column, amount in paid_values if amount > 0), Decimal('0'))
            if principal <= 0 and repayment_total <= 0:
                continue

            issues = []
            mapping = account_map.get(workbook_label.lower())
            if not mapping:
                issues.append('account label not found in members/account source')

            issue_date, date_note = self._parse_issue_date(
                worksheet.cell(row_number, columns['issue_col']).value,
                sheet_month=sheet_month,
                default_year=year,
            )
            if not issue_date and principal > 0:
                issues.append('loan issue date is required when principal is present')
            if principal <= 0 and repayment_total > 0:
                issues.append('repayment amount has no same-row principal')

            loan_reference = ''
            if issue_date and principal > 0:
                loan_reference = self._loan_reference(worksheet.title, row_number, workbook_label, issue_date, principal)

            row_repayments = []
            for column, amount in paid_values:
                if amount <= 0:
                    continue
                repayment_reference = self._repayment_reference(
                    worksheet.title,
                    row_number,
                    column,
                    workbook_label,
                    amount,
                )
                row_repayments.append({
                    'repayment_reference': repayment_reference,
                    'deposit_reference': f'DEP-{repayment_reference}',
                    'loan_reference': loan_reference,
                    'username': mapping['username'] if mapping else '',
                    'account_label': mapping['account_label'] if mapping else workbook_label,
                    'paid_on': sheet_month_end.isoformat(),
                    'payment_week': sheet_payment_week.isoformat(),
                    'payment_time': '00:00',
                    'amount': self._format_amount(amount),
                    'notes': f'Historical loan repayment from {worksheet.title}, row {row_number}.',
                    'source_sheet': worksheet.title,
                    'source_row': row_number,
                    'source_column': self._column_letter(column),
                    'workbook_due_after_payment': self._format_amount(
                        self._due_after_paid_column(worksheet, row_number, column, columns['due_cols'])
                    ),
                })

            if principal > 0 and not issues:
                prepared_loans.append({
                    'loan_reference': loan_reference,
                    'username': mapping['username'],
                    'account_label': mapping['account_label'],
                    'issue_date': issue_date.isoformat(),
                    'approved_on': issue_date.isoformat(),
                    'principal': self._format_amount(principal),
                    'monthly_interest_rate': self._format_amount(rate),
                    'duration_months': '',
                    'status': 'APPROVED',
                    'purpose': 'Historical loan import',
                    'remarks': f'Historical loan imported from {worksheet.title}, row {row_number}.',
                    'source_sheet': worksheet.title,
                    'source_row': row_number,
                    'workbook_final_due': self._format_amount(
                        self._last_nonzero_due(worksheet, row_number, columns['due_cols'])
                    ),
                })
                prepared_repayments.extend(row_repayments)

            review_rows.append({
                'status': 'READY' if not issues else 'ERROR',
                'issues': '; '.join(issues),
                'sheet': worksheet.title,
                'source_row': row_number,
                'workbook_account_label': workbook_label,
                'mapped_username': mapping['username'] if mapping else '',
                'mapped_account_label': mapping['account_label'] if mapping else '',
                'loan_reference': loan_reference,
                'issue_date': issue_date.isoformat() if issue_date else '',
                'principal': self._format_amount(principal),
                'repayment_count': len(row_repayments),
                'repayment_total': self._format_amount(repayment_total),
                'date_note': date_note,
            })

        return {
            'loans': prepared_loans,
            'repayments': prepared_repayments,
            'review_rows': review_rows,
        }

    def _find_columns(self, worksheet):
        issue_col = None
        principal_col = None
        paid_cols = []
        due_cols = []
        for column in range(1, worksheet.max_column + 1):
            heading_one = str(worksheet.cell(1, column).value or '').strip().lower()
            heading_two = str(worksheet.cell(2, column).value or '').strip().lower()
            if heading_one == 'date of loan issue':
                issue_col = column
            if heading_one == 'principal':
                principal_col = column
            if heading_two == 'amount paid':
                paid_cols.append(column)
            if heading_two == 'amount due':
                due_cols.append(column)
        return {
            'issue_col': issue_col,
            'principal_col': principal_col,
            'paid_cols': paid_cols,
            'due_cols': due_cols,
        }

    def _load_account_map(self, members_path):
        if members_path:
            path = Path(members_path)
            if not path.exists():
                raise CommandError(f'Members file not found: {path}')
            account_map = {}
            with path.open(newline='', encoding='utf-8-sig') as handle:
                reader = csv.DictReader(handle)
                required = {'username', 'account_labels'}
                missing = required.difference(reader.fieldnames or [])
                if missing:
                    raise CommandError(f'Members file missing column(s): {", ".join(sorted(missing))}')
                for row in reader:
                    username = (row.get('username') or '').strip()
                    for label in re.split(r'[;|\n]+', row.get('account_labels') or ''):
                        clean_label = self._normalize_label(label)
                        if clean_label:
                            account_map[clean_label.lower()] = {
                                'username': username,
                                'account_label': clean_label,
                            }
            return account_map

        return {
            account.label.lower(): {
                'username': account.owner.username,
                'account_label': account.label,
            }
            for account in SavingsAccount.objects.select_related('owner').filter(is_active=True)
        }

    def _parse_issue_date(self, value, sheet_month, default_year):
        if isinstance(value, datetime):
            raw = value.date()
            corrected = self._correct_excel_date(raw, sheet_month, default_year)
            if corrected != raw:
                return corrected, f'corrected Excel date {raw.isoformat()} as DD/MM/YYYY'
            return raw, ''
        if isinstance(value, date):
            corrected = self._correct_excel_date(value, sheet_month, default_year)
            if corrected != value:
                return corrected, f'corrected Excel date {value.isoformat()} as DD/MM/YYYY'
            return value, ''

        text = str(value or '').strip()
        if not text:
            return None, ''
        text = re.sub(r'\s+', ' ', text)
        for date_format in DATE_FORMATS:
            try:
                parsed = datetime.strptime(text, date_format).date()
                return parsed, 'parsed as DD/MM/YYYY' if date_format.startswith('%d') else ''
            except ValueError:
                continue
        return None, f'unparseable date: {text}'

    def _correct_excel_date(self, raw_date, sheet_month, default_year):
        if raw_date.month != sheet_month and raw_date.day == sheet_month and 1 <= raw_date.month <= 31:
            try:
                return date(raw_date.year or default_year, sheet_month, raw_date.month)
            except ValueError:
                return raw_date
        return raw_date

    def _balance_row(self, loan, repayments, rate):
        repayment_rows = [
            {
                'paid_on': datetime.strptime(repayment['paid_on'], '%Y-%m-%d').date(),
                'amount': self._parse_amount(repayment['amount']),
                'workbook_due_after_payment': self._parse_amount(repayment['workbook_due_after_payment']),
            }
            for repayment in repayments
        ]
        issue_date = datetime.strptime(loan['issue_date'], '%Y-%m-%d').date()
        comparison_date = max([issue_date] + [row['paid_on'] for row in repayment_rows])
        system_balance = self._simulate_balance(
            principal=self._parse_amount(loan['principal']),
            rate=rate,
            anchor=issue_date,
            repayments=repayment_rows,
            as_of_date=comparison_date,
        )
        workbook_due = Decimal('0')
        if repayment_rows:
            workbook_due = repayment_rows[-1]['workbook_due_after_payment']
        difference = system_balance - workbook_due
        return {
            'loan_reference': loan['loan_reference'],
            'account_label': loan['account_label'],
            'issue_date': loan['issue_date'],
            'principal': loan['principal'],
            'repayment_total': self._format_amount(
                sum((row['amount'] for row in repayment_rows), Decimal('0'))
            ),
            'comparison_date': comparison_date.isoformat(),
            'system_balance': self._format_amount(system_balance),
            'workbook_due_after_last_payment': self._format_amount(workbook_due),
            'difference': self._format_amount(difference),
            'workbook_final_due': loan['workbook_final_due'],
        }

    def _simulate_balance(self, principal, rate, anchor, repayments, as_of_date):
        repayments = sorted(
            [row for row in repayments if row['paid_on'] <= as_of_date],
            key=lambda row: row['paid_on'],
        )
        monthly_rate = rate / Decimal('100.00')
        months_elapsed = self._elapsed_full_months(anchor, as_of_date)
        balance = principal
        repayment_index = 0

        for month_number in range(1, months_elapsed + 2):
            if balance > 0 and monthly_rate > 0:
                balance += balance * monthly_rate
            month_end = self._add_months(anchor, month_number) if month_number <= months_elapsed else as_of_date
            while repayment_index < len(repayments) and repayments[repayment_index]['paid_on'] <= month_end:
                balance -= repayments[repayment_index]['amount']
                if balance < 0:
                    balance = Decimal('0.00')
                repayment_index += 1
        return balance

    def _last_friday(self, year, month):
        current = date(year, month, monthrange(year, month)[1])
        while current.weekday() != 4:
            current = current.replace(day=current.day - 1)
        return current

    def _add_months(self, base_date, months):
        year = base_date.year + (base_date.month - 1 + months) // 12
        month = (base_date.month - 1 + months) % 12 + 1
        day = min(base_date.day, 28)
        return base_date.replace(year=year, month=month, day=day)

    def _elapsed_full_months(self, start_date, end_date):
        months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
        if end_date.day < start_date.day:
            months -= 1
        return max(months, 0)

    def _due_after_paid_column(self, worksheet, row_number, paid_column, due_columns):
        due_column = next((column for column in due_columns if column > paid_column), None)
        if due_column is None:
            return Decimal('0')
        return self._parse_amount(worksheet.cell(row_number, due_column).value)

    def _last_nonzero_due(self, worksheet, row_number, due_columns):
        latest = Decimal('0')
        for column in due_columns:
            amount = self._parse_amount(worksheet.cell(row_number, column).value)
            if amount != 0:
                latest = amount
        return latest

    def _loan_reference(self, sheet, row_number, account_label, issue_date, principal):
        raw_key = f'{sheet}|{row_number}|{account_label}|{issue_date.isoformat()}|{self._format_amount(principal)}'
        digest = hashlib.sha1(raw_key.encode('utf-8')).hexdigest()[:8]
        return f'LOAN26-{slugify(sheet)}-{row_number}-{slugify(account_label)[:40]}-{digest}'

    def _repayment_reference(self, sheet, row_number, column, account_label, amount):
        raw_key = f'{sheet}|{row_number}|{column}|{account_label}|{self._format_amount(amount)}'
        digest = hashlib.sha1(raw_key.encode('utf-8')).hexdigest()[:8]
        return f'LRP26-{slugify(sheet)}-{row_number}-{self._column_letter(column)}-{slugify(account_label)[:34]}-{digest}'

    def _column_letter(self, column):
        result = ''
        while column:
            column, remainder = divmod(column - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _normalize_label(self, value):
        return re.sub(r'\s+', ' ', str(value or '')).strip()

    def _parse_amount(self, value):
        if value in (None, ''):
            return Decimal('0')
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value).replace(',', '').strip())
        except (InvalidOperation, AttributeError):
            return Decimal('0')

    def _format_amount(self, value):
        amount = self._parse_amount(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amount == amount.to_integral_value():
            return str(amount.to_integral_value())
        return format(amount.normalize(), 'f')

    def _write_csv(self, path, fieldnames, rows):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open('w', newline='', encoding='utf-8') as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
