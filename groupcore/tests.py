import csv
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.exceptions import PermissionDenied, ValidationError
from django.contrib.messages import get_messages
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone

from deposits.models import (
    DepositFineAllocation,
    DepositSubmission,
    DepositWelfareAllocation,
)
from deposits.rules import saving_year_weeks
from fines.models import Fine
from groupcore.models import (
    AccountSettlement,
    FinancialRecordRevision,
    FinancialYearClose,
    GroupSettings,
    MemberProfile,
    SavingsAccount,
    SettlementLoanAllocation,
)
from groupcore.week_cycle import current_saving_week, first_friday_of_year
from groupcore.savings_calendar import (
    apply_year_end_fine_relief,
    build_weekly_calendar,
    clear_on_time_missed_fine,
    ensure_overdue_fines,
    second_friday_of_december,
)
from groupcore.year_close import (
    ensure_automatic_year_lock,
    finalize_financial_year,
    lock_financial_year,
    record_payout,
)
from loans.models import LoanGuarantorApproval, LoanRepayment, LoanRequest
from Assets_Expenditures.models import Asset


class SavingWeekCycleTests(SimpleTestCase):
    def test_first_configured_year_uses_group_week_one_start(self):
        saving_week = current_saving_week(
            week_one_start=date(2026, 1, 2),
            today=date(2026, 1, 2),
        )

        self.assertEqual(saving_week.week_start, date(2026, 1, 2))
        self.assertEqual(saving_week.week_number, 1)
        self.assertEqual(saving_week.saving_year, 2026)

    def test_week_number_resets_for_next_saving_year(self):
        saving_week = current_saving_week(
            week_one_start=date(2026, 1, 2),
            today=date(2027, 7, 3),
        )

        self.assertEqual(saving_week.cycle_start, date(2027, 1, 1))
        self.assertEqual(saving_week.week_start, date(2027, 7, 2))
        self.assertEqual(saving_week.week_number, 27)
        self.assertEqual(saving_week.saving_year, 2027)

    def test_saving_year_ends_on_second_friday_of_december(self):
        _active, weeks = saving_year_weeks(
            week_one_start=date(2026, 1, 2),
            today=date(2026, 7, 26),
        )

        self.assertEqual(weeks[-1], date(2026, 12, 11))
        self.assertNotIn(date(2026, 12, 18), weeks)
        self.assertNotIn(date(2026, 12, 25), weeks)

    def test_active_week_stays_on_closing_friday_after_year_is_closed(self):
        saving_week = current_saving_week(
            week_one_start=date(2026, 1, 2),
            today=date(2026, 12, 25),
        )

        self.assertEqual(saving_week.week_start, date(2026, 12, 11))


class FinancialRecordAuditTests(TestCase):
    def setUp(self):
        self.treasurer = MemberProfile.objects.create_user(
            username='audit-treasurer', password='pass12345', role='TREASURER'
        )
        self.member = MemberProfile.objects.create_user(
            username='audit-member', password='pass12345', role='MEMBER'
        )
        self.asset = Asset.objects.create(
            name='Old asset', value=Decimal('10000'),
            date_acquired=date(2026, 1, 1), source='SAVINGS',
        )

    def test_treasurer_edit_creates_revision_and_member_cannot_inspect_unowned_record(self):
        self.client.login(username=self.treasurer.username, password='pass12345')
        response = self.client.post(
            reverse('financial_record_edit', args=['asset', self.asset.id]),
            {
                'name': 'Corrected asset', 'value': '12000',
                'date_acquired': '2026-01-02', 'source': 'SAVINGS',
                'remarks': 'Corrected', 'edit_reason': 'Correcting receipt details',
            },
        )
        self.assertRedirects(
            response,
            reverse('financial_record_history', args=['asset', self.asset.id]),
        )
        revision = FinancialRecordRevision.objects.get(
            record_type='asset', object_id=self.asset.id
        )
        self.assertEqual(revision.edited_by, self.treasurer)
        self.assertEqual(revision.before_data['name'], 'Old asset')
        self.assertEqual(revision.after_data['name'], 'Corrected asset')
        history = self.client.get(
            reverse('financial_record_history', args=['asset', self.asset.id])
        )
        self.assertContains(history, 'Name')
        self.assertContains(history, 'Old asset')
        self.assertContains(history, 'Corrected asset')
        self.assertContains(history, 'Before')
        self.assertContains(history, 'After')
        self.assertNotContains(history, "{'name':")

        self.client.logout()
        self.client.login(username=self.member.username, password='pass12345')
        denied = self.client.get(
            reverse('financial_record_history', args=['asset', self.asset.id])
        )
        self.assertRedirects(denied, reverse('member_dashboard'))

    def test_approved_deposit_edit_reconciles_welfare_fine_and_loan(self):
        member = MemberProfile.objects.create_user(
            username='audit-depositor', password='pass12345', role='MEMBER'
        )
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        loan = LoanRequest.objects.create(
            member=member, account=account, principal=Decimal('100000'),
            duration_months=6, status='APPROVED', approved_on=timezone.now(),
        )
        fine = Fine.objects.create(
            member=member, account=account, fine_type='MISSED_WEEKLY_SAVING',
            reference_week=date(2026, 1, 2), reason='Missed', amount=Decimal('2000'),
            amount_paid=Decimal('2000'), is_paid=True,
        )
        deposit = DepositSubmission.objects.create(
            member=member, account=account, submitted_by=self.treasurer,
            reviewed_by=self.treasurer, payment_week=date(2026, 1, 2),
            welfare_amount=Decimal('1000'), fine_amount=Decimal('2000'),
            loan_repayment_amount=Decimal('10000'), loan_repayment_loan=loan,
            payment_date=date(2026, 7, 1), payment_time=time(10, 0), status='APPROVED',
        )
        DepositWelfareAllocation.objects.create(
            deposit=deposit, account=account, welfare_week=date(2026, 1, 2),
        )
        DepositFineAllocation.objects.create(
            deposit=deposit, fine=fine, amount=Decimal('2000'),
        )
        LoanRepayment.objects.create(
            loan=loan, source_deposit=deposit, amount=Decimal('10000'),
            paid_on=deposit.payment_date, recorded_by=self.treasurer,
        )

        self.client.login(username=self.treasurer.username, password='pass12345')
        response = self.client.post(
            reverse('financial_record_edit', args=['deposit', deposit.id]),
            {
                'member': member.id, 'account': account.id,
                'payment_week': '2026-01-02', 'payment_date': '2026-07-01',
                'payment_time': '10:00', 'saving_amount': '0',
                'annual_subscription_amount': '0', 'membership_amount': '0',
                'shares_amount': '0', 'loan_repayment_loan': loan.id,
                'loan_repayment_amount': '5000', 'remarks': 'Corrected',
                'welfare_weeks': ['2026-01-09'], 'selected_fines': [fine.id],
                'edit_reason': 'Correcting allocated weeks and repayment',
            },
        )
        self.assertEqual(response.status_code, 302)
        deposit.refresh_from_db()
        fine.refresh_from_db()
        self.assertEqual(deposit.welfare_amount, Decimal('1000'))
        self.assertEqual(
            deposit.welfare_allocations.get().welfare_week, date(2026, 1, 9)
        )
        self.assertEqual(fine.amount_paid, Decimal('2000'))
        self.assertTrue(fine.is_paid)
        self.assertEqual(
            deposit.generated_loan_repayments.get().amount, Decimal('5000')
        )
        revision = FinancialRecordRevision.objects.get(
            record_type='deposit', object_id=deposit.id
        )
        self.assertEqual(revision.before_data['welfare_weeks'][0]['week'], '2026-01-02')
        self.assertEqual(revision.after_data['welfare_weeks'][0]['week'], '2026-01-09')

    def test_deposit_edit_welfare_choices_stop_at_annual_closing(self):
        member = MemberProfile.objects.create_user(
            username='closing-edit-member', password='pass12345', role='MEMBER'
        )
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        deposit = DepositSubmission.objects.create(
            member=member, account=account, submitted_by=self.treasurer,
            payment_week=date(2026, 7, 3), saving_amount=Decimal('10000'),
            payment_date=date(2026, 7, 3), payment_time=time(10, 0),
            status='APPROVED',
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(
            reverse('financial_record_edit', args=['deposit', deposit.id])
        )

        choices = dict(response.context['form'].fields['welfare_weeks'].choices)
        self.assertIn('2026-12-11', choices)
        self.assertNotIn('2026-12-18', choices)
        self.assertNotIn('2026-12-25', choices)

    def test_deposit_edit_hides_welfare_weeks_used_by_other_deposits(self):
        member = MemberProfile.objects.create_user(
            username='occupied-welfare-member', password='pass12345', role='MEMBER'
        )
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        deposit = DepositSubmission.objects.create(
            member=member, account=account, submitted_by=self.treasurer,
            payment_week=date(2026, 7, 3), welfare_amount=Decimal('1000'),
            payment_date=date(2026, 7, 3), payment_time=time(10, 0),
            status='APPROVED',
        )
        DepositWelfareAllocation.objects.create(
            deposit=deposit, account=account, welfare_week=date(2026, 7, 3)
        )
        other = DepositSubmission.objects.create(
            member=member, account=account, submitted_by=self.treasurer,
            payment_week=date(2026, 7, 10), welfare_amount=Decimal('1000'),
            payment_date=date(2026, 7, 10), payment_time=time(10, 0),
            status='PENDING',
        )
        DepositWelfareAllocation.objects.create(
            deposit=other, account=account, welfare_week=date(2026, 7, 10)
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(
            reverse('financial_record_edit', args=['deposit', deposit.id])
        )

        choices = dict(response.context['form'].fields['welfare_weeks'].choices)
        self.assertIn('2026-07-03', choices)
        self.assertIn('currently on this deposit', choices['2026-07-03'])
        self.assertNotIn('2026-07-10', choices)
        self.assertContains(response, 'Only available weeks are shown')
        self.assertContains(response, 'correction-choice-grid')


class AutomaticWeeklyFineTests(TestCase):
    def test_ten_thousand_is_paid_but_subthreshold_week_is_fined(self):
        paid_member = MemberProfile.objects.create_user(username='threshold-paid', password='pass12345')
        unpaid_member = MemberProfile.objects.create_user(username='threshold-unpaid', password='pass12345')
        paid_account = SavingsAccount.objects.create(owner=paid_member, label='A')
        unpaid_account = SavingsAccount.objects.create(owner=unpaid_member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        for member, account, amount in (
            (paid_member, paid_account, Decimal('10000')),
            (unpaid_member, unpaid_account, Decimal('9999')),
        ):
            DepositSubmission.objects.create(
                member=member, account=account, submitted_by=member,
                payment_week=date(2026, 1, 2), saving_amount=amount,
                payment_date=date(2026, 1, 2), payment_time=time(9, 0), status='APPROVED',
            )
        after_deadline = timezone.make_aware(datetime(2026, 1, 5, 0, 1))

        ensure_overdue_fines(now=after_deadline)

        self.assertFalse(Fine.objects.filter(member=paid_member).exists())
        fine = Fine.objects.get(member=unpaid_member, reference_week=date(2026, 1, 2))
        self.assertEqual(fine.amount, Decimal('2000'))

    def test_year_end_relief_only_applies_when_savings_remain_unpaid(self):
        member = MemberProfile.objects.create_user(username='relief-member', password='pass12345')
        paid_member = MemberProfile.objects.create_user(username='late-paid-member', password='pass12345')
        account = SavingsAccount.objects.create(owner=member, label='A')
        paid_account = SavingsAccount.objects.create(owner=paid_member, label='A')
        week = date(2026, 1, 2)
        closing = second_friday_of_december(2026)
        unpaid_fine = Fine.objects.create(
            member=member, account=account, fine_type='MISSED_WEEKLY_SAVING',
            reference_week=week, reason='Missed', amount=Decimal('2000'),
        )
        paid_fine = Fine.objects.create(
            member=paid_member, account=paid_account, fine_type='MISSED_WEEKLY_SAVING',
            reference_week=week, reason='Late', amount=Decimal('2000'),
        )
        DepositSubmission.objects.create(
            member=paid_member, account=paid_account, submitted_by=paid_member,
            payment_week=week, saving_amount=Decimal('10000'),
            payment_date=date(2026, 2, 1), payment_time=time(9, 0), status='APPROVED',
        )

        self.assertEqual(apply_year_end_fine_relief(2026, closing), 1)
        self.assertEqual(apply_year_end_fine_relief(2026, closing), 0)
        unpaid_fine.refresh_from_db()
        paid_fine.refresh_from_db()
        self.assertEqual(unpaid_fine.amount, Decimal('1000'))
        self.assertEqual(unpaid_fine.original_amount, Decimal('2000'))
        self.assertEqual(unpaid_fine.relief_applied_on, closing)
        self.assertEqual(paid_fine.amount, Decimal('2000'))

    def test_calendar_marks_valid_late_allocation_paid_late(self):
        member = MemberProfile.objects.create_user(username='calendar-late', password='pass12345')
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), saving_amount=Decimal('30000'),
            payment_date=date(2026, 1, 6), payment_time=time(9, 0), status='APPROVED',
        )

        calendar = build_weekly_calendar(member, account, today=date(2026, 7, 14))

        self.assertEqual(calendar['weeks'][0]['status'], 'paid_late')
        self.assertEqual(calendar['weeks'][0]['required'], Decimal('10000'))

    def test_pending_on_time_payment_does_not_create_fine(self):
        member = MemberProfile.objects.create_user(username='pending-on-time', password='pass12345')
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), saving_amount=Decimal('10000'),
            payment_date=date(2026, 1, 3), payment_time=time(10, 0), status='PENDING',
        )

        self.assertEqual(
            ensure_overdue_fines(
                member, account,
                timezone.make_aware(datetime(2026, 1, 6, 0, 1)),
            ),
            0,
        )
        self.assertFalse(Fine.objects.filter(member=member).exists())

    def test_on_time_approval_removes_previously_created_fine(self):
        member = MemberProfile.objects.create_user(username='approval-on-time', password='pass12345')
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        fine = Fine.objects.create(
            member=member, account=account, fine_type='MISSED_WEEKLY_SAVING',
            reference_week=date(2026, 1, 2), reason='Generated before approval',
            amount=Decimal('1000'), issued_by=None,
        )
        DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), saving_amount=Decimal('10000'),
            payment_date=date(2026, 1, 3), payment_time=time(10, 0), status='APPROVED',
        )

        self.assertEqual(clear_on_time_missed_fine(member, account, date(2026, 1, 2)), 1)
        self.assertFalse(Fine.objects.filter(pk=fine.pk).exists())

    def test_on_time_approval_voids_referenced_fine_instead_of_raising_protected_error(self):
        member = MemberProfile.objects.create_user(
            username='protected-fine-member', password='pass12345'
        )
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        fine = Fine.objects.create(
            member=member, account=account, fine_type='MISSED_WEEKLY_SAVING',
            reference_week=date(2026, 1, 2), reason='Generated before approval',
            amount=Decimal('2000'),
        )
        fine_deposit = DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), fine_amount=Decimal('2000'),
            payment_date=date(2026, 1, 4), payment_time=time(9, 0), status='PENDING',
        )
        DepositFineAllocation.objects.create(
            deposit=fine_deposit, fine=fine, amount=Decimal('2000'),
        )
        DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), saving_amount=Decimal('10000'),
            payment_date=date(2026, 1, 3), payment_time=time(10, 0), status='APPROVED',
        )

        self.assertEqual(
            clear_on_time_missed_fine(member, account, date(2026, 1, 2)),
            1,
        )
        fine.refresh_from_db()
        self.assertTrue(fine.is_voided)
        self.assertEqual(fine.outstanding_amount, Decimal('0.00'))
        self.assertTrue(fine.deposit_allocations.filter(deposit=fine_deposit).exists())

    def test_generation_is_idempotent_and_late_savings_do_not_remove_fine(self):
        member = MemberProfile.objects.create_user(username='calendar-member', password='pass12345')
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        after_deadline = timezone.make_aware(datetime(2026, 1, 5, 0, 1))

        self.assertEqual(ensure_overdue_fines(member, account, after_deadline), 1)
        self.assertEqual(ensure_overdue_fines(member, account, after_deadline), 0)
        fine = Fine.objects.get(member=member, account=account, reference_week=date(2026, 1, 2))
        DepositSubmission.objects.create(
            member=member, account=account, submitted_by=member,
            payment_week=date(2026, 1, 2), saving_amount=Decimal('20000'),
            payment_date=date(2026, 1, 6), payment_time=time(10, 0), status='APPROVED',
        )
        self.assertTrue(Fine.objects.filter(pk=fine.pk, is_paid=False).exists())

    def test_fine_generation_stops_at_annual_closing_week(self):
        member = MemberProfile.objects.create_user(
            username='closing-fine-member', password='pass12345'
        )
        account = SavingsAccount.objects.create(owner=member, label='A')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))

        ensure_overdue_fines(
            member,
            account,
            timezone.make_aware(datetime(2026, 12, 28, 9, 0)),
        )

        fine_weeks = Fine.objects.filter(member=member).values_list(
            'reference_week', flat=True
        )
        self.assertIn(date(2026, 12, 11), fine_weeks)
        self.assertNotIn(date(2026, 12, 18), fine_weeks)
        self.assertNotIn(date(2026, 12, 25), fine_weeks)

    def test_new_year_waits_for_the_first_friday_saving_week(self):
        saving_week = current_saving_week(
            week_one_start=date(2026, 1, 2),
            today=date(2028, 1, 3),
        )

        self.assertEqual(saving_week.cycle_start, date(2028, 1, 7))
        self.assertEqual(saving_week.week_start, date(2028, 1, 7))
        self.assertEqual(saving_week.week_number, 1)


class RootUrlTests(SimpleTestCase):
    def test_root_redirects_to_login(self):
        response = self.client.get('/')

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], '/login/')


class TreasurerMemberDashboardPreviewTests(TestCase):
    def setUp(self):
        self.treasurer = MemberProfile.objects.create_user(
            username='preview-treasurer', password='pass12345', role='TREASURER'
        )
        self.member = MemberProfile.objects.create_user(
            username='preview-member', first_name='Amina', last_name='Nansubuga',
            email='amina@example.com', phone_number='0700000000', password='pass12345',
        )
        self.account_a = SavingsAccount.objects.create(owner=self.member, label='A')
        self.account_b = SavingsAccount.objects.create(owner=self.member, label='B')
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        for account, amount in ((self.account_a, Decimal('20000')), (self.account_b, Decimal('40000'))):
            DepositSubmission.objects.create(
                member=self.member,
                account=account,
                submitted_by=self.member,
                payment_week=date(2026, 1, 2),
                saving_amount=amount,
                payment_date=date(2026, 1, 2),
                payment_time=time(9, 0),
                status='APPROVED',
            )

    def test_treasurer_can_search_for_a_member(self):
        MemberProfile.objects.create_superuser(
            username='hidden-admin', email='admin@example.com', password='pass12345'
        )
        self.client.login(username='preview-treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_member_preview_index'), {'q': 'Amina'})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'groupcore/treasurer_member_preview_select.html')
        self.assertContains(response, 'Amina Nansubuga')
        self.assertNotContains(response, 'hidden-admin')

    def test_selected_member_uses_the_real_member_dashboard_in_read_only_mode(self):
        self.client.login(username='preview-treasurer', password='pass12345')

        response = self.client.get(
            reverse('treasurer_member_preview_index'),
            {'member': self.member.id, 'account': self.account_a.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'groupcore/member_dashboard.html')
        self.assertTrue(response.context['is_treasurer_preview'])
        self.assertEqual(response.context['preview_member'], self.member)
        self.assertEqual(response.context['active_account'], self.account_a)
        self.assertEqual(response.context['user_savings_total'], Decimal('20000'))
        self.assertContains(response, 'Viewing Amina Nansubuga')
        self.assertNotContains(response, 'New Deposit')

    def test_preview_can_switch_between_a_members_accounts(self):
        self.client.login(username='preview-treasurer', password='pass12345')

        response = self.client.get(
            reverse('treasurer_member_preview_index'),
            {'member': self.member.id, 'account': self.account_b.id},
        )

        self.assertEqual(response.context['active_account'], self.account_b)
        self.assertEqual(response.context['user_savings_total'], Decimal('40000'))

    def test_existing_member_preview_link_uses_the_same_dashboard(self):
        self.client.login(username='preview-treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_member_preview', args=[self.member.id]))

        self.assertTemplateUsed(response, 'groupcore/member_dashboard.html')
        self.assertEqual(response.context['preview_member'], self.member)

    def test_non_treasurer_cannot_preview_other_members(self):
        outsider = MemberProfile.objects.create_user(
            username='preview-outsider', password='pass12345', role='MEMBER'
        )
        self.client.login(username=outsider.username, password='pass12345')

        response = self.client.get(reverse('treasurer_member_preview_index'))

        self.assertRedirects(response, reverse('member_dashboard'))


class TreasurerPersonalMemberPanelTests(TestCase):
    def setUp(self):
        self.treasurer = MemberProfile.objects.create_user(
            username='personal-panel-treasurer',
            password='pass12345',
            role='TREASURER',
        )
        SavingsAccount.objects.create(owner=self.treasurer, label='Personal')

    def test_treasurer_sidebar_contains_personal_member_tools(self):
        self.client.login(
            username=self.treasurer.username, password='pass12345'
        )

        response = self.client.get(reverse('treasurer_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="treasurer-member-panel"')
        self.assertContains(response, 'Personal Member Panel')
        self.assertContains(response, reverse('member_dashboard'))
        self.assertContains(response, reverse('submit_deposit'))
        self.assertContains(response, reverse('my_contributions'))
        self.assertContains(response, reverse('request_loan'))
        self.assertContains(response, reverse('my_loans'))
        self.assertContains(response, reverse('my_fines'))


class GroupSettingsSetupTests(TestCase):
    def setUp(self):
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.chairman = MemberProfile.objects.create_user(
            username='chairman',
            password='pass12345',
            role='CHAIRMAN',
        )
        self.member = MemberProfile.objects.create_user(
            username='member',
            password='pass12345',
            role='MEMBER',
        )

    def test_group_settings_defaults_to_first_friday_of_current_year(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('group_settings'))

        expected_start = first_friday_of_year(timezone.localdate().year)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['settings_exists'])
        self.assertEqual(response.context['form'].initial['week_one_start'], expected_start)
        self.assertEqual(response.context['saving_week'].cycle_start, expected_start)

    def test_treasurer_can_create_group_settings(self):
        self.client.login(username='treasurer', password='pass12345')
        week_one_start = first_friday_of_year(timezone.localdate().year)

        response = self.client.post(
            reverse('group_settings'),
            {'week_one_start': week_one_start.isoformat()},
        )

        self.assertRedirects(response, reverse('group_settings'))
        self.assertEqual(GroupSettings.objects.count(), 1)
        self.assertEqual(GroupSettings.get_active().week_one_start, week_one_start)

    def test_chairman_can_update_group_settings(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        self.client.login(username='chairman', password='pass12345')

        response = self.client.post(
            reverse('group_settings'),
            {'week_one_start': '2027-01-01'},
        )

        self.assertRedirects(response, reverse('group_settings'))
        self.assertEqual(GroupSettings.objects.count(), 1)
        self.assertEqual(GroupSettings.get_active().week_one_start, date(2027, 1, 1))

    def test_member_cannot_access_group_settings(self):
        self.client.login(username='member', password='pass12345')

        response = self.client.get(reverse('group_settings'))

        self.assertRedirects(response, reverse('member_dashboard'))

    def test_group_settings_save_reuses_existing_record(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        GroupSettings.objects.create(week_one_start=date(2027, 1, 1))

        self.assertEqual(GroupSettings.objects.count(), 1)
        self.assertEqual(GroupSettings.get_active().week_one_start, date(2027, 1, 1))

    def test_member_deposit_submission_has_friendly_message_when_cycle_missing(self):
        self.client.login(username='member', password='pass12345')

        response = self.client.get(reverse('submit_deposit'))

        self.assertRedirects(response, reverse('member_dashboard'))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn(
            "The saving cycle has not been opened yet. Please contact the Treasurer.",
            messages,
        )

    def test_chairman_deposit_submission_redirects_to_setup_when_cycle_missing(self):
        self.client.login(username='chairman', password='pass12345')

        response = self.client.get(reverse('submit_deposit'))

        self.assertRedirects(response, reverse('group_settings'))


class LeadershipAccountSelectionTests(TestCase):
    def setUp(self):
        GroupSettings.objects.create(week_one_start=first_friday_of_year(timezone.localdate().year))
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.secretary = MemberProfile.objects.create_user(
            username='secretary',
            password='pass12345',
            role='SECRETARY',
        )
        SavingsAccount.objects.create(owner=self.treasurer, label='T1')
        SavingsAccount.objects.create(owner=self.secretary, label='A1')
        SavingsAccount.objects.create(owner=self.secretary, label='A2')

    def test_secretary_management_dashboard_does_not_require_account_selection(self):
        self.client.login(username='secretary', password='pass12345')

        response = self.client.get(reverse('secretary_dashboard'))

        self.assertEqual(response.status_code, 200)

    def test_secretary_member_feature_requires_account_selection(self):
        self.client.login(username='secretary', password='pass12345')

        response = self.client.get(reverse('submit_deposit'))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response['Location'].startswith(reverse('select_savings_account')))
        self.assertIn(reverse('submit_deposit'), response['Location'])

    def test_treasurer_can_access_personal_deposit_submission(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('submit_deposit'))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'deposits/submit_deposit.html')

    def test_treasurer_mobile_deposit_shortcut_opens_deposit_management(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_dashboard'))

        self.assertContains(response, f'href="{reverse("manage_deposits")}"')
        self.assertContains(response, '<i class="bi bi-clipboard-check"></i><span>Deposits</span>', html=True)

    def test_account_selection_returns_to_requested_member_feature(self):
        self.client.login(username='secretary', password='pass12345')
        account = SavingsAccount.objects.get(owner=self.secretary, label='A2')

        response = self.client.post(
            reverse('select_savings_account'),
            {'account_id': account.id, 'next': reverse('my_contributions')},
        )

        self.assertRedirects(response, reverse('my_contributions'))


class MemberWeekProgressTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        self.week_one_start = first_friday_of_year(today.year)
        GroupSettings.objects.create(week_one_start=self.week_one_start)
        self.member = MemberProfile.objects.create_user(
            username='member',
            password='pass12345',
            role='MEMBER',
        )
        self.account = SavingsAccount.objects.create(owner=self.member, label='A1')

    def test_member_dashboard_reports_missing_due_weeks(self):
        self.client.login(username='member', password='pass12345')

        response = self.client.get(reverse('member_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['week_progress']['status'], 'Behind')
        self.assertGreater(response.context['week_progress']['missing_weeks_count'], 0)

    def test_dashboard_separates_paid_to_date_prepaid_and_total_covered(self):
        saving_week = current_saving_week(
            self.week_one_start,
            timezone.localdate(),
        )
        due_week = saving_week.cycle_start
        future_week = saving_week.week_start + timedelta(weeks=1)
        for week in (due_week, future_week):
            DepositSubmission.objects.create(
                member=self.member,
                account=self.account,
                submitted_by=self.member,
                payment_week=week,
                saving_amount=Decimal('10000.00'),
                payment_date=week,
                payment_time=time(9, 0),
                status='APPROVED',
            )
        self.client.login(username='member', password='pass12345')

        response = self.client.get(reverse('member_dashboard'))

        progress = response.context['week_progress']
        self.assertEqual(progress['paid_weeks_count'], 1)
        self.assertEqual(progress['future_weeks_count'], 1)
        self.assertEqual(progress['total_paid_weeks_count'], 2)
        self.assertEqual(response.context['weeks_paid'], 2)
        self.assertContains(response, 'Total Covered Weeks')
        self.assertContains(response, 'Paid to Date')
        self.assertContains(response, 'Prepaid')
        self.assertContains(response, 'Total Covered')
        self.assertNotContains(response, 'Approved Weeks')


class YearEndSettlementYearFilterTests(TestCase):
    def setUp(self):
        self.current_year = timezone.localdate().year
        self.previous_year = self.current_year - 1
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.member = MemberProfile.objects.create_user(
            username='member',
            password='pass12345',
            role='MEMBER',
        )
        self._deposit(
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.current_year, 1, 2),
            saving_amount=Decimal('1000.00'),
        )
        self._deposit(
            payment_week=date(self.current_year, 1, 5),
            payment_date=date(self.current_year, 1, 5),
            saving_amount=Decimal('9000.00'),
        )
        self.previous_loan = self._loan(
            principal=Decimal('1000.00'),
            approved_on=timezone.make_aware(datetime(self.previous_year, 6, 1, 9, 0)),
        )
        self._loan(
            principal=Decimal('5000.00'),
            approved_on=timezone.make_aware(datetime(self.current_year, 6, 1, 9, 0)),
        )

    def _deposit(self, payment_week, payment_date, saving_amount):
        return DepositSubmission.objects.create(
            member=self.member,
            submitted_by=self.treasurer,
            payment_week=payment_week,
            starting_week=payment_week,
            weeks_covered=1,
            saving_amount=saving_amount,
            proof='proofs/test.jpg',
            payment_date=payment_date,
            payment_time=time(9, 0),
            status='APPROVED',
        )

    def _loan(self, principal, approved_on):
        return LoanRequest.objects.create(
            member=self.member,
            principal=principal,
            monthly_interest_rate=Decimal('2.00'),
            duration_months=1,
            status=LoanRequest.STATUS_APPROVED,
            approved_on=approved_on,
        )

    def test_year_end_settlement_can_display_previous_year(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(
            reverse('year_end_settlement'),
            {'year': self.previous_year},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['target_year'], self.previous_year)
        self.assertEqual(
            response.context['state'].state,
            FinancialYearClose.STATE_LOCKED,
        )
        self.assertContains(response, 'Locked — Review')
        self.assertContains(response, 'No finalized settlement version exists')


class ControlledFinancialYearClosingTests(TestCase):
    year = 2026

    def setUp(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        self.treasurer = MemberProfile.objects.create_user(
            username='close-treasurer', password='pass12345', role='TREASURER'
        )
        self.member = MemberProfile.objects.create_user(
            username='close-member', password='pass12345', role='MEMBER'
        )
        self.account = SavingsAccount.objects.create(
            owner=self.member, label='Main'
        )
        self.cutoff = timezone.make_aware(datetime(2026, 12, 12, 9, 0))

    def _deposit(self, member=None, account=None, status='APPROVED', **amounts):
        return DepositSubmission.objects.create(
            member=member or self.member,
            account=account or self.account,
            submitted_by=member or self.member,
            payment_week=date(2026, 1, 2),
            payment_date=date(2026, 1, 2),
            payment_time=time(9, 0),
            status=status,
            **amounts,
        )

    def test_only_treasurer_can_lock_and_manual_lock_cannot_be_early(self):
        with self.assertRaises(PermissionDenied):
            lock_financial_year(self.year, self.member, cutoff_at=self.cutoff)
        with self.assertRaises(ValidationError):
            lock_financial_year(
                self.year,
                self.treasurer,
                cutoff_at=timezone.make_aware(datetime(2026, 12, 1, 9, 0)),
            )

        state = lock_financial_year(
            self.year, self.treasurer, cutoff_at=self.cutoff
        )

        self.assertEqual(state.state, FinancialYearClose.STATE_LOCKED)
        self.assertEqual(state.locked_by, self.treasurer)
        self.assertFalse(state.auto_locked)

    def test_pending_queue_blocks_finalization_until_resolved(self):
        pending = self._deposit(status='PENDING', saving_amount=Decimal('10000'))
        state = lock_financial_year(
            self.year, self.treasurer, cutoff_at=self.cutoff
        )

        with self.assertRaises(ValidationError):
            finalize_financial_year(self.year, self.treasurer)

        pending.status = 'REJECTED'
        pending.save(update_fields=['status'])
        version = finalize_financial_year(self.year, self.treasurer)
        state.refresh_from_db()
        self.assertEqual(state.state, FinancialYearClose.STATE_FINALIZED)
        self.assertEqual(version.version, 1)

    def test_pre_cutoff_pending_repayment_can_be_approved_after_lock(self):
        loan = LoanRequest.objects.create(
            member=self.member,
            account=self.account,
            principal=Decimal('100000'),
            monthly_interest_rate=Decimal('2'),
            duration_months=6,
            status=LoanRequest.STATUS_APPROVED,
            approved_on=timezone.make_aware(datetime(2026, 6, 1, 9, 0)),
        )
        deposit = self._deposit(
            status='PENDING',
            loan_repayment_amount=Decimal('10000'),
            loan_repayment_loan=loan,
        )
        state = lock_financial_year(
            self.year, self.treasurer, cutoff_at=self.cutoff
        )
        frozen_before = state.loan_snapshots.get(loan=loan).frozen_balance
        self.client.login(
            username=self.treasurer.username, password='pass12345'
        )

        response = self.client.post(reverse('approve_deposit', args=[deposit.id]))

        self.assertEqual(response.status_code, 302)
        deposit.refresh_from_db()
        snapshot = state.loan_snapshots.get(loan=loan)
        repayment = deposit.generated_loan_repayments.get()
        self.assertEqual(deposit.status, 'APPROVED')
        self.assertEqual(
            snapshot.frozen_balance,
            frozen_before - repayment.amount,
        )
        self.assertEqual(
            snapshot.collected_interest,
            repayment.interest_component,
        )

    def test_automatic_lock_uses_january_first_friday_cutoff(self):
        state = ensure_automatic_year_lock(date(2027, 1, 1))

        self.assertEqual(state.year, 2026)
        self.assertEqual(state.state, FinancialYearClose.STATE_LOCKED)
        self.assertTrue(state.auto_locked)
        self.assertEqual(
            timezone.localtime(state.cutoff_at),
            timezone.make_aware(datetime(2027, 1, 1, 0, 0)),
        )

    def test_settlement_deducts_only_outstanding_obligations(self):
        _active, weeks = saving_year_weeks(date(2026, 1, 2), date(2026, 12, 11))
        deposit = self._deposit(
            saving_amount=Decimal('100000'),
            welfare_amount=Decimal('1000'),
            annual_subscription_amount=Decimal('10000'),
        )
        DepositWelfareAllocation.objects.create(
            deposit=deposit,
            account=self.account,
            welfare_week=weeks[0],
            amount=Decimal('1000'),
        )
        Fine.objects.create(
            member=self.member,
            account=self.account,
            fine_type='MISSED_WEEKLY_SAVING',
            reference_week=weeks[1],
            reason='Unpaid after relief',
            amount=Decimal('1000'),
        )
        paid_fine = Fine.objects.create(
            member=self.member,
            account=self.account,
            fine_type='OTHER',
            reason='Already paid',
            amount=Decimal('3000'),
            amount_paid=Decimal('3000'),
            is_paid=True,
        )
        self.assertEqual(paid_fine.outstanding_amount, Decimal('0'))
        lock_financial_year(self.year, self.treasurer, cutoff_at=self.cutoff)

        version = finalize_financial_year(self.year, self.treasurer)
        row = version.account_settlements.get(account=self.account)

        self.assertEqual(row.savings_total, Decimal('100000'))
        self.assertEqual(
            row.welfare_due,
            Decimal('1000') * (len(weeks) - 1),
        )
        self.assertEqual(row.fine_due, Decimal('1000'))
        self.assertEqual(row.annual_due, Decimal('0'))
        self.assertEqual(
            row.net_payout,
            Decimal('100000') - row.welfare_due - row.fine_due,
        )

    def test_loan_offsets_borrower_then_guarantors_and_records_loss(self):
        guarantors = [
            MemberProfile.objects.create_user(
                username=f'close-guarantor-{number}',
                password='pass12345',
            )
            for number in (1, 2)
        ]
        guarantor_accounts = [
            SavingsAccount.objects.create(owner=user, label='Main')
            for user in guarantors
        ]
        self._deposit(saving_amount=Decimal('100000'))
        for user, account in zip(guarantors, guarantor_accounts):
            self._deposit(
                member=user, account=account, saving_amount=Decimal('100000')
            )
        loan = LoanRequest.objects.create(
            member=self.member,
            account=self.account,
            principal=Decimal('300000'),
            monthly_interest_rate=Decimal('2'),
            duration_months=12,
            status=LoanRequest.STATUS_APPROVED,
            approved_on=timezone.make_aware(datetime(2026, 12, 1, 9, 0)),
        )
        for guarantor in guarantors:
            LoanGuarantorApproval.objects.create(
                loan=loan,
                guarantor=guarantor,
                status=LoanGuarantorApproval.STATUS_APPROVED,
            )
        lock_financial_year(self.year, self.treasurer, cutoff_at=self.cutoff)

        version = finalize_financial_year(self.year, self.treasurer)
        loan.refresh_from_db()
        allocations = version.loan_allocations.filter(
            loan_snapshot__loan=loan
        )

        self.assertEqual(loan.status, LoanRequest.STATUS_SETTLED)
        self.assertTrue(
            allocations.filter(
                allocation_type=SettlementLoanAllocation.TYPE_BORROWER
            ).exists()
        )
        self.assertEqual(
            allocations.filter(
                allocation_type=SettlementLoanAllocation.TYPE_GUARANTOR
            ).count(),
            2,
        )
        self.assertGreater(version.total_group_loss, Decimal('0'))
        self.assertEqual(loan.settlement_loss, version.total_group_loss)

    def test_finalization_is_idempotent_until_an_audited_correction(self):
        self._deposit(saving_amount=Decimal('100000'))
        state = lock_financial_year(
            self.year, self.treasurer, cutoff_at=self.cutoff
        )
        first = finalize_financial_year(self.year, self.treasurer)
        repeated = finalize_financial_year(self.year, self.treasurer)
        self.assertEqual(repeated.pk, first.pk)
        self.assertEqual(state.versions.count(), 1)

        state.needs_regeneration = True
        state.save(update_fields=['needs_regeneration'])
        corrected = finalize_financial_year(self.year, self.treasurer)
        self.assertEqual(corrected.version, 2)
        self.assertEqual(state.versions.count(), 2)

    def test_payout_permissions_and_partial_status(self):
        self._deposit(saving_amount=Decimal('100000'))
        lock_financial_year(self.year, self.treasurer, cutoff_at=self.cutoff)
        row = finalize_financial_year(
            self.year, self.treasurer
        ).account_settlements.get(account=self.account)

        with self.assertRaises(PermissionDenied):
            record_payout(
                row, Decimal('1'), date(2026, 12, 20), '', '', self.member
            )
        payment = record_payout(
            row,
            row.net_payout / 2,
            date(2026, 12, 20),
            'BANK-001',
            '',
            self.treasurer,
        )
        row.refresh_from_db()
        self.assertEqual(row.payout_status, AccountSettlement.STATUS_PARTIAL)
        self.assertEqual(payment.recorded_by, self.treasurer)


class HistoricalDataImportCommandTests(TestCase):
    def setUp(self):
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )

    def _write_import_files(self, directory, expected_total='21000', payment_date='2026-06-26'):
        members_path = Path(directory) / 'members.csv'
        transactions_path = Path(directory) / 'transactions.csv'
        members_path.write_text(
            '\n'.join([
                'username,first_name,last_name,email,phone_number,role,account_labels,is_active',
                'jane_member,Jane,Member,jane@example.com,+256700000001,MEMBER,A1,true',
            ]),
            encoding='utf-8',
        )
        transactions_path.write_text(
            '\n'.join([
                (
                    'transaction_reference,username,account_label,payment_week,payment_date,payment_time,'
                    'saving_amount,welfare_amount,annual_subscription_amount,fine_amount,shares_amount,'
                    'loan_repayment_amount,expected_total,status,remarks,proof_reference'
                ),
                (
                    f'OLD-001,jane_member,A1,2026-06-26,{payment_date},09:30,'
                    f'20000,1000,0,0,0,0,{expected_total},APPROVED,Historical import,proofs/old-001.jpg'
                ),
            ]),
            encoding='utf-8',
        )
        return members_path, transactions_path

    def test_dry_run_validates_without_writing_data(self):
        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(directory)
            report_path = Path(directory) / 'report.csv'

            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(report_path),
            )

            self.assertFalse(MemberProfile.objects.filter(username='jane_member').exists())
            self.assertIn('VALID_NEW_MEMBER', report_path.read_text(encoding='utf-8'))
            self.assertIn('VALID_NEW_TRANSACTION', report_path.read_text(encoding='utf-8'))

    def test_commit_imports_members_accounts_and_historical_transactions(self):
        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(directory)
            report_path = Path(directory) / 'report.csv'

            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(report_path),
                commit=True,
            )

            member = MemberProfile.objects.get(username='jane_member')
            account = SavingsAccount.objects.get(owner=member, label='A1')
            deposit = DepositSubmission.objects.get(import_reference='OLD-001')
            self.assertEqual(deposit.member, member)
            self.assertEqual(deposit.account, account)
            self.assertEqual(deposit.payment_week, date(2026, 6, 26))
            self.assertEqual(deposit.payment_date, date(2026, 6, 26))
            self.assertEqual(deposit.payment_time, time(9, 30))
            self.assertEqual(deposit.amount, Decimal('21000.00'))
            self.assertEqual(deposit.status, 'APPROVED')
            self.assertIn('CREATED_TRANSACTION', report_path.read_text(encoding='utf-8'))

    def test_commit_import_preserves_existing_weekly_fine_independently(self):
        member = MemberProfile.objects.create_user(
            username='jane_member',
            password='pass12345',
            role='MEMBER',
        )
        account = SavingsAccount.objects.create(owner=member, label='A1')
        fine = Fine.objects.create(
            member=member,
            account=account,
            fine_type='MISSED_WEEKLY_SAVING',
            reference_week=date(2026, 6, 26),
            reason='Failed to save for account A1 for week closing 2026-06-26',
            amount=Decimal('2000.00'),
            issued_by=self.treasurer,
        )

        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(directory)
            report_path = Path(directory) / 'report.csv'

            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(report_path),
                commit=True,
            )

        self.assertTrue(Fine.objects.filter(id=fine.id, is_paid=False).exists())

    def test_commit_import_keeps_matching_missed_week_fine_when_paid_late(self):
        member = MemberProfile.objects.create_user(
            username='jane_member',
            password='pass12345',
            role='MEMBER',
        )
        account = SavingsAccount.objects.create(owner=member, label='A1')
        fine = Fine.objects.create(
            member=member,
            account=account,
            fine_type='MISSED_WEEKLY_SAVING',
            reference_week=date(2026, 6, 26),
            reason='Failed to save for account A1 for week closing 2026-06-26',
            amount=Decimal('2000.00'),
            issued_by=self.treasurer,
        )

        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(
                directory,
                payment_date='2026-06-29',
            )
            report_path = Path(directory) / 'report.csv'

            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(report_path),
                commit=True,
            )

        fine.refresh_from_db()
        self.assertFalse(fine.is_paid)

    def test_members_import_accepts_full_account_name_labels(self):
        with TemporaryDirectory() as directory:
            members_path = Path(directory) / 'members.csv'
            report_path = Path(directory) / 'report.csv'
            members_path.write_text(
                '\n'.join([
                    'username,first_name,last_name,email,phone_number,role,account_labels,is_active',
                    'long_label_member,Long,Label,long@example.com,+256700000005,MEMBER,Kolyangha Martin Luther,true',
                ]),
                encoding='utf-8',
            )

            call_command(
                'import_historical_data',
                members=str(members_path),
                report=str(report_path),
                commit=True,
            )

            member = MemberProfile.objects.get(username='long_label_member')
            self.assertTrue(
                SavingsAccount.objects.filter(
                    owner=member,
                    label='Kolyangha Martin Luther',
                ).exists()
            )
            self.assertIn('CREATED_MEMBER', report_path.read_text(encoding='utf-8'))

    def test_rerun_skips_duplicate_transactions_by_reference(self):
        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(directory)
            report_path = Path(directory) / 'report.csv'
            duplicate_report_path = Path(directory) / 'duplicate-report.csv'

            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(report_path),
                commit=True,
            )
            call_command(
                'import_historical_data',
                members=str(members_path),
                transactions=str(transactions_path),
                submitted_by='treasurer',
                report=str(duplicate_report_path),
                commit=True,
            )

            self.assertEqual(DepositSubmission.objects.filter(import_reference='OLD-001').count(), 1)
            self.assertIn('SKIPPED_DUPLICATE', duplicate_report_path.read_text(encoding='utf-8'))

    def test_invalid_transaction_report_blocks_import(self):
        with TemporaryDirectory() as directory:
            members_path, transactions_path = self._write_import_files(directory, expected_total='99999')
            report_path = Path(directory) / 'report.csv'

            with self.assertRaises(CommandError):
                call_command(
                    'import_historical_data',
                    members=str(members_path),
                    transactions=str(transactions_path),
                    submitted_by='treasurer',
                    report=str(report_path),
                    commit=True,
                )

            self.assertFalse(MemberProfile.objects.filter(username='jane_member').exists())
            self.assertFalse(DepositSubmission.objects.filter(import_reference='OLD-001').exists())
            report_text = report_path.read_text(encoding='utf-8')
            self.assertIn('ERROR', report_text)
            self.assertIn('does not match amount sum', report_text)


class SavingsWorkbookPreparationCommandTests(SimpleTestCase):
    def _write_members_csv(self, directory):
        members_path = Path(directory) / 'members.csv'
        members_path.write_text(
            '\n'.join([
                'username,first_name,last_name,email,phone_number,role,account_labels,is_active',
                'jane_user,Jane,Member,jane@example.com,+256700000001,MEMBER,Member Jane,true',
                'joe_user,Joe,Member,joe@example.com,+256700000002,MEMBER,Member Joe,true',
            ]),
            encoding='utf-8',
        )
        return members_path

    def _write_savings_workbook(self, directory):
        from openpyxl import Workbook

        workbook_path = Path(directory) / 'savings.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'January'

        headers = ['Saving', 'Fine', 'Welfare', 'Shares', 'Membership', 'Annual Subscription']
        worksheet.cell(row=1, column=3, value=date(2026, 1, 2))
        worksheet.cell(row=1, column=9, value=date(2026, 1, 9))
        worksheet.cell(row=2, column=1, value='S/No')
        worksheet.cell(row=2, column=2, value='Name')
        for index, header in enumerate(headers):
            worksheet.cell(row=2, column=3 + index, value=header)
            worksheet.cell(row=2, column=9 + index, value=header)

        worksheet.cell(row=3, column=1, value=1)
        worksheet.cell(row=3, column=2, value='Jane Member')
        worksheet.cell(row=3, column=3, value=20000)
        worksheet.cell(row=3, column=4, value=2000)
        worksheet.cell(row=3, column=5, value=1000)

        worksheet.cell(row=4, column=1, value=2)
        worksheet.cell(row=4, column=2, value='Member Joe')
        worksheet.cell(row=4, column=3, value=20000)
        worksheet.cell(row=4, column=7, value=10000)

        worksheet.cell(row=5, column=1, value=3)
        worksheet.cell(row=5, column=2, value='Unknown Person')
        worksheet.cell(row=5, column=3, value=20000)

        worksheet.cell(row=3, column=9, value=50000)

        summary = workbook.create_sheet('Summary')
        summary.cell(row=1, column=1, value='Summary should be ignored')
        workbook.save(workbook_path)
        return workbook_path

    def test_prepare_savings_workbook_writes_import_ready_rows_and_review(self):
        with TemporaryDirectory() as directory:
            members_path = self._write_members_csv(directory)
            workbook_path = self._write_savings_workbook(directory)
            transactions_path = Path(directory) / 'transactions.csv'
            review_path = Path(directory) / 'review.csv'
            mapping_review_path = Path(directory) / 'mapping.csv'

            call_command(
                'prepare_savings_workbook_import',
                workbook=str(workbook_path),
                members=str(members_path),
                transactions=str(transactions_path),
                review=str(review_path),
                mapping_review=str(mapping_review_path),
                cutoff_date='2026-01-05',
            )

            with transactions_path.open(newline='', encoding='utf-8') as handle:
                transactions = list(csv.DictReader(handle))
            self.assertEqual(len(transactions), 2)
            transaction = transactions[0]
            self.assertEqual(transaction['username'], 'jane_user')
            self.assertEqual(transaction['account_label'], 'Member Jane')
            self.assertEqual(transaction['payment_week'], '2026-01-02')
            self.assertEqual(transaction['payment_date'], '2026-01-05')
            self.assertEqual(transaction['expected_total'], '23000')
            membership_transaction = transactions[1]
            self.assertEqual(membership_transaction['username'], 'joe_user')
            self.assertEqual(membership_transaction['account_label'], 'Member Joe')
            self.assertEqual(membership_transaction['membership_amount'], '10000')
            self.assertEqual(membership_transaction['expected_total'], '30000')

            review_text = review_path.read_text(encoding='utf-8')
            self.assertIn('READY', review_text)
            self.assertIn('SKIPPED_UNCONFIRMED_ACCOUNT_MATCH', review_text)
            self.assertIn('SKIPPED_FUTURE_WEEK', review_text)
            self.assertIn('Jane Member', mapping_review_path.read_text(encoding='utf-8'))


class UsernameUpdateCommandTests(TestCase):
    def test_dry_run_does_not_change_username(self):
        MemberProfile.objects.create_user(
            username='kolyangha_martin_luther',
            email='martin@example.com',
            password='pass12345',
        )
        with TemporaryDirectory() as directory:
            mapping_path = Path(directory) / 'mapping.csv'
            report_path = Path(directory) / 'report.csv'
            mapping_path.write_text(
                '\n'.join([
                    'old_username,new_username,email,name',
                    'kolyangha_martin_luther,KolyanghaM,martin@example.com,Kolyangha Martin Luther',
                ]),
                encoding='utf-8',
            )

            call_command(
                'update_usernames_from_csv',
                file=str(mapping_path),
                report=str(report_path),
            )

            self.assertTrue(MemberProfile.objects.filter(username='kolyangha_martin_luther').exists())
            self.assertFalse(MemberProfile.objects.filter(username='KolyanghaM').exists())
            self.assertIn('VALID_UPDATE', report_path.read_text(encoding='utf-8'))

    def test_commit_updates_username(self):
        MemberProfile.objects.create_user(
            username='kolyangha_martin_luther',
            email='martin@example.com',
            password='pass12345',
        )
        with TemporaryDirectory() as directory:
            mapping_path = Path(directory) / 'mapping.csv'
            report_path = Path(directory) / 'report.csv'
            mapping_path.write_text(
                '\n'.join([
                    'old_username,new_username,email,name',
                    'kolyangha_martin_luther,KolyanghaM,martin@example.com,Kolyangha Martin Luther',
                ]),
                encoding='utf-8',
            )

            call_command(
                'update_usernames_from_csv',
                file=str(mapping_path),
                report=str(report_path),
                commit=True,
            )

            self.assertFalse(MemberProfile.objects.filter(username='kolyangha_martin_luther').exists())
            self.assertTrue(MemberProfile.objects.filter(username='KolyanghaM').exists())
            self.assertIn('UPDATED_USERNAME', report_path.read_text(encoding='utf-8'))

    def test_conflicting_new_username_blocks_updates(self):
        MemberProfile.objects.create_user(
            username='kolyangha_martin_luther',
            email='martin@example.com',
            password='pass12345',
        )
        MemberProfile.objects.create_user(
            username='KolyanghaM',
            email='other@example.com',
            password='pass12345',
        )
        with TemporaryDirectory() as directory:
            mapping_path = Path(directory) / 'mapping.csv'
            report_path = Path(directory) / 'report.csv'
            mapping_path.write_text(
                '\n'.join([
                    'old_username,new_username,email,name',
                    'kolyangha_martin_luther,KolyanghaM,martin@example.com,Kolyangha Martin Luther',
                ]),
                encoding='utf-8',
            )

            with self.assertRaises(CommandError):
                call_command(
                    'update_usernames_from_csv',
                    file=str(mapping_path),
                    report=str(report_path),
                    commit=True,
                )

            self.assertTrue(MemberProfile.objects.filter(username='kolyangha_martin_luther').exists())
            self.assertIn('already belongs to another user', report_path.read_text(encoding='utf-8'))
