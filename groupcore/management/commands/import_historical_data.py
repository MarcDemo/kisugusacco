import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from deposits.models import DepositSubmission
from groupcore.models import MemberProfile, SavingsAccount


ACCOUNT_LABEL_RE = re.compile(r'^[A-Za-z0-9 ._-]+$')
DATE_FORMATS = ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y')
TIME_FORMATS = ('%H:%M:%S', '%H:%M')
PURPOSE_FIELDS = (
    'saving_amount',
    'welfare_amount',
    'annual_subscription_amount',
    'fine_amount',
    'shares_amount',
    'loan_repayment_amount',
)
HEADER_ALIASES = {
    'account': 'account_label',
    'savings_account': 'account_label',
    'savings_account_label': 'account_label',
    'member_username': 'username',
    'user_name': 'username',
    'transaction_id': 'transaction_reference',
    'reference': 'transaction_reference',
    'payment_reference': 'transaction_reference',
    'week': 'payment_week',
    'saving_week': 'payment_week',
    'total': 'expected_total',
    'total_amount': 'expected_total',
}


@dataclass
class MemberPayload:
    row_number: int
    report_index: int
    username: str
    fields: dict
    account_labels: list
    exists: bool


@dataclass
class TransactionPayload:
    row_number: int
    report_index: int
    import_reference: str
    username: str
    account_label: str
    payment_week: date
    payment_date: date
    payment_time: time
    amounts: dict
    status: str
    remarks: str
    proof_reference: str
    total: Decimal


class Command(BaseCommand):
    help = (
        "Validate and import historical member accounts and savings transactions. "
        "Runs in dry-run mode unless --commit is supplied."
    )

    def add_arguments(self, parser):
        parser.add_argument('--members', help='CSV/XLSX file containing member accounts.')
        parser.add_argument('--transactions', help='CSV/XLSX file containing historical deposits.')
        parser.add_argument('--members-sheet', default='Members')
        parser.add_argument('--transactions-sheet', default='Transactions')
        parser.add_argument('--report', help='Path for the import validation report CSV.')
        parser.add_argument('--commit', action='store_true', help='Write validated rows to the database.')
        parser.add_argument(
            '--update-existing-members',
            action='store_true',
            help='Update existing member profile fields from the members file.',
        )
        parser.add_argument(
            '--submitted-by',
            help='Username to store as submitted/reviewed by for imported transactions.',
        )
        parser.add_argument(
            '--batch',
            help='Batch label stored on imported transactions. Defaults to historical-import-YYYYMMDD-HHMMSS.',
        )

    def handle(self, *args, **options):
        if not options['members'] and not options['transactions']:
            raise CommandError('Provide --members, --transactions, or both.')

        report_path = Path(
            options['report']
            or f"import_report_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        batch_label = options['batch'] or f"historical-import-{timezone.localtime().strftime('%Y%m%d-%H%M%S')}"
        report_rows = []

        member_rows = self._load_rows(options['members'], options['members_sheet']) if options['members'] else []
        transaction_rows = (
            self._load_rows(options['transactions'], options['transactions_sheet'])
            if options['transactions']
            else []
        )

        member_payloads, planned_accounts, member_errors = self._validate_members(
            member_rows,
            report_rows,
        )
        transaction_payloads, transaction_errors = self._validate_transactions(
            transaction_rows,
            planned_accounts,
            report_rows,
        )

        errors = member_errors + transaction_errors
        if errors:
            self._write_report(report_path, report_rows)
            raise CommandError(
                f"Import blocked: {len(errors)} row(s) need correction. "
                f"Review {report_path}."
            )

        if not options['commit']:
            self._write_report(report_path, report_rows)
            self.stdout.write(
                self.style.SUCCESS(
                    f"Dry run passed. No data was changed. Review {report_path}, "
                    "then rerun with --commit to import."
                )
            )
            return

        submitted_by = self._get_submitted_by(options.get('submitted_by'))
        with transaction.atomic():
            self._import_members(
                member_payloads,
                report_rows,
                update_existing=options['update_existing_members'],
            )
            self._import_transactions(
                transaction_payloads,
                report_rows,
                submitted_by=submitted_by,
                batch_label=batch_label,
            )

        self._write_report(report_path, report_rows)
        self.stdout.write(
            self.style.SUCCESS(
                f"Import completed. Review {report_path} for created, existing, skipped, and error rows."
            )
        )

    def _load_rows(self, file_path, sheet_name):
        path = Path(file_path)
        if not path.exists():
            raise CommandError(f'Import file not found: {path}')

        suffix = path.suffix.lower()
        if suffix == '.csv':
            return self._load_csv_rows(path)
        if suffix in {'.xlsx', '.xlsm'}:
            return self._load_xlsx_rows(path, sheet_name)
        raise CommandError(f'Unsupported import file type: {path.suffix}. Use CSV or XLSX.')

    def _load_csv_rows(self, path):
        with path.open(newline='', encoding='utf-8-sig') as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise CommandError(f'{path} has no header row.')
            rows = []
            for row_number, row in enumerate(reader, start=2):
                normalised = {
                    self._normalise_header(key): self._clean_value(value)
                    for key, value in row.items()
                    if key is not None
                }
                normalised['__row_number'] = row_number
                rows.append(normalised)
            return rows

    def _load_xlsx_rows(self, path, sheet_name):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('openpyxl is required to import XLSX files.') from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise CommandError(f'{path} is empty.') from exc

        normalised_headers = [self._normalise_header(header) for header in headers]
        data = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, '') for value in values):
                continue
            row = {
                header: self._clean_value(value)
                for header, value in zip(normalised_headers, values)
                if header
            }
            row['__row_number'] = row_number
            data.append(row)
        return data

    def _validate_members(self, rows, report_rows):
        payloads = []
        errors = []
        seen_usernames = set()
        planned_accounts = {}
        roles = {key for key, _label in MemberProfile.ROLE_CHOICES}

        for row in rows:
            row_number = row['__row_number']
            username = row.get('username', '').strip()
            row_errors = []

            if not username:
                row_errors.append('username is required')
            elif username.lower() in seen_usernames:
                row_errors.append(f'duplicate username in file: {username}')
            else:
                seen_usernames.add(username.lower())

            if len(username) > 150:
                row_errors.append('username must be 150 characters or fewer')

            role = (row.get('role') or 'MEMBER').strip().upper()
            if role not in roles:
                row_errors.append(f'role must be one of: {", ".join(sorted(roles))}')

            account_labels = self._parse_account_labels(row)
            if not account_labels:
                row_errors.append('account_labels is required')
            for label in account_labels:
                if len(label) > 20:
                    row_errors.append(f'account label "{label}" must be 20 characters or fewer')
                if not ACCOUNT_LABEL_RE.match(label):
                    row_errors.append(f'account label "{label}" has invalid characters')

            exists = MemberProfile.objects.filter(username=username).exists() if username else False
            fields = {
                'first_name': row.get('first_name', ''),
                'last_name': row.get('last_name', ''),
                'email': row.get('email', ''),
                'phone_number': row.get('phone_number', ''),
                'next_of_kin_name': row.get('next_of_kin_name', ''),
                'next_of_kin_contact': row.get('next_of_kin_contact', ''),
                'role': role,
                'is_active': self._parse_bool(row.get('is_active', 'true')),
            }

            status = 'ERROR' if row_errors else ('VALID_EXISTING_MEMBER' if exists else 'VALID_NEW_MEMBER')
            report_index = self._add_report_row(
                report_rows,
                section='members',
                row_number=row_number,
                status=status,
                username=username,
                account_label=';'.join(account_labels),
                message='; '.join(row_errors),
            )

            if row_errors:
                errors.extend(row_errors)
                continue

            payloads.append(
                MemberPayload(
                    row_number=row_number,
                    report_index=report_index,
                    username=username,
                    fields=fields,
                    account_labels=account_labels,
                    exists=exists,
                )
            )
            for label in account_labels:
                planned_accounts[(username.lower(), label.lower())] = label

        return payloads, planned_accounts, errors

    def _validate_transactions(self, rows, planned_accounts, report_rows):
        payloads = []
        errors = []
        seen_references = set()
        statuses = {key for key, _label in DepositSubmission.STATUS_CHOICES}
        existing_references = {
            reference.lower()
            for reference in (
                DepositSubmission.objects
                .filter(import_reference__isnull=False)
                .values_list('import_reference', flat=True)
            )
            if reference
        }
        planned_members = {username for username, _label in planned_accounts}

        for row in rows:
            row_number = row['__row_number']
            row_errors = []
            import_reference = row.get('transaction_reference', '').strip()
            username = row.get('username', '').strip()
            account_label = row.get('account_label', '').strip()

            if not import_reference:
                row_errors.append('transaction_reference is required')
            elif len(import_reference) > 120:
                row_errors.append('transaction_reference must be 120 characters or fewer')
            elif import_reference.lower() in seen_references:
                row_errors.append(f'duplicate transaction_reference in file: {import_reference}')
            else:
                seen_references.add(import_reference.lower())

            if not username:
                row_errors.append('username is required')

            if not account_label:
                row_errors.append('account_label is required')

            member_exists = MemberProfile.objects.filter(username=username).exists() if username else False
            planned_member_exists = username.lower() in planned_members
            planned_account_exists = (username.lower(), account_label.lower()) in planned_accounts
            db_account_exists = SavingsAccount.objects.filter(
                owner__username=username,
                label__iexact=account_label,
            ).exists() if username and account_label else False
            if username and not member_exists and not planned_member_exists:
                row_errors.append(f'member "{username}" does not exist and is not in the members file')
            if username and account_label and not planned_account_exists and not db_account_exists:
                row_errors.append(
                    f'account "{account_label}" for member "{username}" does not exist and is not in the members file'
                )

            payment_week = self._parse_date(row.get('payment_week'), 'payment_week', row_errors)
            if payment_week and payment_week.weekday() != 0:
                row_errors.append('payment_week must be a Monday/week-start date')
            payment_date = self._parse_date(row.get('payment_date'), 'payment_date', row_errors)
            payment_time = self._parse_time(row.get('payment_time') or '00:00', 'payment_time', row_errors)

            amounts = {}
            for field in PURPOSE_FIELDS:
                amounts[field] = self._parse_decimal(row.get(field, '0'), field, row_errors)

            total = sum(amounts.values(), Decimal('0.00'))
            if total <= 0:
                row_errors.append('at least one amount column must be greater than zero')

            expected_total_raw = row.get('expected_total', '')
            if expected_total_raw not in ('', None):
                expected_total = self._parse_decimal(expected_total_raw, 'expected_total', row_errors)
                if expected_total != total:
                    row_errors.append(f'expected_total {expected_total} does not match amount sum {total}')

            status = (row.get('status') or 'APPROVED').strip().upper()
            if status not in statuses:
                row_errors.append(f'status must be one of: {", ".join(sorted(statuses))}')

            existing_duplicate = import_reference.lower() in existing_references
            if existing_duplicate and not row_errors:
                self._add_report_row(
                    report_rows,
                    section='transactions',
                    row_number=row_number,
                    status='SKIPPED_DUPLICATE',
                    username=username,
                    account_label=account_label,
                    transaction_reference=import_reference,
                    amount=total,
                    message='transaction_reference already exists',
                )
                continue

            report_index = self._add_report_row(
                report_rows,
                section='transactions',
                row_number=row_number,
                status='ERROR' if row_errors else 'VALID_NEW_TRANSACTION',
                username=username,
                account_label=account_label,
                transaction_reference=import_reference,
                amount=total,
                message='; '.join(row_errors),
            )

            if row_errors:
                errors.extend(row_errors)
                continue

            payloads.append(
                TransactionPayload(
                    row_number=row_number,
                    report_index=report_index,
                    import_reference=import_reference,
                    username=username,
                    account_label=account_label,
                    payment_week=payment_week,
                    payment_date=payment_date,
                    payment_time=payment_time,
                    amounts=amounts,
                    status=status,
                    remarks=row.get('remarks', ''),
                    proof_reference=row.get('proof_reference', '') or 'historical_import/no-proof.jpg',
                    total=total,
                )
            )

        return payloads, errors

    def _import_members(self, payloads, report_rows, update_existing=False):
        for payload in payloads:
            user = MemberProfile.objects.filter(username=payload.username).first()
            if user:
                if update_existing:
                    for field, value in payload.fields.items():
                        setattr(user, field, value)
                    user.save()
                    status = 'UPDATED_MEMBER'
                else:
                    status = 'EXISTING_MEMBER'
            else:
                user = MemberProfile(username=payload.username, **payload.fields)
                user.set_unusable_password()
                user.save()
                status = 'CREATED_MEMBER'

            created_accounts = []
            existing_accounts = []
            for label in payload.account_labels:
                _account, created = SavingsAccount.objects.get_or_create(
                    owner=user,
                    label=label,
                    defaults={'is_active': True},
                )
                if created:
                    created_accounts.append(label)
                else:
                    existing_accounts.append(label)

            message_parts = []
            if created_accounts:
                message_parts.append(f"created accounts: {', '.join(created_accounts)}")
            if existing_accounts:
                message_parts.append(f"existing accounts: {', '.join(existing_accounts)}")
            self._update_report_row(
                report_rows,
                payload.report_index,
                status=status,
                message='; '.join(message_parts),
            )

    def _import_transactions(self, payloads, report_rows, submitted_by, batch_label):
        now = timezone.now()
        deposits = []
        for payload in payloads:
            if DepositSubmission.objects.filter(import_reference=payload.import_reference).exists():
                self._update_report_row(
                    report_rows,
                    payload.report_index,
                    status='SKIPPED_DUPLICATE',
                    message='transaction_reference already exists',
                )
                continue

            member = MemberProfile.objects.get(username=payload.username)
            account = SavingsAccount.objects.get(owner=member, label__iexact=payload.account_label)
            reviewed_by = submitted_by if payload.status in {'APPROVED', 'REJECTED'} else None
            date_reviewed = now if reviewed_by else None
            deposits.append(
                DepositSubmission(
                    member=member,
                    account=account,
                    payment_week=payload.payment_week,
                    starting_week=payload.payment_week,
                    weeks_covered=1,
                    saving_amount=payload.amounts['saving_amount'],
                    welfare_amount=payload.amounts['welfare_amount'],
                    annual_subscription_amount=payload.amounts['annual_subscription_amount'],
                    fine_amount=payload.amounts['fine_amount'],
                    shares_amount=payload.amounts['shares_amount'],
                    loan_repayment_amount=payload.amounts['loan_repayment_amount'],
                    amount=payload.total,
                    proof=payload.proof_reference,
                    remarks=payload.remarks,
                    payment_date=payload.payment_date,
                    payment_time=payload.payment_time,
                    status=payload.status,
                    submitted_by=submitted_by,
                    reviewed_by=reviewed_by,
                    date_submitted=now,
                    date_reviewed=date_reviewed,
                    import_reference=payload.import_reference,
                    import_batch=batch_label,
                    imported_at=now,
                )
            )
            self._update_report_row(
                report_rows,
                payload.report_index,
                status='CREATED_TRANSACTION',
                message='imported',
            )

        if deposits:
            DepositSubmission.objects.bulk_create(deposits)

    def _get_submitted_by(self, username):
        if not username:
            return None
        try:
            return MemberProfile.objects.get(username=username)
        except MemberProfile.DoesNotExist as exc:
            raise CommandError(f'--submitted-by user not found: {username}') from exc

    def _parse_account_labels(self, row):
        raw = row.get('account_labels', '')
        if not raw:
            raw = row.get('account_label', '')
        labels = []
        seen = set()
        for item in re.split(r'[;|\n]+', str(raw)):
            label = item.strip()
            key = label.lower()
            if label and key not in seen:
                labels.append(label)
                seen.add(key)
        return labels

    def _parse_date(self, value, field_name, errors):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        value = str(value or '').strip()
        if not value:
            errors.append(f'{field_name} is required')
            return None
        for date_format in DATE_FORMATS:
            try:
                return datetime.strptime(value, date_format).date()
            except ValueError:
                continue
        errors.append(f'{field_name} must be a date like 2026-06-22')
        return None

    def _parse_time(self, value, field_name, errors):
        if isinstance(value, datetime):
            return value.time().replace(microsecond=0)
        if isinstance(value, time):
            return value.replace(microsecond=0)
        value = str(value or '').strip()
        if not value:
            return time(0, 0)
        for time_format in TIME_FORMATS:
            try:
                return datetime.strptime(value, time_format).time()
            except ValueError:
                continue
        errors.append(f'{field_name} must be a time like 09:30')
        return None

    def _parse_decimal(self, value, field_name, errors):
        value = str(value or '0').strip().replace(',', '').replace('UGX', '').strip()
        try:
            amount = Decimal(value or '0').quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            errors.append(f'{field_name} must be a valid number')
            return Decimal('0.00')
        if amount < 0:
            errors.append(f'{field_name} cannot be negative')
        return amount

    def _parse_bool(self, value):
        return str(value).strip().lower() not in {'0', 'false', 'no', 'n', 'inactive'}

    def _normalise_header(self, value):
        key = str(value or '').strip().lower()
        key = re.sub(r'[^a-z0-9]+', '_', key).strip('_')
        return HEADER_ALIASES.get(key, key)

    def _clean_value(self, value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return value

    def _add_report_row(
        self,
        report_rows,
        section,
        row_number,
        status,
        username='',
        account_label='',
        transaction_reference='',
        amount='',
        message='',
    ):
        report_rows.append({
            'section': section,
            'row': row_number,
            'status': status,
            'username': username,
            'account_label': account_label,
            'transaction_reference': transaction_reference,
            'amount': amount,
            'message': message,
        })
        return len(report_rows) - 1

    def _update_report_row(self, report_rows, index, status, message=''):
        report_rows[index]['status'] = status
        report_rows[index]['message'] = message

    def _write_report(self, report_path, report_rows):
        report_path.parent.mkdir(parents=True, exist_ok=True)
        with report_path.open('w', newline='', encoding='utf-8') as handle:
            fieldnames = [
                'section',
                'row',
                'status',
                'username',
                'account_label',
                'transaction_reference',
                'amount',
                'message',
            ]
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(report_rows)
