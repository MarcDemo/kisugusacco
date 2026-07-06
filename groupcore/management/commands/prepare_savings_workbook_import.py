import csv
import difflib
import hashlib
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.utils.text import slugify


AMOUNT_HEADERS = (
    'Saving',
    'Fine',
    'Welfare',
    'Shares',
    'Membership',
    'Annual Subscription',
)
SUPPORTED_IMPORT_FIELDS = {
    'Saving': 'saving_amount',
    'Welfare': 'welfare_amount',
    'Annual Subscription': 'annual_subscription_amount',
    'Membership': 'membership_amount',
    'Fine': 'fine_amount',
    'Shares': 'shares_amount',
}
TRANSACTION_FIELDNAMES = [
    'transaction_reference',
    'username',
    'account_label',
    'payment_week',
    'payment_date',
    'payment_time',
    'saving_amount',
    'welfare_amount',
    'annual_subscription_amount',
    'membership_amount',
    'fine_amount',
    'shares_amount',
    'loan_repayment_amount',
    'expected_total',
    'status',
    'remarks',
    'proof_reference',
]
REVIEW_FIELDNAMES = [
    'status',
    'issues',
    'sheet',
    'source_row',
    'payment_week',
    'workbook_account_label',
    'mapped_username',
    'mapped_account_label',
    'mapping_status',
    'suggested_username',
    'suggested_account_label',
    'suggestion_score',
    'saving_amount',
    'welfare_amount',
    'annual_subscription_amount',
    'fine_amount',
    'shares_amount',
    'membership_amount',
    'loan_repayment_amount',
    'supported_total',
    'workbook_total',
    'payment_date_for_import',
    'transaction_reference',
    'remarks',
]
MAPPING_REVIEW_FIELDNAMES = [
    'workbook_account_label',
    'mapping_status',
    'mapped_username',
    'mapped_account_label',
    'suggested_username',
    'suggested_account_label',
    'suggestion_score',
    'occurrences',
    'supported_total',
    'membership_total',
    'workbook_total',
    'action_needed',
]


class Command(BaseCommand):
    help = (
        "Convert the group's monthly savings workbook into the historical "
        "transactions CSV format. The generated CSV is import-ready only for "
        "rows whose account labels can be matched safely."
    )

    def add_arguments(self, parser):
        parser.add_argument('--workbook', required=True, help='Savings workbook, for example Savings 2026.xlsx.')
        parser.add_argument('--members', required=True, help='Members CSV containing username and account_labels.')
        parser.add_argument('--transactions', default='imports/transactions.csv')
        parser.add_argument('--review', default='imports/transactions_import_review.csv')
        parser.add_argument('--mapping-review', default='imports/savings_account_mapping_review.csv')
        parser.add_argument(
            '--cutoff-date',
            default=timezone.localdate().isoformat(),
            help='Latest payment_week to export. Defaults to today.',
        )
        parser.add_argument(
            '--include-future',
            action='store_true',
            help='Include future-dated payment weeks instead of flagging them in the review only.',
        )

    def handle(self, *args, **options):
        workbook_path = Path(options['workbook'])
        members_path = Path(options['members'])
        transactions_path = Path(options['transactions'])
        review_path = Path(options['review'])
        mapping_review_path = Path(options['mapping_review'])
        cutoff_date = self._parse_cutoff_date(options['cutoff_date'])

        if not workbook_path.exists():
            raise CommandError(f'Workbook not found: {workbook_path}')
        if not members_path.exists():
            raise CommandError(f'Members CSV not found: {members_path}')

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('openpyxl is required to read the savings workbook.') from exc

        account_index = self._load_account_index(members_path)
        workbook = load_workbook(workbook_path, data_only=True)
        prepared_rows, review_rows, mapping_rows = self._prepare_rows(
            workbook=workbook,
            account_index=account_index,
            cutoff_date=cutoff_date,
            include_future=options['include_future'],
            workbook_name=workbook_path.name,
        )

        self._write_csv(transactions_path, TRANSACTION_FIELDNAMES, prepared_rows)
        self._write_csv(review_path, REVIEW_FIELDNAMES, review_rows)
        self._write_csv(mapping_review_path, MAPPING_REVIEW_FIELDNAMES, mapping_rows)

        status_counts = Counter(row['status'] for row in review_rows)
        summary = ', '.join(f'{status}: {count}' for status, count in sorted(status_counts.items()))
        self.stdout.write(
            self.style.SUCCESS(
                f'Prepared {len(prepared_rows)} transaction row(s) in {transactions_path}. '
                f'Review {len(review_rows)} parsed row(s) in {review_path}. '
                f'Status counts: {summary or "none"}.'
            )
        )
        self.stdout.write(f'Account mapping review written to {mapping_review_path}.')

    def _load_account_index(self, members_path):
        accounts = []
        with members_path.open(newline='', encoding='utf-8-sig') as handle:
            reader = csv.DictReader(handle)
            required = {'username', 'account_labels'}
            missing = required - set(reader.fieldnames or [])
            if missing:
                raise CommandError(f'Members CSV missing required columns: {", ".join(sorted(missing))}')
            for row in reader:
                username = (row.get('username') or '').strip()
                if not username:
                    continue
                for label in re.split(r'[;|\n]+', row.get('account_labels') or ''):
                    account_label = label.strip()
                    if account_label:
                        accounts.append({
                            'username': username,
                            'account_label': account_label,
                            'normalised': self._normalise_name(account_label),
                            'token_key': self._token_key(account_label),
                        })

        by_normalised = defaultdict(list)
        by_token = defaultdict(list)
        for account in accounts:
            by_normalised[account['normalised']].append(account)
            by_token[account['token_key']].append(account)
        return {
            'accounts': accounts,
            'by_normalised': by_normalised,
            'by_token': by_token,
        }

    def _prepare_rows(self, workbook, account_index, cutoff_date, include_future, workbook_name):
        prepared_rows = []
        review_rows = []
        mapping_stats = defaultdict(lambda: {
            'occurrences': 0,
            'supported_total': Decimal('0'),
            'membership_total': Decimal('0'),
            'workbook_total': Decimal('0'),
            'mapping': None,
        })
        seen_references = set()

        for worksheet in workbook.worksheets:
            if worksheet.title.strip().lower() == 'summary':
                continue
            for start_column, payment_week in self._dated_week_blocks(worksheet):
                for row_number in range(3, worksheet.max_row + 1):
                    raw_label = worksheet.cell(row=row_number, column=2).value
                    workbook_label = str(raw_label or '').strip()
                    if not workbook_label or workbook_label.lower() == 'total':
                        continue

                    amount_errors = []
                    amounts = self._read_amounts(worksheet, row_number, start_column, amount_errors)
                    supported_total = sum(
                        amounts[header]
                        for header in SUPPORTED_IMPORT_FIELDS
                    )
                    membership_amount = amounts['Membership']
                    workbook_total = supported_total
                    if workbook_total <= 0:
                        continue

                    mapping = self._match_account(workbook_label, account_index)
                    mapping_stats[workbook_label]['occurrences'] += 1
                    mapping_stats[workbook_label]['supported_total'] += supported_total
                    mapping_stats[workbook_label]['membership_total'] += membership_amount
                    mapping_stats[workbook_label]['workbook_total'] += workbook_total
                    mapping_stats[workbook_label]['mapping'] = mapping

                    issues = list(amount_errors)
                    if payment_week.weekday() != 4:
                        issues.append('payment_week is not a Friday')
                    if payment_week > cutoff_date and not include_future:
                        issues.append(f'future payment_week after cutoff {cutoff_date.isoformat()}')
                    if mapping['status'] not in {'EXACT', 'TOKEN_REORDER'}:
                        issues.append('account match needs confirmation')
                    if supported_total <= 0:
                        issues.append('no supported deposit amount to import')

                    transaction_reference = ''
                    payment_date = self._inferred_payment_date(payment_week, amounts['Fine'])
                    if not issues:
                        transaction_reference = self._transaction_reference(
                            payment_week=payment_week,
                            username=mapping['username'],
                            account_label=mapping['account_label'],
                        )
                        if transaction_reference in seen_references:
                            issues.append('duplicate generated transaction_reference')
                        else:
                            seen_references.add(transaction_reference)

                    status = 'READY' if not issues else self._skipped_status(issues)
                    review_row = self._review_row(
                        status=status,
                        issues=issues,
                        worksheet=worksheet,
                        row_number=row_number,
                        payment_week=payment_week,
                        workbook_label=workbook_label,
                        mapping=mapping,
                        amounts=amounts,
                        supported_total=supported_total,
                        workbook_total=workbook_total,
                        payment_date=payment_date,
                        transaction_reference=transaction_reference,
                    )
                    review_rows.append(review_row)

                    if status == 'READY':
                        prepared_rows.append(self._transaction_row(
                            transaction_reference=transaction_reference,
                            mapping=mapping,
                            payment_week=payment_week,
                            payment_date=payment_date,
                            amounts=amounts,
                            supported_total=supported_total,
                            workbook_name=workbook_name,
                            worksheet=worksheet,
                            row_number=row_number,
                        ))

        mapping_rows = self._mapping_review_rows(mapping_stats)
        return prepared_rows, review_rows, mapping_rows

    def _dated_week_blocks(self, worksheet):
        for column in range(3, worksheet.max_column + 1):
            payment_week = self._parse_week_header(worksheet.cell(row=1, column=column).value)
            if not payment_week:
                continue
            headers = [
                self._normalise_header(worksheet.cell(row=2, column=column + offset).value)
                for offset in range(len(AMOUNT_HEADERS))
                if column + offset <= worksheet.max_column
            ]
            if headers == list(AMOUNT_HEADERS):
                yield column, payment_week

    def _read_amounts(self, worksheet, row_number, start_column, errors):
        amounts = {}
        for offset, header in enumerate(AMOUNT_HEADERS):
            value = worksheet.cell(row=row_number, column=start_column + offset).value
            amounts[header] = self._parse_amount(value, header, errors)
        return amounts

    def _match_account(self, workbook_label, account_index):
        normalised = self._normalise_name(workbook_label)
        token_key = self._token_key(workbook_label)

        exact_matches = account_index['by_normalised'].get(normalised, [])
        if len(exact_matches) == 1:
            return self._mapping('EXACT', exact_matches[0], score='1.00')

        token_matches = account_index['by_token'].get(token_key, [])
        if len(token_matches) == 1:
            return self._mapping('TOKEN_REORDER', token_matches[0], score='1.00')

        suggestion = self._best_suggestion(token_key, account_index['accounts'])
        if suggestion:
            score, account = suggestion
            return {
                'status': f'SUGGESTED_{score:.2f}',
                'username': '',
                'account_label': '',
                'suggested_username': account['username'],
                'suggested_account_label': account['account_label'],
                'suggestion_score': f'{score:.2f}',
            }
        return {
            'status': 'UNMATCHED',
            'username': '',
            'account_label': '',
            'suggested_username': '',
            'suggested_account_label': '',
            'suggestion_score': '',
        }

    def _best_suggestion(self, token_key, accounts):
        if not accounts:
            return None
        suggestions = sorted(
            (
                (
                    difflib.SequenceMatcher(None, token_key, account['token_key']).ratio(),
                    account,
                )
                for account in accounts
            ),
            key=lambda item: item[0],
            reverse=True,
        )
        return suggestions[0]

    def _mapping(self, status, account, score):
        return {
            'status': status,
            'username': account['username'],
            'account_label': account['account_label'],
            'suggested_username': '',
            'suggested_account_label': '',
            'suggestion_score': score,
        }

    def _review_row(
        self,
        status,
        issues,
        worksheet,
        row_number,
        payment_week,
        workbook_label,
        mapping,
        amounts,
        supported_total,
        workbook_total,
        payment_date,
        transaction_reference,
    ):
        return {
            'status': status,
            'issues': '; '.join(issues),
            'sheet': worksheet.title,
            'source_row': row_number,
            'payment_week': payment_week.isoformat(),
            'workbook_account_label': workbook_label,
            'mapped_username': mapping['username'],
            'mapped_account_label': mapping['account_label'],
            'mapping_status': mapping['status'],
            'suggested_username': mapping['suggested_username'],
            'suggested_account_label': mapping['suggested_account_label'],
            'suggestion_score': mapping['suggestion_score'],
            'saving_amount': self._format_amount(amounts['Saving']),
            'welfare_amount': self._format_amount(amounts['Welfare']),
            'annual_subscription_amount': self._format_amount(amounts['Annual Subscription']),
            'membership_amount': self._format_amount(amounts['Membership']),
            'fine_amount': self._format_amount(amounts['Fine']),
            'shares_amount': self._format_amount(amounts['Shares']),
            'loan_repayment_amount': '0',
            'supported_total': self._format_amount(supported_total),
            'workbook_total': self._format_amount(workbook_total),
            'payment_date_for_import': payment_date.isoformat(),
            'transaction_reference': transaction_reference,
            'remarks': 'Fine present, payment date inferred after Sunday grace'
            if amounts['Fine'] > 0 else 'Payment date inferred as saving Friday',
        }

    def _transaction_row(
        self,
        transaction_reference,
        mapping,
        payment_week,
        payment_date,
        amounts,
        supported_total,
        workbook_name,
        worksheet,
        row_number,
    ):
        return {
            'transaction_reference': transaction_reference,
            'username': mapping['username'],
            'account_label': mapping['account_label'],
            'payment_week': payment_week.isoformat(),
            'payment_date': payment_date.isoformat(),
            'payment_time': '00:00',
            'saving_amount': self._format_amount(amounts['Saving']),
            'welfare_amount': self._format_amount(amounts['Welfare']),
            'annual_subscription_amount': self._format_amount(amounts['Annual Subscription']),
            'membership_amount': self._format_amount(amounts['Membership']),
            'fine_amount': self._format_amount(amounts['Fine']),
            'shares_amount': self._format_amount(amounts['Shares']),
            'loan_repayment_amount': '0',
            'expected_total': self._format_amount(supported_total),
            'status': 'APPROVED',
            'remarks': (
                f'Historical import from {workbook_name}, sheet {worksheet.title}, row {row_number}. '
                + (
                    'Payment date inferred after Sunday grace because a fine is recorded.'
                    if amounts['Fine'] > 0 else
                    'Payment date inferred as the Friday saving close date.'
                )
            ),
            'proof_reference': 'historical_import/no-proof.jpg',
        }

    def _mapping_review_rows(self, mapping_stats):
        rows = []
        for workbook_label, stats in sorted(mapping_stats.items(), key=lambda item: item[0].lower()):
            mapping = stats['mapping'] or {}
            action_needed = ''
            if mapping.get('status') not in {'EXACT', 'TOKEN_REORDER'}:
                action_needed = 'Confirm or add this account/member before importing its transactions.'
            rows.append({
                'workbook_account_label': workbook_label,
                'mapping_status': mapping.get('status', ''),
                'mapped_username': mapping.get('username', ''),
                'mapped_account_label': mapping.get('account_label', ''),
                'suggested_username': mapping.get('suggested_username', ''),
                'suggested_account_label': mapping.get('suggested_account_label', ''),
                'suggestion_score': mapping.get('suggestion_score', ''),
                'occurrences': stats['occurrences'],
                'supported_total': self._format_amount(stats['supported_total']),
                'membership_total': self._format_amount(stats['membership_total']),
                'workbook_total': self._format_amount(stats['workbook_total']),
                'action_needed': action_needed,
            })
        return rows

    def _transaction_reference(self, payment_week, username, account_label):
        raw_key = f'{payment_week.isoformat()}|{username}|{account_label}'
        digest = hashlib.sha1(raw_key.encode('utf-8')).hexdigest()[:8]
        label_slug = slugify(account_label)[:60] or 'account'
        return f'SAV26-{payment_week:%Y%m%d}-{label_slug}-{digest}'

    def _skipped_status(self, issues):
        joined = '; '.join(issues)
        if 'future payment_week' in joined:
            return 'SKIPPED_FUTURE_WEEK'
        if 'account match needs confirmation' in issues:
            return 'SKIPPED_UNCONFIRMED_ACCOUNT_MATCH'
        if 'duplicate generated transaction_reference' in issues:
            return 'SKIPPED_DUPLICATE_REFERENCE'
        return 'SKIPPED_REVIEW_NEEDED'

    def _inferred_payment_date(self, payment_week, fine_amount):
        if fine_amount > 0:
            return payment_week + timedelta(days=3)
        return payment_week

    def _parse_week_header(self, value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or '').strip()
        if not text:
            return None
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r',\s*', ', ', text)
        for date_format in ('%A, %B %d, %Y', '%A, %b %d, %Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                continue
        return None

    def _parse_cutoff_date(self, value):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError as exc:
            raise CommandError('--cutoff-date must be in YYYY-MM-DD format') from exc

    def _parse_amount(self, value, header, errors):
        if value in (None, ''):
            return Decimal('0')
        if isinstance(value, str) and not value.strip():
            return Decimal('0')
        try:
            return Decimal(str(value).replace(',', '').strip())
        except (InvalidOperation, AttributeError):
            errors.append(f'{header} amount is not numeric: {value}')
            return Decimal('0')

    def _normalise_header(self, value):
        text = self._normalise_name(value)
        aliases = {
            'saving': 'Saving',
            'fine': 'Fine',
            'welfare': 'Welfare',
            'shares': 'Shares',
            'membership': 'Membership',
            'members': 'Membership',
            'annual subscription': 'Annual Subscription',
            'annual subscriptions': 'Annual Subscription',
            'annual sub': 'Annual Subscription',
            'annual subs': 'Annual Subscription',
        }
        return aliases.get(text, str(value or '').strip())

    def _normalise_name(self, value):
        return re.sub(r'[^a-z0-9]+', ' ', str(value or '').lower()).strip()

    def _token_key(self, value):
        return ' '.join(sorted(self._normalise_name(value).split()))

    def _format_amount(self, value):
        decimal_value = Decimal(value)
        if decimal_value == decimal_value.to_integral_value():
            return str(int(decimal_value))
        return format(decimal_value, 'f')

    def _write_csv(self, path, fieldnames, rows):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open('w', newline='', encoding='utf-8') as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
