from datetime import date, time, timedelta
from decimal import Decimal
from io import BytesIO
from io import StringIO
from unittest.mock import patch
import uuid

from django.core.management import call_command
from django.test import TestCase
from django.db import connection
from django.http import HttpResponse, QueryDict
from django.urls import reverse
from django.utils import timezone
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook

from deposits.models import (
    DepositFineAllocation,
    DepositSubmission,
    DepositWelfareAllocation,
)
from deposits.forms import DepositSubmissionForm, DirectDepositForm
from fines.models import Fine
from groupcore.account_context import SESSION_KEY_ACTIVE_ACCOUNT
from groupcore.models import (
    FinancialRecordRevision,
    GroupSettings,
    MemberProfile,
    SavingsAccount,
)
from groupcore.week_cycle import current_saving_week, saving_year_closing_date
from loans.models import LoanRequest
from loans.models import LoanRepayment
from deposits.welfare_calendar import build_welfare_calendar, welfare_totals_by_week


class WelfareAllocationBackfillTests(TestCase):
    def setUp(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        self.member = MemberProfile.objects.create_user(
            username='legacy-welfare', password='pass12345'
        )
        self.account = SavingsAccount.objects.create(
            owner=self.member, label='Main'
        )

    def _deposit(self, week, amount):
        return DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.member,
            payment_week=week,
            payment_date=week,
            payment_time=time(9, 0),
            welfare_amount=Decimal(amount),
            status='APPROVED',
        )

    def test_dry_run_does_not_write_and_commit_preserves_recorded_weeks(self):
        first = self._deposit(date(2026, 1, 2), '1000')
        later = self._deposit(date(2026, 1, 16), '1000')
        bulk = self._deposit(date(2026, 1, 30), '3000')
        output = StringIO()

        call_command('backfill_welfare_allocations', stdout=output)

        self.assertEqual(DepositWelfareAllocation.objects.count(), 0)
        self.assertIn('Would create 5 welfare-week allocation(s)', output.getvalue())

        call_command(
            'backfill_welfare_allocations', '--commit', stdout=StringIO()
        )

        self.assertEqual(
            list(first.welfare_allocations.values_list('welfare_week', flat=True)),
            [date(2026, 1, 2)],
        )
        self.assertEqual(
            list(later.welfare_allocations.values_list('welfare_week', flat=True)),
            [date(2026, 1, 16)],
        )
        self.assertEqual(
            set(bulk.welfare_allocations.values_list('welfare_week', flat=True)),
            {date(2026, 1, 9), date(2026, 1, 23), date(2026, 1, 30)},
        )
        self.assertEqual(
            sum(
                DepositWelfareAllocation.objects.values_list(
                    'amount', flat=True
                ),
                Decimal('0.00'),
            ),
            Decimal('5000.00'),
        )

    def test_commit_is_idempotent_and_invalid_amount_is_reported(self):
        self._deposit(date(2026, 1, 2), '1000')
        invalid = self._deposit(date(2026, 1, 9), '1500')
        call_command(
            'backfill_welfare_allocations', '--commit', stdout=StringIO()
        )
        output = StringIO()

        call_command(
            'backfill_welfare_allocations', '--commit', stdout=output
        )

        self.assertEqual(DepositWelfareAllocation.objects.count(), 1)
        self.assertFalse(invalid.welfare_allocations.exists())
        self.assertIn('Invalid deposits: 1', output.getvalue())

    def test_unique_account_is_inferred_and_excess_carries_to_next_year(self):
        deposit = DepositSubmission.objects.create(
            member=self.member,
            account=None,
            submitted_by=self.member,
            payment_week=date(2026, 7, 13),
            payment_date=date(2026, 7, 13),
            payment_time=time(9, 0),
            welfare_amount=Decimal('51000'),
            status='APPROVED',
        )
        output = StringIO()

        call_command(
            'backfill_welfare_allocations', '--commit', stdout=output
        )

        allocations = deposit.welfare_allocations.order_by('welfare_week')
        self.assertEqual(allocations.count(), 51)
        self.assertTrue(
            allocations.filter(
                account=self.account,
                welfare_week=date(2026, 7, 17),
            ).exists()
        )
        self.assertEqual(
            allocations.filter(welfare_week__year=2027).count(), 1
        )
        self.assertIn('Inferred unique accounts: 1', output.getvalue())
        self.assertIn('Normalized legacy weeks to Friday: 1', output.getvalue())
        self.assertIn('Carried into later saving years: 1', output.getvalue())


class VariableWeeklySavingsAllocationTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        first_day = date(today.year, 1, 1)
        first_friday = first_day + timedelta(days=(4 - first_day.weekday()) % 7)
        GroupSettings.objects.create(week_one_start=first_friday)
        self.saving_week = current_saving_week(first_friday, today)
        self.treasurer = MemberProfile.objects.create_user(
            username='variable-treasurer', password='pass12345', role='TREASURER'
        )
        self.member = MemberProfile.objects.create_user(
            username='variable-member', password='pass12345', role='MEMBER'
        )
        self.account = SavingsAccount.objects.create(owner=self.member, label='A')

    def direct_data(self, weeks, amounts, amount_received=None):
        data = QueryDict('', mutable=True)
        data.update({
            'member': str(self.member.id),
            'account': str(self.account.id),
            'payment_date': timezone.localdate().isoformat(),
            'payment_time': '10:00',
            'saving_amount': str(sum(amounts)),
        })
        data.setlist('selected_purposes', ['saving'])
        data.setlist('selected_weeks', [week.isoformat() for week in weeks])
        for week, amount in zip(weeks, amounts):
            data[f'week_amount_{week.isoformat()}'] = str(amount)
        if amount_received is not None:
            data['amount_received'] = str(amount_received)
        return data

    def test_each_boundary_and_midrange_weekly_amount_is_valid(self):
        week = self.saving_week.cycle_start
        for amount in (10000, 30000, 50000):
            with self.subTest(amount=amount):
                form = DirectDepositForm(self.direct_data([week], [amount], amount))
                self.assertTrue(form.is_valid(), form.errors)
                self.assertEqual(form.cleaned_data['weekly_allocations'], [(week, Decimal(amount))])

    def test_below_minimum_and_above_maximum_are_rejected(self):
        week = self.saving_week.cycle_start
        for amount in (9999, 50001):
            with self.subTest(amount=amount):
                form = DirectDepositForm(self.direct_data([week], [amount], amount))
                self.assertFalse(form.is_valid())
                self.assertIn('selected_weeks', form.errors)

    def test_treasurer_can_allocate_different_amounts_to_multiple_weeks(self):
        weeks = [self.saving_week.cycle_start, self.saving_week.cycle_start + timedelta(weeks=1)]
        data = self.direct_data(weeks, [20000, 10000], 35000)
        data['welfare_amount'] = '1000'
        post_data = data.dict()
        post_data['selected_purposes'] = ['saving', 'welfare']
        post_data['selected_weeks'] = [week.isoformat() for week in weeks]
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.post(reverse('manage_deposits'), post_data)

        self.assertEqual(response.status_code, 302, response.context['form'].errors if response.context else '')
        self.assertRedirects(response, reverse('manage_deposits'))
        saved = list(DepositSubmission.objects.filter(member=self.member).order_by('payment_week'))
        self.assertEqual([item.saving_amount for item in saved], [Decimal('20000'), Decimal('10000')])
        self.assertEqual([item.welfare_amount for item in saved], [Decimal('1000'), Decimal('0')])
        self.assertEqual(sum((item.amount for item in saved), Decimal('0')), Decimal('31000'))
        allocation = DepositWelfareAllocation.objects.get(deposit=saved[0])
        self.assertEqual(allocation.amount, Decimal('1000'))

    def test_welfare_calendar_tracks_pending_and_paid_weeks(self):
        week = self.saving_week.week_start
        deposit = DepositSubmission.objects.create(
            member=self.member, account=self.account, submitted_by=self.member,
            payment_week=week, welfare_amount=Decimal('1000'),
            payment_date=timezone.localdate(), payment_time=time(9, 0), status='PENDING',
        )
        DepositWelfareAllocation.objects.create(
            deposit=deposit, account=self.account, welfare_week=week,
        )
        pending = build_welfare_calendar(self.member, self.account)
        card = next(item for item in pending['weeks'] if item['friday'] == week)
        self.assertEqual(card['status'], 'pending')
        deposit.status = 'APPROVED'
        deposit.save(update_fields=['status'])
        paid = build_welfare_calendar(self.member, self.account)
        card = next(item for item in paid['weeks'] if item['friday'] == week)
        self.assertEqual(card['status'], 'paid')

    def test_direct_repayment_targets_selected_loan_and_reduces_balance(self):
        loan = LoanRequest.objects.create(
            member=self.member, account=self.account, principal=Decimal('100000'),
            monthly_interest_rate=Decimal('2'), duration_months=6, status='APPROVED',
            approved_on=timezone.now(),
        )
        before = loan.outstanding_balance
        data = QueryDict('', mutable=True)
        data.update({
            'member': str(self.member.id), 'account': str(self.account.id),
            'payment_date': timezone.localdate().isoformat(), 'payment_time': '10:00',
            'loan_repayment_loan': str(loan.id), 'loan_repayment_amount': '20000',
        })
        data.setlist('selected_purposes', ['loan_repayment'])
        self.client.login(username=self.treasurer.username, password='pass12345')
        response = self.client.post(reverse('manage_deposits'), data)
        self.assertEqual(response.status_code, 302)
        repayment = LoanRepayment.objects.get(loan=loan)
        self.assertIsNotNone(repayment.source_deposit_id)
        loan.refresh_from_db()
        self.assertEqual(loan.outstanding_balance, before - Decimal('20000'))

    def test_paid_week_is_locked_and_duplicate_selection_is_rejected(self):
        week = self.saving_week.cycle_start
        DepositSubmission.objects.create(
            member=self.member, account=self.account, submitted_by=self.treasurer,
            payment_week=week, saving_amount=Decimal('10000'),
            payment_date=week, payment_time=time(9, 0), status='APPROVED',
        )
        locked_form = DirectDepositForm(self.direct_data([week], [30000], 30000))
        self.assertFalse(locked_form.is_valid())
        self.assertContainsError(locked_form, 'already paid and locked')

        duplicate_data = self.direct_data([week + timedelta(weeks=1)] * 2, [10000, 10000], 20000)
        duplicate_form = DirectDepositForm(duplicate_data)
        self.assertFalse(duplicate_form.is_valid())
        self.assertContainsError(duplicate_form, 'Select each week only once')

    def assertContainsError(self, form, text):
        self.assertIn(text, str(form.errors))

    def test_member_form_enforces_same_range_and_pending_lock(self):
        week = self.saving_week.week_start
        base = QueryDict('', mutable=True)
        base.update({
            'account': str(self.account.id), 'payment_date': timezone.localdate().isoformat(),
            'payment_time': '10:00', 'saving_amount': '10000',
        })
        base.setlist('selected_purposes', ['saving'])
        valid = DepositSubmissionForm(base, user=self.member, payment_week=week)
        self.assertTrue(valid.is_valid(), valid.errors)

        too_low = base.copy()
        too_low['saving_amount'] = '9999'
        self.assertFalse(DepositSubmissionForm(too_low, user=self.member, payment_week=week).is_valid())

        DepositSubmission.objects.create(
            member=self.member, account=self.account, submitted_by=self.member,
            payment_week=week, saving_amount=Decimal('10000'),
            payment_date=week, payment_time=time(9, 0), status='PENDING',
        )
        locked = DepositSubmissionForm(base, user=self.member, payment_week=week)
        self.assertFalse(locked.is_valid())
        self.assertContainsError(locked, 'already paid or awaiting approval')

    def test_member_cannot_backdate_a_past_week(self):
        week = self.saving_week.cycle_start
        data = QueryDict('', mutable=True)
        data.update({
            'account': str(self.account.id),
            'payment_date': (timezone.localdate() - timedelta(days=1)).isoformat(),
            'payment_time': '10:00',
            'saving_amount': '10000',
        })
        data.setlist('selected_purposes', ['saving'])
        data.setlist('selected_weeks', [week.isoformat()])
        data[f'week_amount_{week.isoformat()}'] = '10000'

        form = DepositSubmissionForm(
            data,
            user=self.member,
            payment_week=week,
            allow_backdated_payment=False,
        )
        self.assertFalse(form.is_valid())
        self.assertIn('cannot backdate', str(form.errors))

    def test_manage_page_and_status_api_expose_week_allocations_and_locking(self):
        week = self.saving_week.cycle_start
        DepositSubmission.objects.create(
            member=self.member, account=self.account, submitted_by=self.treasurer,
            payment_week=week, saving_amount=Decimal('10000'),
            payment_date=week, payment_time=time(9, 0), status='APPROVED',
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        page = self.client.get(reverse('manage_deposits'))
        api = self.client.get(reverse('treasurer_week_options'), {
            'member': self.member.id, 'account': self.account.id,
        })

        self.assertEqual(page.status_code, 200)
        self.assertContains(page, 'UGX 10,000–50,000')
        self.assertContains(page, 'week_amount_')
        self.assertEqual(api.status_code, 200)
        week_status = next(item for item in api.json()['weeks'] if item['date'] == week.isoformat())
        self.assertTrue(week_status['paid'])

    def test_manage_page_includes_approved_and_rejected_submissions(self):
        week = self.saving_week.cycle_start
        DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.treasurer,
            payment_week=week,
            saving_amount=Decimal('10000'),
            payment_date=week,
            payment_time=time(9, 0),
            status='APPROVED',
        )
        rejected = DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.member,
            payment_week=week + timedelta(weeks=1),
            saving_amount=Decimal('10000'),
            payment_date=week + timedelta(weeks=1),
            payment_time=time(9, 0),
            status='REJECTED',
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(reverse('manage_deposits'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Deposit Submissions')
        self.assertContains(response, 'Approved')
        self.assertContains(response, 'Rejected')
        self.assertIn(rejected, response.context['deposit_submissions'])

    def test_manage_page_orders_by_latest_review_then_latest_submission(self):
        week = self.saving_week.cycle_start
        latest_review = DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.treasurer,
            payment_week=week,
            saving_amount=Decimal('10000'),
            payment_date=week,
            payment_time=time(9, 0),
            status='APPROVED',
        )
        older_review = DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.treasurer,
            payment_week=week + timedelta(weeks=1),
            saving_amount=Decimal('10000'),
            payment_date=week + timedelta(weeks=1),
            payment_time=time(9, 0),
            status='APPROVED',
        )
        newer_pending = DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.member,
            payment_week=week + timedelta(weeks=2),
            saving_amount=Decimal('10000'),
            payment_date=week + timedelta(weeks=2),
            payment_time=time(9, 0),
            status='PENDING',
        )
        older_pending = DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.member,
            payment_week=week + timedelta(weeks=3),
            saving_amount=Decimal('10000'),
            payment_date=week + timedelta(weeks=3),
            payment_time=time(9, 0),
            status='PENDING',
        )
        now = timezone.now()
        DepositSubmission.objects.filter(pk=latest_review.pk).update(
            date_reviewed=now,
            date_submitted=now - timedelta(days=4),
        )
        DepositSubmission.objects.filter(pk=older_review.pk).update(
            date_reviewed=now - timedelta(days=1),
            date_submitted=now - timedelta(days=3),
        )
        DepositSubmission.objects.filter(pk=newer_pending.pk).update(
            date_submitted=now - timedelta(hours=1),
        )
        DepositSubmission.objects.filter(pk=older_pending.pk).update(
            date_submitted=now - timedelta(days=2),
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(reverse('manage_deposits'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [deposit.pk for deposit in response.context['deposit_submissions']],
            [
                latest_review.pk,
                older_review.pk,
                newer_pending.pk,
                older_pending.pk,
            ],
        )

    def test_manage_page_status_filters_show_counts_and_only_selected_records(self):
        week = self.saving_week.cycle_start
        records = {}
        for index, status in enumerate(('PENDING', 'APPROVED', 'REJECTED')):
            records[status] = DepositSubmission.objects.create(
                member=self.member,
                account=self.account,
                submitted_by=self.treasurer,
                payment_week=week + timedelta(weeks=index),
                saving_amount=Decimal('10000'),
                payment_date=week + timedelta(weeks=index),
                payment_time=time(9, 0),
                status=status,
                remarks=f'{status.lower()}-record',
            )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(
            reverse('manage_deposits'),
            {'status': 'PENDING'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['status_filter'], 'PENDING')
        self.assertEqual(
            response.context['status_counts'],
            {'all': 3, 'pending': 1, 'approved': 1, 'rejected': 1},
        )
        displayed = list(response.context['deposit_submissions'])
        self.assertEqual(displayed, [records['PENDING']])
        self.assertContains(response, 'Pending')
        self.assertContains(response, '?status=APPROVED')
        self.assertContains(response, '?status=REJECTED')

    def test_manage_page_search_and_pagination_preserve_status_filter(self):
        week = self.saving_week.cycle_start
        DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.treasurer,
            payment_week=week,
            saving_amount=Decimal('10000'),
            payment_date=week,
            payment_time=time(9, 0),
            status='PENDING',
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(
            reverse('manage_deposits'),
            {'status': 'PENDING', 'q': self.member.username},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['pagination_query'], 'status=PENDING&q=variable-member')
        self.assertContains(response, 'type="hidden" name="status" value="PENDING"')
        self.assertContains(response, 'status=APPROVED&amp;q=variable-member')

    def test_manage_page_uses_compact_picker_after_purpose_amounts(self):
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(reverse('manage_deposits'))

        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        purpose_position = html.index('Deposit Purpose')
        savings_amount_position = html.index('Savings total (UGX)')
        picker_position = html.index('id="ddSavingsAllocation"')
        self.assertLess(purpose_position, savings_amount_position)
        self.assertLess(savings_amount_position, picker_position)
        self.assertContains(response, 'id="ddWeekPickerModal"')
        self.assertContains(response, 'modal-fullscreen-sm-down')
        self.assertContains(response, 'data-week-option')
        self.assertContains(response, 'memberLiveSearch')
        self.assertContains(response, 'Search this table')
        self.assertContains(response, 'ddWelfareSelectedCount')
        self.assertContains(response, 'ddWelfareModalTotal')
        self.assertContains(response, 'Paid and pending weeks are locked')
        self.assertNotContains(response, 'class="week-card')

    def test_legacy_amount_received_is_ignored(self):
        week = self.saving_week.cycle_start
        form = DirectDepositForm(self.direct_data([week], [10000], 30000))
        self.assertTrue(form.is_valid(), form.errors)
        self.assertNotIn('amount_received', form.fields)

    def test_savings_total_is_authoritative_and_must_match_week_allocations(self):
        weeks = [self.saving_week.cycle_start, self.saving_week.cycle_start + timedelta(weeks=1)]
        data = self.direct_data(weeks, [10000, 20000], 40000)
        data['saving_amount'] = '40000'

        form = DirectDepositForm(data)

        self.assertFalse(form.is_valid())
        self.assertContainsError(form, 'Weekly allocations must equal the Savings total')
        self.assertContainsError(form, '10,000 is remaining')
        selected_options = [option for option in form.week_options if option['selected']]
        self.assertEqual([option['value'] for option in selected_options], [
            week.isoformat() for week in weeks
        ])
        self.assertEqual([option['amount'] for option in selected_options], ['10000', '20000'])

    def test_future_weeks_are_selectable_but_out_of_year_weeks_are_rejected(self):
        controlled_today = self.saving_week.cycle_start + timedelta(weeks=4, days=2)
        future_week = self.saving_week.cycle_start + timedelta(weeks=5)
        with patch('deposits.forms.timezone.localdate', return_value=controlled_today):
            future_form = DirectDepositForm(self.direct_data([future_week], [10000], 10000))

        self.assertTrue(future_form.is_valid(), future_form.errors)

        outside_week = self.saving_week.cycle_start - timedelta(weeks=1)
        outside_form = DirectDepositForm(self.direct_data([outside_week], [10000], 10000))
        self.assertFalse(outside_form.is_valid())
        self.assertContainsError(outside_form, 'active saving year only')

    def test_non_savings_deposit_ignores_calendar_post_and_uses_payment_date_week(self):
        payment_date = self.saving_week.cycle_start + timedelta(weeks=2, days=3)
        expected_week = self.saving_week.cycle_start + timedelta(weeks=2)
        data = QueryDict('', mutable=True)
        data.update({
            'member': str(self.member.id),
            'account': str(self.account.id),
            'payment_date': payment_date.isoformat(),
            'payment_time': '10:00',
            'payment_week': 'not-a-date',
            'welfare_amount': '15000',
            'amount_received': '15000',
        })
        data.setlist('selected_purposes', ['welfare'])
        data.setlist('selected_weeks', ['1900-01-01', '2999-12-31'])
        data['week_amount_1900-01-01'] = '50000'

        form = DirectDepositForm(data)

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['selected_week_dates'], [expected_week])
        self.assertEqual(form.cleaned_data['payment_week'], expected_week)
        self.assertEqual(form.cleaned_data['weekly_allocations'], [])
        self.assertEqual(form.cleaned_data['saving_amount'], Decimal('0.00'))

        self.client.login(username=self.treasurer.username, password='pass12345')
        response = self.client.post(reverse('manage_deposits'), data)
        self.assertRedirects(response, reverse('manage_deposits'))
        saved = DepositSubmission.objects.get(member=self.member)
        self.assertEqual(saved.payment_week, expected_week)
        self.assertEqual(saved.saving_amount, Decimal('0.00'))
        self.assertEqual(saved.welfare_amount, Decimal('1000.00'))

    def test_full_saving_year_metadata_and_status_api_use_bounded_queries(self):
        paid_week = self.saving_week.cycle_start
        DepositSubmission.objects.create(
            member=self.member, account=self.account, submitted_by=self.treasurer,
            payment_week=paid_week, saving_amount=Decimal('10000'),
            payment_date=paid_week, payment_time=time(9, 0), status='APPROVED',
        )

        form = DirectDepositForm(initial={'member': self.member, 'account': self.account})
        self.assertEqual(
            date.fromisoformat(form.week_options[-1]['value']),
            saving_year_closing_date(self.saving_week.saving_year),
        )
        first = form.week_options[0]
        self.assertEqual(first['value'], paid_week.isoformat())
        self.assertEqual(first['week_number'], 1)
        self.assertEqual(first['month'], paid_week.month)
        self.assertEqual(first['day'], paid_week.day)
        self.assertIn('accessible_label', first)
        self.assertTrue(first['is_paid'])
        self.assertFalse(first['selectable'])

        self.client.login(username=self.treasurer.username, password='pass12345')
        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse('treasurer_week_options'), {
                'member': self.member.id,
                'account': self.account.id,
            })

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries), 12)
        weeks = response.json()['weeks']
        self.assertEqual(len(weeks), len(form.week_options))
        api_week = weeks[0]
        self.assertEqual(api_week['date'], paid_week.isoformat())
        self.assertEqual(api_week['paid_amount'], '10000')
        self.assertTrue(api_week['paid'])
        self.assertFalse(api_week['selectable'])
        self.assertIn('date_label', api_week)


class AuditedDepositDeletionTests(TestCase):
    def setUp(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        self.treasurer = MemberProfile.objects.create_user(
            username='delete-treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.member = MemberProfile.objects.create_user(
            username='delete-member',
            password='pass12345',
            role='MEMBER',
        )
        self.account = SavingsAccount.objects.create(
            owner=self.member,
            label='Main account',
        )
        self.week = date(2026, 7, 24)

    def _deposit(self, **overrides):
        values = {
            'member': self.member,
            'account': self.account,
            'submitted_by': self.treasurer,
            'payment_week': self.week,
            'payment_date': self.week,
            'payment_time': time(9, 0),
            'saving_amount': Decimal('10000'),
            'status': 'APPROVED',
        }
        values.update(overrides)
        return DepositSubmission.objects.create(**values)

    def test_treasurer_hard_delete_reverses_linked_effects_and_records_revision(self):
        loan = LoanRequest.objects.create(
            member=self.member,
            account=self.account,
            principal=Decimal('100000'),
            monthly_interest_rate=Decimal('2'),
            duration_months=6,
            status='APPROVED',
            approved_on=timezone.now(),
        )
        fine = Fine.objects.create(
            member=self.member,
            account=self.account,
            reason='Test fine',
            amount=Decimal('2000'),
            amount_paid=Decimal('1000'),
            issued_by=self.treasurer,
        )
        deposit = self._deposit(
            saving_amount=Decimal('0'),
            welfare_amount=Decimal('1000'),
            fine_amount=Decimal('1000'),
            loan_repayment_amount=Decimal('10000'),
            loan_repayment_loan=loan,
        )
        DepositWelfareAllocation.objects.create(
            deposit=deposit,
            account=self.account,
            welfare_week=self.week,
        )
        DepositFineAllocation.objects.create(
            deposit=deposit,
            fine=fine,
            amount=Decimal('1000'),
        )
        repayment = LoanRepayment.objects.create(
            loan=loan,
            amount=Decimal('10000'),
            paid_on=self.week,
            recorded_by=self.treasurer,
            source_deposit=deposit,
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': 'Duplicate imported payment'},
        )

        self.assertRedirects(response, reverse('manage_deposits'))
        fine.refresh_from_db()
        self.assertFalse(DepositSubmission.objects.filter(pk=deposit.pk).exists())
        self.assertEqual(fine.amount_paid, Decimal('0'))
        self.assertFalse(fine.is_paid)
        self.assertFalse(LoanRepayment.objects.filter(pk=repayment.pk).exists())
        self.assertFalse(
            DepositWelfareAllocation.objects.filter(deposit_id=deposit.pk).exists()
        )
        self.assertNotIn(
            self.week,
            welfare_totals_by_week(
                self.member,
                self.account,
                [self.week],
            ),
        )
        revision = FinancialRecordRevision.objects.get(
            record_type='deposit',
            object_id=deposit.id,
        )
        self.assertEqual(revision.reason, 'Duplicate imported payment')
        self.assertEqual(revision.before_data['status'], 'APPROVED')
        self.assertEqual(revision.after_data['record_state'], 'Deleted')

    def test_reason_is_required_and_member_cannot_delete(self):
        deposit = self._deposit(status='PENDING')
        self.client.login(username=self.treasurer.username, password='pass12345')
        response = self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': ''},
        )
        self.assertRedirects(response, reverse('manage_deposits'))
        deposit.refresh_from_db()
        self.assertEqual(deposit.status, 'PENDING')
        self.assertFalse(FinancialRecordRevision.objects.exists())

        self.client.login(username=self.member.username, password='pass12345')
        response = self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': 'Member trying deletion'},
        )
        self.assertRedirects(response, reverse('member_dashboard'))
        deposit.refresh_from_db()
        self.assertEqual(deposit.status, 'PENDING')

    def test_legacy_unlinked_fine_payment_is_not_deleted_ambiguously(self):
        deposit = self._deposit(
            saving_amount=Decimal('0'),
            fine_amount=Decimal('1000'),
        )
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': 'Incorrect legacy entry'},
            follow=True,
        )

        deposit.refresh_from_db()
        self.assertEqual(deposit.status, 'APPROVED')
        self.assertContains(response, 'not linked to a specific fine')
        self.assertFalse(FinancialRecordRevision.objects.exists())

    def test_manage_page_has_delete_form_audit_buttons_and_deletion_audit(self):
        deposit = self._deposit(status='PENDING')
        self.client.login(username=self.treasurer.username, password='pass12345')

        response = self.client.get(reverse('manage_deposits'))
        self.assertContains(response, 'Reason for deletion')
        self.assertContains(response, 'Deletion Audit')
        self.assertContains(response, 'Audit')
        self.assertContains(
            response,
            f'data-delete-url="{reverse("delete_deposit", args=[deposit.id])}"',
        )

        self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': 'Previously removed'},
        )
        audit = self.client.get(reverse('deposit_deletion_audit'))
        self.assertEqual(audit.status_code, 200)
        self.assertContains(audit, f'Deposit #{deposit.id}')
        self.assertContains(audit, 'Previously removed')
        self.assertContains(audit, self.treasurer.username)

    def test_deleted_deposit_is_read_only_but_audit_remains_visible(self):
        deposit = self._deposit(status='PENDING')
        self.client.login(username=self.treasurer.username, password='pass12345')
        self.client.post(
            reverse('delete_deposit', args=[deposit.id]),
            {'deletion_reason': 'Entered for wrong member'},
        )

        audit_response = self.client.get(
            reverse('financial_record_history', args=['deposit', deposit.id]),
        )

        self.assertEqual(audit_response.status_code, 200)
        self.assertContains(audit_response, 'Entered for wrong member')
        self.assertContains(audit_response, 'Deleted')
        self.assertContains(audit_response, 'deleted from the database')
        self.assertNotContains(audit_response, 'Edit record')


class TreasurerReportYearFilterTests(TestCase):
    def setUp(self):
        self.current_year = timezone.localdate().year
        self.previous_year = self.current_year - 1
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.member = MemberProfile.objects.create_user(
            username='member',
            password='pass12345',
            role='MEMBER',
        )
        self.previous_deposit = self._deposit(
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.current_year, 1, 2),
            saving_amount=Decimal('10000.00'),
        )
        self.current_deposit = self._deposit(
            payment_week=date(self.current_year, 1, 5),
            payment_date=date(self.current_year, 1, 5),
            saving_amount=Decimal('50000.00'),
        )

    def _deposit(self, payment_week, payment_date, saving_amount):
        return DepositSubmission.objects.create(
            member=self.member,
            submitted_by=self.treasurer,
            payment_week=payment_week,
            starting_week=payment_week,
            weeks_covered=1,
            saving_amount=saving_amount,
            proof='proofs/test.jpg',
            payment_date=payment_date,
            payment_time=time(9, 0),
            status='APPROVED',
        )

    def _member_row(self, response):
        return next(
            row for row in response.context['report_data']
            if row['member'].id == self.member.id
        )

    def test_treasurer_report_filters_totals_by_selected_payment_week_year(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_reports'), {'year': self.previous_year})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_year'], self.previous_year)
        row = self._member_row(response)
        self.assertEqual(row['total_saving'], Decimal('10000'))
        self.assertEqual(row['total_amount'], Decimal('10000'))
        self.assertEqual(row['total_weeks'], 1)

    def test_treasurer_report_includes_savings_account_column(self):
        account = SavingsAccount.objects.create(owner=self.member, label='A2')
        self.previous_deposit.account = account
        self.previous_deposit.save(update_fields=['account'])
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_reports'), {'year': self.previous_year})

        self.assertContains(response, 'Savings Account')
        self.assertEqual(self._member_row(response)['savings_account'], 'A2')

    def test_treasurer_report_defaults_to_current_year(self):
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_reports'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_year'], self.current_year)
        row = self._member_row(response)
        self.assertEqual(row['total_saving'], Decimal('50000'))
        self.assertEqual(row['total_amount'], Decimal('50000'))
        self.assertEqual(row['total_weeks'], 1)

    def test_treasurer_report_ignores_approved_loans_without_approval_date_in_year_options(self):
        LoanRequest.objects.create(
            member=self.member,
            principal=Decimal('100000.00'),
            monthly_interest_rate=Decimal('2.00'),
            status=LoanRequest.STATUS_APPROVED,
            approved_on=None,
        )
        self.client.login(username='treasurer', password='pass12345')

        response = self.client.get(reverse('treasurer_reports'))

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.current_year, response.context['years'])

    def test_member_export_filenames_include_selected_year(self):
        pdf_response = self.client.get(
            reverse('download_member_report', args=[self.member.id, 'pdf']),
            {'year': self.previous_year},
        )
        excel_response = self.client.get(
            reverse('download_member_report', args=[self.member.id, 'excel']),
            {'year': self.previous_year},
        )

        self.assertIn(
            f'{self.member.username}_report_{self.previous_year}.pdf',
            pdf_response['Content-Disposition'],
        )
        self.assertIn(
            f'{self.member.username}_report_{self.previous_year}.xlsx',
            excel_response['Content-Disposition'],
        )

    def test_all_member_excel_export_uses_selected_payment_week_year(self):
        response = self.client.get(
            reverse('download_all_reports', args=['excel']),
            {'year': self.previous_year},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f'all_member_reports_{self.previous_year}.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        values = [
            value
            for row in workbook.active.iter_rows(values_only=True)
            for value in row
        ]
        self.assertIn(self.previous_deposit.payment_week.strftime('%Y-%m-%d'), values)
        self.assertNotIn(self.current_deposit.payment_week.strftime('%Y-%m-%d'), values)


class MyContributionsAccountExportTests(TestCase):
    def setUp(self):
        self.current_year = timezone.localdate().year
        self.previous_year = self.current_year - 1
        self.member = MemberProfile.objects.create_user(
            username='member_account_owner',
            password='pass12345',
            role='MEMBER',
        )
        self.other_member = MemberProfile.objects.create_user(
            username='other_member',
            password='pass12345',
            role='MEMBER',
        )
        self.account_a1 = SavingsAccount.objects.create(owner=self.member, label='A1')
        self.account_a2 = SavingsAccount.objects.create(owner=self.member, label='A2')
        self.other_account = SavingsAccount.objects.create(owner=self.other_member, label='A2')

        self.a2_previous_approved = self._deposit(
            member=self.member,
            account=self.account_a2,
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.current_year, 1, 2),
            saving_amount=Decimal('100.00'),
            status='APPROVED',
        )
        self.a2_previous_pending = self._deposit(
            member=self.member,
            account=self.account_a2,
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.previous_year, 12, 30),
            saving_amount=Decimal('200.00'),
            status='PENDING',
        )
        self.a2_previous_rejected = self._deposit(
            member=self.member,
            account=self.account_a2,
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.previous_year, 12, 31),
            saving_amount=Decimal('300.00'),
            status='REJECTED',
        )
        self.a2_current = self._deposit(
            member=self.member,
            account=self.account_a2,
            payment_week=date(self.current_year, 1, 5),
            payment_date=date(self.current_year, 1, 5),
            saving_amount=Decimal('400.00'),
            status='APPROVED',
        )
        self.a1_previous = self._deposit(
            member=self.member,
            account=self.account_a1,
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.previous_year, 12, 29),
            saving_amount=Decimal('500.00'),
            status='APPROVED',
        )
        self.other_previous = self._deposit(
            member=self.other_member,
            account=self.other_account,
            payment_week=date(self.previous_year, 12, 29),
            payment_date=date(self.previous_year, 12, 29),
            saving_amount=Decimal('600.00'),
            status='APPROVED',
        )

    def _deposit(self, member, account, payment_week, payment_date, saving_amount, status):
        return DepositSubmission.objects.create(
            member=member,
            account=account,
            submitted_by=member,
            payment_week=payment_week,
            starting_week=payment_week,
            weeks_covered=1,
            saving_amount=saving_amount,
            proof='proofs/test.jpg',
            payment_date=payment_date,
            payment_time=time(9, 0),
            status=status,
        )

    def _login_with_active_account(self, account):
        self.client.login(username='member_account_owner', password='pass12345')
        session = self.client.session
        session[SESSION_KEY_ACTIVE_ACCOUNT] = account.id
        session.save()

    def test_my_contributions_page_filters_active_account_by_payment_week_year(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(reverse('my_contributions'), {'year': self.previous_year})

        self.assertEqual(response.status_code, 200)
        deposits = list(response.context['deposits'])
        self.assertIn(self.a2_previous_approved, deposits)
        self.assertIn(self.a2_previous_pending, deposits)
        self.assertIn(self.a2_previous_rejected, deposits)
        self.assertNotIn(self.a2_current, deposits)
        self.assertNotIn(self.a1_previous, deposits)
        self.assertNotIn(self.other_previous, deposits)
        self.assertEqual(response.context['approved_totals']['total'], Decimal('100'))

    def test_my_contributions_defaults_to_current_year(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(reverse('my_contributions'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_year'], self.current_year)
        self.assertIn(self.current_year, response.context['years'])
        deposits = list(response.context['deposits'])
        self.assertIn(self.a2_current, deposits)
        self.assertNotIn(self.a2_previous_approved, deposits)
        self.assertNotIn(self.a2_previous_pending, deposits)
        self.assertNotIn(self.a2_previous_rejected, deposits)

    def test_excel_export_matches_active_account_filter_and_approved_totals(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(
            reverse('export_my_contributions', args=['excel']),
            {'year': self.previous_year},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(
            f'my_contributions_A2_{self.previous_year}.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        values = [
            value
            for row in sheet.iter_rows(values_only=True)
            for value in row
        ]

        self.assertEqual(sheet['A12'].value, 100)
        self.assertIn('Pending', values)
        self.assertIn('Rejected', values)
        self.assertIn(self.a2_previous_approved.payment_week.strftime('%Y-%m-%d'), values)
        self.assertNotIn('A1', values)
        self.assertNotIn('other_member', values)
        self.assertNotIn(self.a2_current.payment_week.strftime('%Y-%m-%d'), values)

    def test_excel_export_defaults_to_current_year(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(reverse('export_my_contributions', args=['excel']))

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f'my_contributions_A2_{self.current_year}.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        values = [
            value
            for row in sheet.iter_rows(values_only=True)
            for value in row
        ]
        self.assertEqual(sheet['A12'].value, 400)
        self.assertIn(self.a2_current.payment_week.strftime('%Y-%m-%d'), values)
        self.assertNotIn(self.a2_previous_approved.payment_week.strftime('%Y-%m-%d'), values)

    def test_pdf_export_uses_active_account_filename(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(
            reverse('export_my_contributions', args=['pdf']),
            {'year': self.previous_year},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn(
            f'my_contributions_A2_{self.previous_year}.pdf',
            response['Content-Disposition'],
        )

    def test_invalid_year_export_defaults_to_current_year(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(
            reverse('export_my_contributions', args=['excel']),
            {'year': 'bad-year'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f'my_contributions_A2_{self.current_year}.xlsx',
            response['Content-Disposition'],
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.active['A12'].value, 400)

    def test_empty_excel_export_is_valid_with_zero_approved_totals_for_selected_year(self):
        self._login_with_active_account(self.account_a2)

        response = self.client.get(
            reverse('export_my_contributions', args=['excel']),
            {'year': 1900},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.active['A12'].value, 0)


class CurrentWeekStatusExportTests(TestCase):
    def setUp(self):
        self.today = date(2026, 7, 5)
        self.monday_after_grace = date(2026, 7, 6)
        self.week_one_start = date(2026, 1, 2)
        GroupSettings.objects.create(week_one_start=self.week_one_start)
        self.treasurer = MemberProfile.objects.create_user(
            username='treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.member = MemberProfile.objects.create_user(
            username='member',
            password='pass12345',
            role='MEMBER',
            first_name='Test',
            last_name='Member',
        )
        self.account = SavingsAccount.objects.create(owner=self.member, label='A1')
        self.unpaid_member = MemberProfile.objects.create_user(
            username='unpaid_member',
            password='pass12345',
            role='MEMBER',
            first_name='Late',
            last_name='Member',
        )
        self.unpaid_account = SavingsAccount.objects.create(owner=self.unpaid_member, label='B1')
        saving_week = current_saving_week(self.week_one_start, self.today)
        DepositSubmission.objects.create(
            member=self.member,
            account=self.account,
            submitted_by=self.member,
            payment_week=saving_week.week_start,
            starting_week=saving_week.week_start,
            weeks_covered=1,
            saving_amount=Decimal('50000.00'),
            proof='proofs/test.jpg',
            payment_date=saving_week.week_start,
            payment_time=time(9, 0),
            status='APPROVED',
        )

    def test_current_week_status_page_lists_account_level_statuses(self):
        self.client.login(username='treasurer', password='pass12345')

        with patch('deposits.views.timezone.localdate', return_value=self.today):
            response = self.client.get(reverse('current_week_status'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Paid Accounts')
        self.assertContains(response, 'Test Member')
        self.assertContains(response, 'A1')

    def test_current_week_status_does_not_create_fines_before_monday_after_grace(self):
        self.client.login(username='treasurer', password='pass12345')

        with patch('deposits.views.timezone.localdate', return_value=self.today):
            response = self.client.get(reverse('current_week_status'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Late Member')
        self.assertEqual(Fine.objects.count(), 0)

    def test_current_week_status_creates_fines_after_sunday_closes(self):
        self.client.login(username='treasurer', password='pass12345')

        with patch('deposits.views.timezone.localdate', return_value=self.monday_after_grace):
            response = self.client.get(reverse('current_week_status'))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            Fine.objects.filter(
                member=self.unpaid_member,
                account=self.unpaid_account,
                fine_type='MISSED_WEEKLY_SAVING',
            ).exists()
        )

    def test_current_week_status_export_excel_includes_account_status(self):
        self.client.login(username='treasurer', password='pass12345')

        with patch('deposits.views.timezone.localdate', return_value=self.today):
            response = self.client.get(reverse('export_current_week_status', args=['excel']))

        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        values = list(workbook.active.values)
        flattened = [cell for row in values for cell in row]
        self.assertIn('Test Member', flattened)
        self.assertIn('A1', flattened)
        self.assertIn('Paid', flattened)

    def test_current_week_status_export_pdf_returns_pdf(self):
        self.client.login(username='treasurer', password='pass12345')

        with patch('deposits.views.timezone.localdate', return_value=self.today):
            response = self.client.get(reverse('export_current_week_status', args=['pdf']))

        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response['Content-Disposition'].endswith('.pdf"'))

    def test_paid_and_unpaid_pdf_exports_are_separate(self):
        self.client.login(username='treasurer', password='pass12345')

        with (
            patch('deposits.views.timezone.localdate', return_value=self.today),
            patch(
                'deposits.views._export_current_week_status_pdf',
                side_effect=lambda data: HttpResponse(data['export_scope']),
            ) as exporter,
        ):
            paid_response = self.client.get(
                reverse('export_current_week_status', args=['pdf']),
                {'scope': 'paid'},
            )
            unpaid_response = self.client.get(
                reverse('export_current_week_status', args=['pdf']),
                {'scope': 'unpaid'},
            )

        self.assertEqual(paid_response.content, b'paid')
        self.assertEqual(unpaid_response.content, b'unpaid')
        self.assertEqual(
            [call.args[0]['export_scope'] for call in exporter.call_args_list],
            ['paid', 'unpaid'],
        )


class DepositBatchWorkflowTests(TestCase):
    def setUp(self):
        GroupSettings.objects.create(week_one_start=date(2026, 1, 2))
        self.treasurer = MemberProfile.objects.create_user(
            username='batch-treasurer',
            password='pass12345',
            role='TREASURER',
        )
        self.member = MemberProfile.objects.create_user(
            username='batch-member',
            password='pass12345',
        )
        self.account = SavingsAccount.objects.create(
            owner=self.member,
            label='Main',
        )
        self.batch = uuid.uuid4()
        self.weeks = [date(2026, 7, 3), date(2026, 7, 10)]
        for week, amount in zip(self.weeks, ('10000', '20000')):
            DepositSubmission.objects.create(
                submission_batch=self.batch,
                member=self.member,
                account=self.account,
                submitted_by=self.member,
                payment_week=week,
                saving_amount=Decimal(amount),
                payment_date=date(2026, 7, 10),
                payment_time=time(10, 0),
                status='PENDING',
            )
        self.client.login(username='batch-treasurer', password='pass12345')

    def _edit_payload(self):
        return {
            'member': str(self.member.id),
            'account': str(self.account.id),
            'payment_week': self.weeks[0].isoformat(),
            'payment_date': '2026-07-10',
            'payment_time': '10:00',
            'selected_purposes': ['saving'],
            'saving_amount': '30000',
            'selected_weeks': [week.isoformat() for week in self.weeks],
            f'week_amount_{self.weeks[0].isoformat()}': '10000',
            f'week_amount_{self.weeks[1].isoformat()}': '20000',
            'remarks': 'Corrected batch note',
            'edit_reason': 'Correcting the complete submission',
        }

    def test_manage_page_groups_records_and_editor_prefills_weeks(self):
        page = self.client.get(reverse('manage_deposits'))
        edit = self.client.get(
            reverse('edit_deposit_batch', args=[self.batch])
        )

        self.assertEqual(len(page.context['deposit_submissions']), 1)
        summary = page.context['deposit_submissions'][0]
        self.assertEqual(summary.amount, Decimal('30000'))
        self.assertEqual(summary.payment_weeks, self.weeks)
        self.assertContains(edit, 'Edit Deposit Submission')
        self.assertContains(edit, f'value="{self.weeks[0].isoformat()}" checked')
        self.assertContains(edit, 'value="10000.00"')

    def test_edit_and_approve_apply_to_complete_batch(self):
        response = self.client.post(
            reverse('edit_deposit_batch', args=[self.batch]),
            self._edit_payload(),
        )
        self.assertRedirects(
            response,
            reverse('deposit_batch_history', args=[self.batch]),
        )
        records = DepositSubmission.objects.filter(
            submission_batch=self.batch
        ).order_by('payment_week')
        self.assertEqual(records.count(), 2)
        self.assertEqual(records.first().remarks, 'Corrected batch note')
        self.assertEqual(
            FinancialRecordRevision.objects.filter(
                record_type='deposit',
                object_id__in=records.values('id'),
            ).count(),
            2,
        )

        approved = self.client.post(
            reverse('approve_deposit_batch', args=[self.batch])
        )
        self.assertRedirects(approved, reverse('manage_deposits'))
        self.assertFalse(records.exclude(status='APPROVED').exists())

    def test_reject_and_delete_apply_to_complete_batch(self):
        rejected = self.client.post(
            reverse('reject_deposit_batch', args=[self.batch])
        )
        self.assertRedirects(rejected, reverse('manage_deposits'))
        self.assertFalse(
            DepositSubmission.objects.filter(
                submission_batch=self.batch
            ).exclude(status='REJECTED').exists()
        )

        deleted = self.client.post(
            reverse('delete_deposit_batch', args=[self.batch]),
            {'deletion_reason': 'Duplicate complete submission'},
        )
        self.assertRedirects(deleted, reverse('manage_deposits'))
        self.assertFalse(
            DepositSubmission.objects.filter(
                submission_batch=self.batch
            ).exists()
        )

    def test_edit_status_is_context_aware(self):
        DepositSubmission.objects.filter(
            submission_batch=self.batch
        ).update(
            status='APPROVED',
            reviewed_by=self.treasurer,
            date_reviewed=timezone.now(),
        )
        approved_edit = self.client.post(
            reverse('edit_deposit_batch', args=[self.batch]),
            self._edit_payload(),
        )
        self.assertEqual(approved_edit.status_code, 302)
        self.assertFalse(
            DepositSubmission.objects.filter(
                submission_batch=self.batch
            ).exclude(status='APPROVED').exists()
        )

        DepositSubmission.objects.filter(
            submission_batch=self.batch
        ).update(status='REJECTED')
        rejected_edit = self.client.post(
            reverse('edit_deposit_batch', args=[self.batch]),
            self._edit_payload(),
        )
        self.assertEqual(rejected_edit.status_code, 302)
        self.assertFalse(
            DepositSubmission.objects.filter(
                submission_batch=self.batch
            ).exclude(status='PENDING').exists()
        )
