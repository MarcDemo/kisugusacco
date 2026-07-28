from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from .forms import DepositBatchEditForm, DepositSubmissionForm, DirectDepositForm
from .models import DepositFineAllocation, DepositSubmission, DepositWelfareAllocation
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from groupcore.models import FinancialRecordRevision, GroupSettings, MemberProfile
from groupcore.models import SavingsAccount
from groupcore.reporting import merge_year_options, pagination_query, parse_report_year, years_from_dates
from groupcore.week_cycle import current_saving_week
from groupcore.year_close import (
    apply_eligible_locked_repayment,
    financial_year,
    loan_activity_frozen,
    pending_at_cutoff,
    submissions_locked_for_year,
)
from django.utils import timezone
from django.db.models import Count, F, Sum, Q
from django.db import transaction
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime, date, timedelta
from xml.sax.saxutils import escape as xml_escape
import os
import uuid
from django.utils.timezone import now
from openpyxl.utils import get_column_letter
from django.core.mail import send_mail
from django.core.exceptions import ValidationError
from fines.models import Fine
from fines.services import (
    allocate_fine_payment,
    apply_selected_fine_allocations,
    delete_deposit_week_missed_saving_fines,
    missed_saving_fines_can_be_created,
)
from groupcore.account_context import get_active_account, get_user_active_accounts
from groupcore.savings_calendar import LATE_FINE_AMOUNT, completion_for_week, week_deadline
from loans.models import LoanRepayment, LoanRequest
from decimal import Decimal
from deposits.rules import (
    MAX_WEEKLY_SAVINGS,
    MIN_WEEKLY_SAVINGS,
    fine_week_options,
    saving_year_weeks,
    saving_week_statuses,
    weekly_savings_total,
    weekly_savings_totals_by_week,
)
from deposits.welfare_calendar import build_welfare_calendar
from groupcore.member_query import alphabetical_members
from groupcore.financial_records import delete_deposit_with_audit
from groupcore.financial_records import record_revision, snapshot_record

# Create your views here.


def _loan_interest_total(queryset):
    total = 0
    for loan in queryset:
        total += loan.total_interest
    return total


def _apply_loan_repayment(
    member, account, loan, repayment_amount, paid_on, recorded_by,
    source_deposit=None, notes='',
):
    if repayment_amount <= 0:
        return Decimal('0.00'), Decimal('0.00')
    if not loan or loan.member_id != member.id or loan.account_id != getattr(account, 'id', None):
        return Decimal('0.00'), repayment_amount
    locked_loan = (
        LoanRequest.objects.select_for_update().prefetch_related('repayments')
        .get(pk=loan.pk, status='APPROVED')
    )
    outstanding = locked_loan.outstanding_balance_as_of(paid_on)
    if repayment_amount > outstanding:
        return Decimal('0.00'), repayment_amount
    repayment = LoanRepayment(
        loan=locked_loan,
        amount=repayment_amount,
        paid_on=paid_on,
        recorded_by=recorded_by,
        source_deposit=source_deposit,
        notes=notes or 'Recorded from deposit submission.',
    )
    repayment.full_clean()
    repayment.save()
    return repayment_amount, Decimal('0.00')


def _create_welfare_allocations(deposit, account, welfare_weeks):
    if welfare_weeks and not account:
        raise ValidationError('Select an account before allocating welfare weeks.')
    for welfare_week in welfare_weeks:
        occupied = (
            DepositWelfareAllocation.objects.select_for_update()
            .filter(
                account=account,
                welfare_week=welfare_week,
                deposit__status__in=('PENDING', 'APPROVED'),
            )
            .exists()
        )
        if occupied:
            raise ValidationError(f'Welfare for {welfare_week:%d %b %Y} is already paid or pending.')
        DepositWelfareAllocation.objects.create(
            deposit=deposit,
            account=account,
            welfare_week=welfare_week,
        )

@login_required
def submit_deposit(request):
    if not (
        request.user.is_member()
        or request.user.is_secretary()
        or request.user.is_mobilizer()
        or request.user.is_chairman()
        or request.user.is_vice_chairman()
        or request.user.is_overseer()
        or request.user.is_treasurer()
    ):
        return redirect('login')

    group_settings = GroupSettings.get_active()
    if not group_settings:
        if request.user.is_superuser or request.user.is_treasurer() or request.user.is_chairman():
            messages.error(request, "Open the saving cycle in Group Settings before deposits can be submitted.")
            return redirect('group_settings')
        messages.error(request, "The saving cycle has not been opened yet. Please contact the Treasurer.")
        return redirect('member_dashboard')

    active_account = get_active_account(request, request.user)
    if get_user_active_accounts(request.user).count() > 1 and not active_account:
        messages.info(request, "Please select a savings account first.")
        return redirect('select_savings_account')

    saving_week = current_saving_week(group_settings.week_one_start, timezone.localdate())
    current_week_start = saving_week.week_start
    current_year_state = financial_year(saving_week.saving_year)
    if current_year_state.state != current_year_state.STATE_OPEN:
        messages.warning(
            request,
            f'The {saving_week.saving_year} financial year is closed for new submissions.',
        )
        return redirect('member_dashboard')

    if request.method == 'POST':
        form = DepositSubmissionForm(
            request.POST,
            request.FILES,
            user=request.user,
            payment_week=current_week_start,
            allow_backdated_payment=request.user.is_treasurer(),
        )
        if form.is_valid():
            account = active_account or form.cleaned_data.get('account')
            selected_weeks = form.cleaned_data.get('selected_week_dates') or [current_week_start]
            weekly_allocations = dict(form.cleaned_data.get('weekly_allocations') or [])
            payment_week = selected_weeks[0]
            proof = form.cleaned_data.get('proof')
            remarks = form.cleaned_data.get('remarks', '')
            payment_date = form.cleaned_data['payment_date']
            payment_time = form.cleaned_data['payment_time']
            saving_amount = form.cleaned_data.get('saving_amount') or 0
            welfare_amount = form.cleaned_data.get('welfare_amount') or 0
            annual_subscription_amount = form.cleaned_data.get('annual_subscription_amount') or 0
            membership_amount = form.cleaned_data.get('membership_amount') or 0
            fine_amount = form.cleaned_data.get('fine_amount') or 0
            shares_amount = form.cleaned_data.get('shares_amount') or 0
            loan_repayment_amount = form.cleaned_data.get('loan_repayment_amount') or 0
            loan_repayment_loan = form.cleaned_data.get('loan_repayment_loan')
            fine_allocations = form.cleaned_data.get('fine_allocations') or []
            welfare_weeks = form.cleaned_data.get('welfare_week_dates') or []
            selected_years = {week.year for week in selected_weeks + welfare_weeks}
            selected_years.update(
                week.year
                for week in Fine.objects.filter(
                    id__in=[fine_id for fine_id, _amount in fine_allocations],
                    reference_week__isnull=False,
                ).values_list('reference_week', flat=True)
            )
            if any(submissions_locked_for_year(year) for year in selected_years):
                form.add_error(
                    None,
                    'One or more selected weeks belong to a closed financial year.',
                )
            if loan_repayment_loan and loan_activity_frozen(loan_repayment_loan):
                form.add_error(
                    'loan_repayment_amount',
                    'This loan is frozen for financial-year settlement.',
                )
            if form.errors:
                return render(request, 'deposits/submit_deposit.html', {
                    'form': form, 'loan_account_ids_json': '[]',
                })

            deposits_created = []
            submission_batch = uuid.uuid4()
            with transaction.atomic():
                for index, payment_week in enumerate(selected_weeks):
                    deposit = DepositSubmission(
                        submission_batch=submission_batch,
                        member=request.user,
                        account=account,
                        submitted_by=request.user,
                        payment_week=payment_week,
                        starting_week=payment_week,
                        weeks_covered=1,
                        saving_amount=weekly_allocations.get(payment_week, Decimal('0.00')),
                        welfare_amount=welfare_amount if index == 0 else 0,
                        annual_subscription_amount=annual_subscription_amount if index == 0 else 0,
                        membership_amount=membership_amount if index == 0 else 0,
                        fine_amount=fine_amount if index == 0 else 0,
                        shares_amount=shares_amount if index == 0 else 0,
                        loan_repayment_amount=loan_repayment_amount if index == 0 else 0,
                        loan_repayment_loan=loan_repayment_loan if index == 0 else None,
                        proof=proof if index == 0 else None,
                        remarks=remarks if index == 0 else '',
                        payment_date=payment_date,
                        payment_time=payment_time,
                        status='PENDING',
                    )
                    deposit.full_clean()
                    deposit.save()
                    deposits_created.append(deposit)
                first_deposit = deposits_created[0]
                for fine_id, amount in fine_allocations:
                    DepositFineAllocation.objects.create(
                        deposit=first_deposit,
                        fine_id=fine_id,
                        amount=amount,
                    )
                _create_welfare_allocations(first_deposit, account, welfare_weeks)
            amount = sum((deposit.amount for deposit in deposits_created), Decimal('0.00'))

            # Send acknowledgment email
            subject = "Deposit Submission Acknowledgment"
            message = (
                f"Dear {request.user.first_name or request.user.username},\n\n"
                f"Thank you for your deposit submission of UGX {amount:,} "
                f"made on {payment_date.strftime('%d %B %Y')}.\n\n"
                "The Treasury Department will review and approve it soon. "
                "You will receive a confirmation email after it has been approved.\n\n"
                "Regards,\n"
                "Land Investment Group"
            )
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [request.user.email]

            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            # --- Send notification email to treasurer(s) ---
            treasurer_emails = list(
                MemberProfile.objects.filter(role='TREASURER').values_list('email', flat=True)
            )
            if treasurer_emails:
                subject_treasurer = "New Deposit Submission - Action Required"
                message_treasurer = (
                    f"Dear Treasurer,\n\n"
                    f"Member {request.user.get_full_name() or request.user.username} "
                    f"has submitted a deposit of UGX {amount:,} "
                    f"for week of {payment_week.strftime('%d %B %Y')}.\n\n"
                    "Please log in to the system to review and process this deposit.\n\n"
                    "Regards,\n"
                    "Land Investment Group"
                )
                send_mail(
                    subject_treasurer,
                    message_treasurer,
                    settings.DEFAULT_FROM_EMAIL,
                    treasurer_emails,
                    fail_silently=False
                )

            messages.success(request, f"Weekly saving submitted for week of {payment_week}.")
            return redirect('member_dashboard')
    else:
        form = DepositSubmissionForm(
            user=request.user,
            payment_week=current_week_start,
            allow_backdated_payment=request.user.is_treasurer(),
            initial={'account': active_account.id} if active_account else None,
        )
        if active_account:
            form.fields['account'].queryset = SavingsAccount.objects.filter(id=active_account.id)
            form.fields['account'].initial = active_account.id

    # Build the set of account IDs that have an active (approved) loan for this user
    # so the template can show the Loan Repayment checkbox only for those accounts.
    import json as _json
    loan_account_ids = list(
        LoanRequest.objects.filter(member=request.user, status='APPROVED')
        .exclude(account__isnull=True)
        .values_list('account_id', flat=True)
        .distinct()
    )
    return render(request, 'deposits/submit_deposit.html', {
        'form': form,
        'loan_account_ids_json': _json.dumps(loan_account_ids),
    })


@login_required
@transaction.atomic
def approve_deposit(request, deposit_id):
    if not request.user.is_treasurer():
        messages.error(request, "Access denied.")
        return redirect('member_dashboard')

    if request.method != 'POST':
        messages.error(request, 'Use the approval button to approve a deposit submission.')
        return redirect('manage_deposits')
    legacy_deposit = get_object_or_404(DepositSubmission, id=deposit_id)
    return approve_deposit_batch(request, legacy_deposit.submission_batch)

    deposit = get_object_or_404(DepositSubmission, id=deposit_id, status='PENDING')
    close_state = financial_year(deposit.payment_week.year)
    if (
        close_state.cutoff_at
        and deposit.date_submitted > close_state.cutoff_at
    ):
        messages.error(request, 'This deposit was submitted after the financial-year cutoff.')
        return redirect('manage_deposits')
    if deposit.saving_amount > 0 and not (
        MIN_WEEKLY_SAVINGS <= deposit.saving_amount <= MAX_WEEKLY_SAVINGS
    ):
        messages.error(request, 'Weekly savings must be between UGX 10,000 and UGX 50,000.')
        return redirect('manage_deposits')
    if deposit.saving_amount > 0 and weekly_savings_total(
        deposit.member, deposit.account, deposit.payment_week
    ) >= MIN_WEEKLY_SAVINGS:
        messages.error(request, 'This savings week is already paid and locked. The duplicate submission was not approved.')
        return redirect('manage_deposits')
    if deposit.loan_repayment_amount > 0 and not deposit.loan_repayment_loan_id:
        messages.error(request, 'Select the exact loan before approving this repayment deposit.')
        return redirect('financial_record_edit', record_type='deposit', object_id=deposit.id)
    selected_fine_allocations = list(
        deposit.fine_allocations.select_related('fine').all()
    )
    if selected_fine_allocations:
        try:
            apply_selected_fine_allocations(
                (allocation.fine, allocation.amount)
                for allocation in selected_fine_allocations
            )
        except ValidationError as exc:
            messages.error(request, str(exc))
            return redirect('manage_deposits')
    deposit.status = 'APPROVED'
    deposit.reviewed_by = request.user
    deposit.date_reviewed = timezone.now()
    deposit.save()
    if deposit.saving_amount > 0:
        delete_deposit_week_missed_saving_fines(deposit)
    if deposit.fine_amount > 0 and not selected_fine_allocations:
        allocate_fine_payment(deposit.member, deposit.account, deposit.fine_amount)

    if deposit.loan_repayment_amount and deposit.loan_repayment_amount > 0:
        allocated, unallocated = _apply_loan_repayment(
            member=deposit.member,
            account=deposit.account,
            loan=deposit.loan_repayment_loan,
            repayment_amount=deposit.loan_repayment_amount,
            paid_on=deposit.payment_date,
            recorded_by=request.user,
            source_deposit=deposit,
            notes=f'Deposit approval repayment (Deposit #{deposit.id}).',
        )
        if allocated > 0:
            repayment = deposit.generated_loan_repayments.order_by('-id').first()
            if repayment:
                apply_eligible_locked_repayment(repayment)
            messages.success(request, f"UGX {allocated:,.0f} was posted to loan repayment.")
        if unallocated > 0:
            messages.warning(request, f"UGX {unallocated:,.0f} could not be posted because no outstanding approved loan balance was found.")

    # Send approval email
    send_mail(
        subject="Deposit Approved",
        message=(
            f"Dear {deposit.member.get_full_name()},\n\n"
            f"Your deposit of {deposit.amount} made on {deposit.payment_date.strftime('%d %B %Y')} "
            f"has been approved by the Treasury Department.\n\n"
            "Thank you for your continued commitment.\n\n"
            "Regards,\nTreasury Department"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[deposit.member.email],
        fail_silently=True,
    )

    messages.success(request, "Deposit approved and member notified by email.")
    return redirect('treasurer_dashboard')



@login_required
def reject_deposit(request, deposit_id):
    if not request.user.is_treasurer():
        messages.error(request, "Access denied.")
        return redirect('member_dashboard')

    if request.method != 'POST':
        messages.error(request, 'Use the rejection button to reject a deposit submission.')
        return redirect('manage_deposits')
    legacy_deposit = get_object_or_404(DepositSubmission, id=deposit_id)
    return reject_deposit_batch(request, legacy_deposit.submission_batch)

    deposit = get_object_or_404(DepositSubmission, id=deposit_id, status='PENDING')
    deposit.status = 'REJECTED'
    deposit.reviewed_by = request.user
    deposit.date_reviewed = timezone.now()
    deposit.save()

    # Send rejection email
    send_mail(
        subject="Deposit Rejected",
        message=(
            f"Dear {deposit.member.get_full_name()},\n\n"
            f"Your deposit of {deposit.amount} made on {deposit.payment_date.strftime('%d %B %Y')} "
            "has been rejected. Please contact the Treasury Department for more information.\n\n"
            "Regards,\nTreasury Department"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[deposit.member.email],
        fail_silently=True,
    )

    messages.warning(request, "Deposit rejected and member notified by email.")
    return redirect('treasurer_dashboard')


@login_required
def delete_deposit(request, deposit_id):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can delete deposit records.')
        return redirect('member_dashboard')
    if request.method != 'POST':
        messages.error(request, 'Use the delete confirmation form to delete a record.')
        return redirect('manage_deposits')
    legacy_deposit = get_object_or_404(DepositSubmission, id=deposit_id)
    return delete_deposit_batch(request, legacy_deposit.submission_batch)

    try:
        record_id = delete_deposit_with_audit(
            deposit_id,
            request.user,
            request.POST.get('deletion_reason'),
        )
    except DepositSubmission.DoesNotExist:
        messages.error(request, 'That deposit record no longer exists.')
    except ValidationError as exc:
        messages.error(
            request,
            '; '.join(exc.messages) if hasattr(exc, 'messages') else str(exc),
        )
    else:
        messages.success(
            request,
            f'Deposit #{record_id} was deleted from the database. Its audit history was saved.',
        )
    return redirect('manage_deposits')


def _batch_records(batch_id, for_update=False):
    queryset = (
        DepositSubmission.objects
        .filter(submission_batch=batch_id)
        .select_related('member', 'account', 'loan_repayment_loan')
        .prefetch_related('fine_allocations__fine', 'welfare_allocations', 'generated_loan_repayments')
        .order_by('id')
    )
    if for_update:
        queryset = queryset.select_for_update()
    return list(queryset)


def _send_batch_status_email(records, approved):
    first = records[0]
    total = sum((item.amount for item in records), Decimal('0.00'))
    status_word = 'approved' if approved else 'rejected'
    send_mail(
        subject=f"Deposit {status_word.title()}",
        message=(
            f"Dear {first.member.get_full_name() or first.member.username},\n\n"
            f"Your deposit submission of UGX {total:,.0f} made on "
            f"{first.payment_date.strftime('%d %B %Y')} has been {status_word} "
            "by the Treasury Department.\n\nRegards,\nTreasury Department"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[first.member.email],
        fail_silently=True,
    )


@login_required
def approve_deposit_batch(request, batch_id):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can approve deposits.')
        return redirect('member_dashboard')
    if request.method != 'POST':
        messages.error(request, 'Use the approval button to approve a deposit submission.')
        return redirect('manage_deposits')

    try:
        with transaction.atomic():
            records = _batch_records(batch_id, for_update=True)
            if not records:
                raise ValidationError('That deposit submission no longer exists.')
            if any(item.status != 'PENDING' for item in records):
                raise ValidationError('Only a fully pending submission can be approved.')

            for deposit in records:
                close_state = financial_year(deposit.payment_week.year)
                if close_state.cutoff_at and deposit.date_submitted > close_state.cutoff_at:
                    raise ValidationError(
                        'This submission was received after the financial-year cutoff.'
                    )
                if deposit.saving_amount > 0:
                    if not MIN_WEEKLY_SAVINGS <= deposit.saving_amount <= MAX_WEEKLY_SAVINGS:
                        raise ValidationError(
                            'Every weekly saving must be between UGX 10,000 and UGX 50,000.'
                        )
                    if weekly_savings_total(
                        deposit.member,
                        deposit.account,
                        deposit.payment_week,
                    ) >= MIN_WEEKLY_SAVINGS:
                        raise ValidationError(
                            f'The week of {deposit.payment_week:%d %b %Y} is already paid.'
                        )

            for deposit in records:
                selected_fines = list(deposit.fine_allocations.select_related('fine'))
                if selected_fines:
                    apply_selected_fine_allocations(
                        (allocation.fine, allocation.amount)
                        for allocation in selected_fines
                    )
                deposit.status = 'APPROVED'
                deposit.reviewed_by = request.user
                deposit.date_reviewed = timezone.now()
                deposit.save(update_fields=['status', 'reviewed_by', 'date_reviewed', 'amount'])
                if deposit.saving_amount > 0:
                    delete_deposit_week_missed_saving_fines(deposit)
                if deposit.fine_amount > 0 and not selected_fines:
                    allocate_fine_payment(
                        deposit.member, deposit.account, deposit.fine_amount
                    )
                if deposit.loan_repayment_amount > 0:
                    allocated, unallocated = _apply_loan_repayment(
                        member=deposit.member,
                        account=deposit.account,
                        loan=deposit.loan_repayment_loan,
                        repayment_amount=deposit.loan_repayment_amount,
                        paid_on=deposit.payment_date,
                        recorded_by=request.user,
                        source_deposit=deposit,
                        notes=f'Deposit approval repayment (Deposit #{deposit.id}).',
                    )
                    if unallocated or allocated != deposit.loan_repayment_amount:
                        raise ValidationError(
                            'The selected loan no longer has enough outstanding balance.'
                        )
                    repayment = deposit.generated_loan_repayments.order_by('-id').first()
                    if repayment:
                        apply_eligible_locked_repayment(repayment)
            transaction.on_commit(lambda: _send_batch_status_email(records, True))
    except ValidationError as exc:
        messages.error(
            request,
            '; '.join(exc.messages) if hasattr(exc, 'messages') else str(exc),
        )
    else:
        messages.success(request, 'The complete deposit submission was approved.')
    return redirect('manage_deposits')


@login_required
def reject_deposit_batch(request, batch_id):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can reject deposits.')
        return redirect('member_dashboard')
    if request.method != 'POST':
        messages.error(request, 'Use the rejection button to reject a deposit submission.')
        return redirect('manage_deposits')

    with transaction.atomic():
        records = _batch_records(batch_id, for_update=True)
        if not records:
            messages.error(request, 'That deposit submission no longer exists.')
            return redirect('manage_deposits')
        if any(item.status != 'PENDING' for item in records):
            messages.error(request, 'Only a fully pending submission can be rejected.')
            return redirect('manage_deposits')
        reviewed_at = timezone.now()
        DepositSubmission.objects.filter(
            submission_batch=batch_id
        ).update(
            status='REJECTED',
            reviewed_by=request.user,
            date_reviewed=reviewed_at,
        )
        for item in records:
            item.status = 'REJECTED'
        transaction.on_commit(lambda: _send_batch_status_email(records, False))
    messages.success(request, 'The complete deposit submission was rejected.')
    return redirect('manage_deposits')


@login_required
def delete_deposit_batch(request, batch_id):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can delete deposit records.')
        return redirect('member_dashboard')
    if request.method != 'POST':
        messages.error(request, 'Use the delete confirmation form to delete a submission.')
        return redirect('manage_deposits')
    reason = request.POST.get('deletion_reason')
    try:
        with transaction.atomic():
            records = _batch_records(batch_id, for_update=True)
            if not records:
                raise ValidationError('That deposit submission no longer exists.')
            for deposit in records:
                delete_deposit_with_audit(deposit.id, request.user, reason)
    except ValidationError as exc:
        messages.error(
            request,
            '; '.join(exc.messages) if hasattr(exc, 'messages') else str(exc),
        )
    else:
        messages.success(
            request,
            f'The complete deposit submission ({len(records)} record(s)) was deleted and audited.',
        )
    return redirect('manage_deposits')


def _batch_edit_context(records):
    primary = next(
        (
            item for item in records
            if item.proof
            or item.welfare_amount
            or item.annual_subscription_amount
            or item.membership_amount
            or item.fine_amount
            or item.shares_amount
            or item.loan_repayment_amount
        ),
        records[0],
    )
    totals = {
        field_name: sum(
            (getattr(item, field_name) or Decimal('0.00') for item in records),
            Decimal('0.00'),
        )
        for field_name in (
            'saving_amount',
            'welfare_amount',
            'annual_subscription_amount',
            'membership_amount',
            'fine_amount',
            'shares_amount',
            'loan_repayment_amount',
        )
    }
    purpose_by_field = {
        'saving_amount': 'saving',
        'welfare_amount': 'welfare',
        'annual_subscription_amount': 'annual_subscription',
        'membership_amount': 'membership',
        'fine_amount': 'fine',
        'shares_amount': 'shares',
        'loan_repayment_amount': 'loan_repayment',
    }
    selected_weeks = [
        item.payment_week.isoformat()
        for item in records
        if item.saving_amount > 0
    ]
    fine_allocations = [
        allocation
        for item in records
        for allocation in item.fine_allocations.select_related('fine')
    ]
    welfare_weeks = [
        allocation.welfare_week.isoformat()
        for item in records
        for allocation in item.welfare_allocations.all()
    ]
    initial = {
        'member': primary.member_id,
        'account': primary.account_id,
        'payment_week': primary.payment_week,
        'payment_date': primary.payment_date,
        'payment_time': primary.payment_time,
        'loan_repayment_loan': primary.loan_repayment_loan_id,
        'remarks': next((item.remarks for item in records if item.remarks), ''),
        'selected_purposes': [
            purpose_by_field[field_name]
            for field_name, value in totals.items()
            if value > 0
        ],
        'selected_weeks': selected_weeks,
        'selected_fine_weeks': sorted({
            allocation.fine.reference_week.isoformat()
            for allocation in fine_allocations
            if allocation.fine.reference_week
        }),
        'selected_welfare_weeks': welfare_weeks,
        'week_amounts': {
            item.payment_week.isoformat(): str(item.saving_amount)
            for item in records
            if item.saving_amount > 0
        },
        **totals,
    }
    fine_credit = {}
    if all(item.status == 'APPROVED' for item in records):
        for allocation in fine_allocations:
            fine_credit[allocation.fine_id] = (
                fine_credit.get(allocation.fine_id, Decimal('0.00'))
                + allocation.amount
            )
    loan_credit = sum(
        (
            repayment.amount
            for item in records
            for repayment in item.generated_loan_repayments.all()
        ),
        Decimal('0.00'),
    )
    return primary, initial, fine_credit, loan_credit


def _deleted_record_revision(record, before, editor, reason):
    latest = (
        FinancialRecordRevision.objects.select_for_update()
        .filter(record_type='deposit', object_id=record.id)
        .order_by('-revision_number')
        .first()
    )
    after = dict(before)
    after['record_state'] = 'Deleted during submission correction'
    FinancialRecordRevision.objects.create(
        record_type='deposit',
        object_id=record.id,
        revision_number=(latest.revision_number + 1) if latest else 1,
        before_data=before,
        after_data=after,
        reason=reason,
        edited_by=editor,
    )


def _save_deposit_batch_edit(batch_id, form, editor):
    with transaction.atomic():
        records = _batch_records(batch_id, for_update=True)
        if not records:
            raise ValidationError('That deposit submission no longer exists.')
        statuses = {item.status for item in records}
        if len(statuses) != 1:
            raise ValidationError(
                'This historical submission has mixed statuses and must be reviewed manually.'
            )
        original_status = statuses.pop()
        reason = form.cleaned_data['edit_reason'].strip()
        before = {item.id: snapshot_record(item) for item in records}

        if original_status == 'APPROVED':
            for item in records:
                allocations = list(item.fine_allocations.select_related('fine'))
                if item.fine_amount > 0 and not allocations:
                    raise ValidationError(
                        'This approved legacy fine payment is not linked to a specific fine.'
                    )
                for allocation in allocations:
                    fine = Fine.objects.select_for_update().get(pk=allocation.fine_id)
                    fine.amount_paid = max(
                        fine.amount_paid - allocation.amount,
                        Decimal('0.00'),
                    )
                    fine.is_paid = fine.amount_paid >= fine.amount
                    fine.save(update_fields=['amount_paid', 'is_paid'])
                item.generated_loan_repayments.all().delete()

        for item in records:
            item.fine_allocations.all().delete()
            item.welfare_allocations.all().delete()

        selected_weeks = list(form.cleaned_data.get('selected_week_dates') or [])
        weekly_allocations = dict(form.cleaned_data.get('weekly_allocations') or [])
        if not selected_weeks:
            selected_weeks = [form.cleaned_data['payment_week']]

        target_status = 'PENDING' if original_status == 'REJECTED' else original_status
        reviewed_by = editor if target_status == 'APPROVED' else None
        reviewed_at = timezone.now() if target_status == 'APPROVED' else None
        primary_old = records[0]
        existing_proof = next((item.proof for item in records if item.proof), None)
        proof = form.cleaned_data.get('proof')
        if not proof and not form.cleaned_data.get('clear_proof'):
            proof = existing_proof

        unused = list(records)
        saved = []
        created_ids = set()
        for index, payment_week in enumerate(selected_weeks):
            match = next(
                (
                    item for item in unused
                    if item.payment_week == payment_week
                    and (item.saving_amount > 0) == bool(weekly_allocations)
                ),
                None,
            )
            if match is None and unused:
                match = unused[0]
            if match is None:
                match = DepositSubmission(
                    submission_batch=batch_id,
                    submitted_by=primary_old.submitted_by or editor,
                )
            else:
                unused.remove(match)
            first = index == 0
            match.submission_batch = batch_id
            match.member = form.cleaned_data['member']
            match.account = form.cleaned_data.get('account')
            match.payment_week = payment_week
            match.starting_week = payment_week
            match.weeks_covered = 1
            match.saving_amount = weekly_allocations.get(
                payment_week, Decimal('0.00')
            )
            match.welfare_amount = (
                form.cleaned_data.get('welfare_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.annual_subscription_amount = (
                form.cleaned_data.get('annual_subscription_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.membership_amount = (
                form.cleaned_data.get('membership_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.fine_amount = (
                form.cleaned_data.get('fine_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.shares_amount = (
                form.cleaned_data.get('shares_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.loan_repayment_amount = (
                form.cleaned_data.get('loan_repayment_amount') or Decimal('0.00')
            ) if first else Decimal('0.00')
            match.loan_repayment_loan = (
                form.cleaned_data.get('loan_repayment_loan') if first else None
            )
            match.proof = proof if first else None
            match.remarks = form.cleaned_data.get('remarks', '') if first else ''
            match.payment_date = form.cleaned_data['payment_date']
            match.payment_time = form.cleaned_data['payment_time']
            match.status = target_status
            match.reviewed_by = reviewed_by
            match.date_reviewed = reviewed_at
            is_new = match.pk is None
            match.full_clean()
            match.save()
            if is_new:
                created_ids.add(match.id)
            saved.append(match)

        for removed in unused:
            _deleted_record_revision(
                removed, before[removed.id], editor, reason
            )
            removed.delete()

        primary = saved[0]
        for fine_id, amount in form.cleaned_data.get('fine_allocations') or []:
            DepositFineAllocation.objects.create(
                deposit=primary,
                fine_id=fine_id,
                amount=amount,
            )
        _create_welfare_allocations(
            primary,
            primary.account,
            form.cleaned_data.get('welfare_week_dates') or [],
        )

        if target_status == 'APPROVED':
            allocations = list(primary.fine_allocations.select_related('fine'))
            if allocations:
                apply_selected_fine_allocations(
                    (allocation.fine, allocation.amount)
                    for allocation in allocations
                )
            for item in saved:
                if item.saving_amount > 0:
                    delete_deposit_week_missed_saving_fines(item)
            if primary.fine_amount > 0 and not allocations:
                allocate_fine_payment(
                    primary.member, primary.account, primary.fine_amount
                )
            if primary.loan_repayment_amount > 0:
                allocated, unallocated = _apply_loan_repayment(
                    member=primary.member,
                    account=primary.account,
                    loan=primary.loan_repayment_loan,
                    repayment_amount=primary.loan_repayment_amount,
                    paid_on=primary.payment_date,
                    recorded_by=editor,
                    source_deposit=primary,
                    notes=f'Corrected deposit repayment (Deposit #{primary.id}).',
                )
                if unallocated or allocated != primary.loan_repayment_amount:
                    raise ValidationError(
                        'The selected loan no longer has enough outstanding balance.'
                    )

        for item in saved:
            if item.id in created_ids:
                created_before = {
                    'submission_batch': str(batch_id),
                    'record_state': 'Not yet created',
                }
                record_revision(item, created_before, editor, reason)
            else:
                record_revision(item, before[item.id], editor, reason)

        from groupcore.models import FinancialYearClose
        affected_years = {
            item.payment_week.year
            for item in [*records, *saved]
            if item.payment_week
        }
        FinancialYearClose.objects.filter(
            year__in=affected_years,
            state=FinancialYearClose.STATE_FINALIZED,
        ).update(
            needs_regeneration=True,
            last_correction_at=timezone.now(),
        )
    return saved


@login_required
def edit_deposit_batch(request, batch_id):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can edit deposit submissions.')
        return redirect('member_dashboard')
    records = _batch_records(batch_id)
    if not records:
        messages.error(request, 'That deposit submission no longer exists.')
        return redirect('manage_deposits')
    primary, initial, fine_credit, loan_credit = _batch_edit_context(records)
    form_kwargs = {
        'initial': initial,
        'exclude_deposit_ids': [item.id for item in records],
        'extra_saving_weeks': [item.payment_week for item in records],
        'fine_credit_by_id': fine_credit,
        'loan_repayment_credit': loan_credit,
        'loan_credit_loan_id': primary.loan_repayment_loan_id,
    }
    if request.method == 'POST':
        form = DepositBatchEditForm(
            request.POST,
            request.FILES,
            **form_kwargs,
        )
        if form.is_valid():
            try:
                _save_deposit_batch_edit(batch_id, form, request.user)
            except ValidationError as exc:
                form.add_error(
                    None,
                    '; '.join(exc.messages)
                    if hasattr(exc, 'messages')
                    else str(exc),
                )
            else:
                messages.success(
                    request,
                    'Deposit submission corrected and audit history saved.',
                )
                return redirect('deposit_batch_history', batch_id=batch_id)
    else:
        form = DepositBatchEditForm(**form_kwargs)

    loan_account_ids = list(
        LoanRequest.objects.filter(status='APPROVED')
        .exclude(account__isnull=True)
        .values_list('account_id', flat=True)
        .distinct()
    )
    import json as _json
    return render(request, 'deposits/submit_deposit.html', {
        'form': form,
        'page_title': 'Edit Deposit Submission',
        'submit_label': 'Save Corrected Submission',
        'existing_proof': primary.proof or next(
            (item.proof for item in records if item.proof), None
        ),
        'loan_account_ids_json': _json.dumps(loan_account_ids),
    })


@login_required
def deposit_batch_history(request, batch_id):
    records = _batch_records(batch_id)
    current_ids = {item.id for item in records}
    revisions = list(
        FinancialRecordRevision.objects
        .filter(record_type='deposit')
        .select_related('edited_by')
        .order_by('edited_at', 'id')
    )

    def belongs_to_batch(revision):
        if revision.object_id in current_ids:
            return True
        for snapshot in (revision.before_data or {}, revision.after_data or {}):
            if str(snapshot.get('submission_batch') or '') == str(batch_id):
                return True
        return False

    revisions = [item for item in revisions if belongs_to_batch(item)]
    if not records and not revisions:
        raise Http404
    if records:
        allowed = request.user.is_treasurer() or request.user.role in {
            'CHAIRMAN', 'VICE_CHAIRMAN', 'SECRETARY', 'OVERSEER',
        } or request.user.pk == records[0].member_id
        record_display = str(DepositBatchSummary(records).member)
    else:
        member_data = (revisions[-1].before_data or {}).get('member') or {}
        allowed = request.user.role in {
            'TREASURER', 'CHAIRMAN', 'VICE_CHAIRMAN', 'SECRETARY', 'OVERSEER',
        } or request.user.pk == member_data.get('id')
        record_display = member_data.get('label') or 'Deleted submission'
    if not allowed:
        messages.error(request, 'You do not have permission to inspect this submission.')
        return redirect('member_dashboard')
    from groupcore.financial_records import revision_changes
    revision_items = [
        {'revision': revision, 'changes': revision_changes(revision)}
        for revision in revisions
    ]
    return render(request, 'deposits/deposit_batch_history.html', {
        'batch_id': batch_id,
        'record_display': record_display,
        'revision_items': revision_items,
        'can_edit': request.user.is_treasurer() and bool(records),
        'is_deleted': not records,
    })


@login_required
def deposit_deletion_audit(request):
    if not request.user.is_treasurer():
        messages.error(request, 'Only the treasurer can view the deletion audit.')
        return redirect('member_dashboard')
    deletions = (
        FinancialRecordRevision.objects
        .filter(record_type='deposit', after_data__record_state='Deleted')
        .select_related('edited_by')
        .order_by('-edited_at', '-id')
    )
    page = Paginator(deletions, 25).get_page(request.GET.get('page'))
    return render(request, 'deposits/deposit_deletion_audit.html', {
        'deletions': page,
        'page_obj': page,
        'pagination_query': pagination_query(request),
    })





MONTHS = [
    ("01", "January"), ("02", "February"), ("03", "March"), ("04", "April"),
    ("05", "May"), ("06", "June"), ("07", "July"), ("08", "August"),
    ("09", "September"), ("10", "October"), ("11", "November"), ("12", "December")
]


def _optional_int(value, min_value=None, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if min_value is not None and parsed < min_value:
        return None
    if max_value is not None and parsed > max_value:
        return None
    return parsed


def _approved_deposit_totals(queryset):
    totals = queryset.filter(status='APPROVED').aggregate(
        total=Sum('amount'),
        saving=Sum('saving_amount'),
        welfare=Sum('welfare_amount'),
        annual_subscription=Sum('annual_subscription_amount'),
        membership=Sum('membership_amount'),
        fine=Sum('fine_amount'),
        shares=Sum('shares_amount'),
        loan_repayment=Sum('loan_repayment_amount'),
    )
    return {
        key: totals.get(key) or Decimal('0')
        for key in [
            'total',
            'saving',
            'welfare',
            'annual_subscription',
            'membership',
            'fine',
            'shares',
            'loan_repayment',
        ]
    }


def _my_contributions_data(request):
    user = request.user
    active_account = get_active_account(request, user)
    base_deposits = DepositSubmission.objects.filter(member=user)
    if active_account:
        base_deposits = base_deposits.filter(account=active_account)

    current_year = timezone.localdate().year
    selected_year = _optional_int(request.GET.get('year'), min_value=1, max_value=9999) or current_year
    selected_month = _optional_int(request.GET.get('month'), min_value=1, max_value=12)

    deposits = base_deposits.order_by('-payment_week', '-date_submitted')
    deposits = deposits.filter(payment_week__year=selected_year)
    if selected_month:
        deposits = deposits.filter(payment_week__month=selected_month)
    deposits_page = Paginator(deposits, 25).get_page(request.GET.get('page'))
    years = merge_year_options(
        years_from_dates(base_deposits, 'payment_week'),
        selected_year=selected_year,
        default_year=current_year,
    )

    return {
        'active_account': active_account,
        'deposits': deposits,
        'deposits_page': deposits_page,
        'years': years,
        'selected_year': selected_year,
        'selected_month': f'{selected_month:02d}' if selected_month else '',
        'selected_month_number': selected_month,
        'approved_totals': _approved_deposit_totals(deposits),
    }


def _month_label(month_number):
    if not month_number:
        return 'All Months'
    return dict(MONTHS).get(f'{month_number:02d}', 'All Months')


def _safe_filename_part(value):
    value = str(value or 'none')
    return ''.join(char if char.isalnum() or char in ('-', '_') else '_' for char in value)


def _my_contributions_filename(account, selected_year, selected_month, extension):
    account_label = _safe_filename_part(account.label if account else 'no_account')
    if selected_year and selected_month:
        period = f'{selected_year}_{selected_month:02d}'
    elif selected_year:
        period = str(selected_year)
    elif selected_month:
        period = f'month_{selected_month:02d}'
    else:
        period = 'all'
    return f'my_contributions_{account_label}_{period}.{extension}'


def _proof_reference(deposit):
    return deposit.proof.name if deposit.proof else '-'


def _my_contributions_report_meta(user, active_account, contribution_data):
    return {
        'member_name': user.get_full_name() or user.username,
        'username': user.username,
        'account_label': active_account.label if active_account else '-',
        'year_label': contribution_data['selected_year'] or 'All Years',
        'month_label': _month_label(contribution_data['selected_month_number']),
        'generated_at': now().strftime('%Y-%m-%d %H:%M'),
    }


def _my_contributions_detail_rows(deposits):
    rows = []
    for deposit in deposits:
        rows.append([
            deposit.payment_week.strftime('%Y-%m-%d') if deposit.payment_week else '-',
            deposit.account.label if deposit.account else '-',
            deposit.amount,
            deposit.saving_amount,
            deposit.welfare_amount,
            deposit.annual_subscription_amount,
            deposit.membership_amount,
            deposit.fine_amount,
            deposit.shares_amount,
            deposit.loan_repayment_amount,
            deposit.status.title(),
            deposit.payment_date.strftime('%Y-%m-%d') if deposit.payment_date else '-',
            deposit.payment_time.strftime('%H:%M') if deposit.payment_time else '-',
            deposit.submitted_by.username if deposit.submitted_by else '-',
            deposit.remarks or '-',
            _proof_reference(deposit),
        ])
    return rows


def _money(value):
    return f"UGX {Decimal(value or 0):,.0f}"


def _pdf_text(value, style):
    """Create a wrapping, safely escaped PDF cell for long report text."""
    text = '-' if value is None or value == '' else str(value)
    return Paragraph(xml_escape(text).replace('\n', '<br/>'), style)


def _export_my_contributions_pdf(user, active_account, contribution_data, deposits):
    response = HttpResponse(content_type='application/pdf')
    filename = _my_contributions_filename(
        active_account,
        contribution_data['selected_year'],
        contribution_data['selected_month_number'],
        'pdf',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    cell_style = styles['BodyText'].clone('member-contribution-cell')
    cell_style.fontSize = 6.5
    cell_style.leading = 8
    header_style = styles['BodyText'].clone('member-contribution-header')
    header_style.fontSize = 6.5
    header_style.leading = 8
    header_style.textColor = colors.white
    header_style.alignment = 1
    meta = _my_contributions_report_meta(user, active_account, contribution_data)
    totals = contribution_data['approved_totals']

    elements = [
        Paragraph(f"Financial Report for {meta['member_name']}", styles['Title']),
        Paragraph(
            f"Username: {meta['username']} | Savings Account: {meta['account_label']} | "
            f"Period: {meta['year_label']} / {meta['month_label']} | Generated: {meta['generated_at']}",
            styles['Normal'],
        ),
        Spacer(1, 10),
    ]

    summary_data = [
        [_pdf_text(header, header_style) for header in ['Approved Total', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine', 'Shares', 'Loan Repayment']],
        [_pdf_text(value, cell_style) for value in [
            _money(totals['total']),
            _money(totals['saving']),
            _money(totals['welfare']),
            _money(totals['annual_subscription']),
            _money(totals['membership']),
            _money(totals['fine']),
            _money(totals['shares']),
            _money(totals['loan_repayment']),
        ]],
    ]
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    detail_headers = [
        'Saving Week', 'Account', 'Total', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine',
        'Shares', 'Loan Repay', 'Status', 'Payment Date', 'Payment Time',
        'Submitted By', 'Remarks', 'Proof Ref',
    ]
    detail_rows = []
    for row in _my_contributions_detail_rows(deposits):
        detail_rows.append([
            row[0], row[1], _money(row[2]), _money(row[3]), _money(row[4]),
            _money(row[5]), _money(row[6]), _money(row[7]), _money(row[8]),
            _money(row[9]), row[10], row[11], row[12], row[13], row[14], row[15],
        ])
    if not detail_rows:
        detail_rows = [['No matching deposits'] + [''] * (len(detail_headers) - 1)]

    detail_table = Table(
        [[_pdf_text(header, header_style) for header in detail_headers]] + [
            [_pdf_text(value, cell_style) for value in row]
            for row in detail_rows
        ],
        repeatRows=1,
        colWidths=[54, 40, 46, 42, 42, 42, 48, 38, 40, 50, 44, 54, 46, 54, 66, 82],
    )
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5ee')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(detail_table)
    doc.build(elements)
    return response


def _export_my_contributions_excel(user, active_account, contribution_data, deposits):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = _my_contributions_filename(
        active_account,
        contribution_data['selected_year'],
        contribution_data['selected_month_number'],
        'xlsx',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Financial Report'
    meta = _my_contributions_report_meta(user, active_account, contribution_data)
    totals = contribution_data['approved_totals']

    ws.merge_cells('A1:P1')
    ws['A1'] = f"Financial Report for {meta['member_name']}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    meta_rows = [
        ('Username', meta['username']),
        ('Savings Account', meta['account_label']),
        ('Year', meta['year_label']),
        ('Month', meta['month_label']),
        ('Generated', meta['generated_at']),
    ]
    for row_number, (label, value) in enumerate(meta_rows, start=3):
        ws.cell(row=row_number, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_number, column=2, value=value)

    summary_start = 10
    ws.cell(row=summary_start, column=1, value='Approved Totals').font = Font(bold=True)
    summary_headers = ['Total', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine', 'Shares', 'Loan Repayment']
    for column, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_start + 1, column=column, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    summary_values = [
        totals['total'],
        totals['saving'],
        totals['welfare'],
        totals['annual_subscription'],
        totals['membership'],
        totals['fine'],
        totals['shares'],
        totals['loan_repayment'],
    ]
    for column, value in enumerate(summary_values, start=1):
        ws.cell(row=summary_start + 2, column=column, value=float(value))

    detail_start = summary_start + 5
    ws.cell(row=detail_start, column=1, value='Deposit Details').font = Font(bold=True)
    headers = [
        'Saving Week', 'Account', 'Total', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine',
        'Shares', 'Loan Repayment', 'Status', 'Payment Date', 'Payment Time',
        'Submitted By', 'Remarks', 'Proof Reference',
    ]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=detail_start + 1, column=column, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_number, row in enumerate(_my_contributions_detail_rows(deposits), start=detail_start + 2):
        for column, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_number, column=column, value=value)

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 42)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal='center' if cell.row in (11, 12) else 'left',
                vertical='top',
                wrap_text=True,
            )

    wb.save(response)
    return response


@login_required
def export_my_contributions(request, format):
    contribution_data = _my_contributions_data(request)
    active_account = contribution_data['active_account']
    if not active_account:
        if get_user_active_accounts(request.user).exists():
            messages.info(request, "Please select a savings account before downloading your report.")
            return redirect('select_savings_account')
        messages.error(request, "No active savings account was found for your profile.")
        return redirect('my_contributions')

    deposits = list(
        contribution_data['deposits'].select_related('account', 'submitted_by')
    )

    if format == 'pdf':
        return _export_my_contributions_pdf(request.user, active_account, contribution_data, deposits)
    if format == 'excel':
        return _export_my_contributions_excel(request.user, active_account, contribution_data, deposits)
    return HttpResponse("Invalid format", status=400)


@login_required
def my_contributions(request):
    contribution_data = _my_contributions_data(request)
    active_account = contribution_data['active_account']

    context = {
        'deposits': contribution_data['deposits_page'],
        'page_obj': contribution_data['deposits_page'],
        'pagination_query': pagination_query(request),
        'total_approved': contribution_data['approved_totals']['total'],
        'approved_totals': contribution_data['approved_totals'],
        'years': contribution_data['years'],
        'selected_year': contribution_data['selected_year'],
        'selected_month': contribution_data['selected_month'],
        'months': MONTHS,  # add this
        'active_account': active_account,
        'export_querystring': request.GET.urlencode(),
    }
    return render(request, 'deposits/my_contributions.html', context)


class DepositBatchSummary:
    """Display-facing aggregate for the per-week records in one submission."""

    purpose_fields = (
        ('Saving', 'saving_amount'),
        ('Welfare', 'welfare_amount'),
        ('Annual Subscription', 'annual_subscription_amount'),
        ('Membership', 'membership_amount'),
        ('Fine', 'fine_amount'),
        ('Shares', 'shares_amount'),
        ('Loan Repayment', 'loan_repayment_amount'),
    )

    def __init__(self, records):
        self.records = sorted(records, key=lambda item: item.id)
        self.record_ids = [item.id for item in self.records]
        self.submission_batch = self.records[0].submission_batch
        self.id = self.records[0].id
        self.pk = self.id
        self.member = self.records[0].member
        self.account = self.records[0].account
        statuses = {item.status for item in self.records}
        self.status = statuses.pop() if len(statuses) == 1 else 'MIXED'
        self.amount = sum((item.amount for item in self.records), Decimal('0.00'))
        for _label, field_name in self.purpose_fields:
            setattr(
                self,
                field_name,
                sum(
                    (getattr(item, field_name) or Decimal('0.00') for item in self.records),
                    Decimal('0.00'),
                ),
            )
        self.payment_weeks = sorted({
            item.payment_week
            for item in self.records
            if item.payment_week and item.saving_amount > 0
        }) or sorted({
            item.payment_week for item in self.records if item.payment_week
        })
        self.proof = next((item.proof for item in self.records if item.proof), None)
        self.remarks = next((item.remarks for item in self.records if item.remarks), '')
        self.date_submitted = min(item.date_submitted for item in self.records)
        reviewed_dates = [item.date_reviewed for item in self.records if item.date_reviewed]
        self.date_reviewed = max(reviewed_dates) if reviewed_dates else None
        self.payment_date = self.records[0].payment_date
        self.payment_time = self.records[0].payment_time
        self.edits = 0

    def __eq__(self, other):
        if isinstance(other, DepositSubmission):
            return len(self.records) == 1 and self.id == other.id
        if isinstance(other, DepositBatchSummary):
            return self.submission_batch == other.submission_batch
        return NotImplemented

    def purpose_breakdown(self):
        return {
            label: getattr(self, field_name)
            for label, field_name in self.purpose_fields
            if getattr(self, field_name) > 0
        }


def _deposit_batch_summaries(queryset):
    grouped = {}
    for deposit in queryset.order_by('submission_batch', 'id'):
        grouped.setdefault(deposit.submission_batch, []).append(deposit)
    summaries = [DepositBatchSummary(records) for records in grouped.values()]
    summaries.sort(
        key=lambda item: (
            item.date_reviewed is not None,
            item.date_reviewed or item.date_submitted,
            item.date_submitted,
            item.id,
        ),
        reverse=True,
    )
    record_ids = [record_id for item in summaries for record_id in item.record_ids]
    edit_counts = {
        row['object_id']: row['count']
        for row in (
            FinancialRecordRevision.objects
            .filter(record_type='deposit', object_id__in=record_ids)
            .values('object_id')
            .annotate(count=Count('id'))
        )
    }
    for item in summaries:
        item.edits = sum(edit_counts.get(record_id, 0) for record_id in item.record_ids)
    return summaries



@login_required
def manage_deposits(request):
    if not request.user.is_treasurer():
        messages.error(request, "Access denied.")
        return redirect('member_dashboard')

    # Keep the review queue and the submission history together on this page.
    # Pending records still expose approve/reject actions, while approved and
    # rejected records remain visible for a complete audit trail.
    search_query = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip().upper()
    valid_statuses = {value for value, _label in DepositSubmission.STATUS_CHOICES}
    if status_filter not in valid_statuses:
        status_filter = ''

    deposit_submissions_base = (
        DepositSubmission.objects
        .filter(member__is_superuser=False)
        .select_related('member', 'account')
    )
    status_counts = deposit_submissions_base.aggregate(
        all=Count('submission_batch', distinct=True),
        pending=Count(
            'submission_batch', filter=Q(status='PENDING'), distinct=True
        ),
        approved=Count(
            'submission_batch', filter=Q(status='APPROVED'), distinct=True
        ),
        rejected=Count(
            'submission_batch', filter=Q(status='REJECTED'), distinct=True
        ),
    )
    deposit_submissions = deposit_submissions_base
    if search_query:
        deposit_submissions = deposit_submissions.filter(
            Q(member__first_name__icontains=search_query)
            | Q(member__last_name__icontains=search_query)
            | Q(member__username__icontains=search_query)
            | Q(member__email__icontains=search_query)
            | Q(member__phone_number__icontains=search_query)
            | Q(account__label__icontains=search_query)
        )
    if status_filter:
        deposit_submissions = deposit_submissions.filter(status=status_filter)
    deposit_submissions_page = Paginator(
        _deposit_batch_summaries(deposit_submissions),
        25,
    ).get_page(request.GET.get('page'))
    form = DirectDepositForm(request.POST or None, request.FILES or None)
    group_settings = GroupSettings.get_active()
    active_saving_year = (
        current_saving_week(
            group_settings.week_one_start, timezone.localdate()
        ).saving_year
        if group_settings else timezone.localdate().year
    )
    close_state = financial_year(active_saving_year)
    close_pending_count = pending_at_cutoff(close_state).count()

    if request.method == 'POST':
        form = DirectDepositForm(request.POST, request.FILES)
        if form.is_valid():
            member = form.cleaned_data['member']
            account = form.cleaned_data.get('account')
            proof = form.cleaned_data.get('proof')
            remarks = form.cleaned_data.get('remarks', '')
            payment_date = form.cleaned_data['payment_date']
            payment_time = form.cleaned_data['payment_time']
            selected_weeks = form.cleaned_data['selected_week_dates']
            weekly_allocations = form.cleaned_data.get('weekly_allocations') or []
            welfare_amount = form.cleaned_data.get('welfare_amount') or 0
            annual_subscription_amount = form.cleaned_data.get('annual_subscription_amount') or 0
            membership_amount = form.cleaned_data.get('membership_amount') or 0
            fine_amount = form.cleaned_data.get('fine_amount') or 0
            fine_allocations = form.cleaned_data.get('fine_allocations') or []
            shares_amount = form.cleaned_data.get('shares_amount') or 0
            loan_repayment_amount = form.cleaned_data.get('loan_repayment_amount') or 0
            loan_repayment_loan = form.cleaned_data.get('loan_repayment_loan')
            welfare_weeks = form.cleaned_data.get('welfare_week_dates') or []
            selected_years = {week.year for week in selected_weeks + welfare_weeks}
            selected_years.update(
                week.year
                for week in Fine.objects.filter(
                    id__in=[fine_id for fine_id, _amount in fine_allocations],
                    reference_week__isnull=False,
                ).values_list('reference_week', flat=True)
            )
            if any(submissions_locked_for_year(year) for year in selected_years):
                form.add_error(
                    None,
                    'One or more selected weeks belong to a closed financial year.',
                )
            if loan_repayment_loan and loan_activity_frozen(loan_repayment_loan):
                form.add_error(
                    'loan_repayment_amount',
                    'This loan is frozen for financial-year settlement.',
                )
            if form.errors:
                return render(request, 'deposits/manage_deposits.html', {
                    'deposit_submissions': deposit_submissions_page,
                    'page_obj': deposit_submissions_page,
                    'pagination_query': pagination_query(request),
                    'form': form,
                    'search_query': search_query,
                    'status_filter': status_filter,
                    'status_counts': status_counts,
                    'close_state': close_state,
                    'active_saving_year': active_saving_year,
                    'close_pending_count': close_pending_count,
                })

            deposits_created = []
            submission_batch = uuid.uuid4()
            with transaction.atomic():
                allocation_by_week = dict(weekly_allocations)
                for index, payment_week in enumerate(selected_weeks):
                    first = index == 0
                    deposit = DepositSubmission(
                        submission_batch=submission_batch,
                        member=member, account=account, submitted_by=request.user,
                        reviewed_by=request.user, payment_week=payment_week,
                        starting_week=payment_week, weeks_covered=1,
                        saving_amount=allocation_by_week.get(payment_week, Decimal('0.00')),
                        welfare_amount=welfare_amount if first else 0,
                        annual_subscription_amount=annual_subscription_amount if first else 0,
                        membership_amount=membership_amount if first else 0,
                        fine_amount=fine_amount if first else 0,
                        shares_amount=shares_amount if first else 0,
                        loan_repayment_amount=loan_repayment_amount if first else 0,
                        loan_repayment_loan=loan_repayment_loan if first else None,
                        proof=proof if first else None, remarks=remarks,
                        status='APPROVED', payment_date=payment_date,
                        payment_time=payment_time, date_reviewed=timezone.now(),
                    )
                    deposit.full_clean()
                    deposit.save()
                    deposits_created.append(deposit)
                    if first:
                        for fine_id, amount in fine_allocations:
                            DepositFineAllocation.objects.create(
                                deposit=deposit,
                                fine_id=fine_id,
                                amount=amount,
                            )
                        _create_welfare_allocations(deposit, account, welfare_weeks)
                if fine_allocations:
                    apply_selected_fine_allocations(
                        (allocation.fine, allocation.amount)
                        for allocation in deposits_created[0].fine_allocations.select_related('fine').all()
                    )
                for saved_deposit in deposits_created:
                    if saved_deposit.saving_amount > 0:
                        delete_deposit_week_missed_saving_fines(saved_deposit)
                deposit = deposits_created[0]
                if not fine_allocations and deposit.fine_amount > 0:
                    allocate_fine_payment(deposit.member, deposit.account, deposit.fine_amount)
                if deposit.loan_repayment_amount and deposit.loan_repayment_amount > 0:
                    allocated, unallocated = _apply_loan_repayment(
                        member=deposit.member,
                        account=deposit.account,
                        loan=deposit.loan_repayment_loan,
                        repayment_amount=deposit.loan_repayment_amount,
                        paid_on=deposit.payment_date,
                        recorded_by=request.user,
                        source_deposit=deposit,
                        notes=f'Direct deposit repayment (Deposit #{deposit.id}).',
                    )
                    if unallocated > 0:
                        raise ValidationError('Loan repayment exceeds the selected loan balance.')

            if weekly_allocations:
                messages.success(
                    request,
                    f"Savings deposit for {member.username} allocated across "
                    f"{len(selected_weeks)} selected week(s).",
                )
            else:
                messages.success(request, f"Deposit for {member.username} recorded successfully.")
            return redirect('manage_deposits')

    return render(request, 'deposits/manage_deposits.html', {
        'deposit_submissions': deposit_submissions_page,
        'page_obj': deposit_submissions_page,
        'pagination_query': pagination_query(request),
        'form': form,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_counts': status_counts,
        'close_state': close_state,
        'active_saving_year': active_saving_year,
        'close_pending_count': close_pending_count,
    })

@login_required
def treasurer_reports(request):
    selected_year = parse_report_year(request.GET.get('year'))
    search_query = (request.GET.get('q') or '').strip()
    approved_deposits_base = DepositSubmission.objects.filter(status='APPROVED', member__is_superuser=False)
    years = merge_year_options(
        years_from_dates(approved_deposits_base, 'payment_week'),
        years_from_dates(LoanRequest.objects.filter(status='APPROVED'), 'approved_on'),
        selected_year=selected_year,
    )
    members = MemberProfile.objects.exclude(is_superuser=True)
    if search_query:
        members = members.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(savings_accounts__label__icontains=search_query)
        ).distinct()
    members = alphabetical_members(members)
    members_page = Paginator(members, 25).get_page(request.GET.get('page'))

    report_data = []
    for member in members_page:
        approved_deposits = member.deposits.filter(
            status='APPROVED',
            payment_week__year=selected_year,
        ).select_related('account')
        total_amount = approved_deposits.aggregate(Sum('amount'))['amount__sum'] or 0
        total_saving = approved_deposits.aggregate(Sum('saving_amount'))['saving_amount__sum'] or 0
        total_welfare = approved_deposits.aggregate(Sum('welfare_amount'))['welfare_amount__sum'] or 0
        total_annual = approved_deposits.aggregate(Sum('annual_subscription_amount'))['annual_subscription_amount__sum'] or 0
        total_membership = approved_deposits.aggregate(Sum('membership_amount'))['membership_amount__sum'] or 0
        total_fine = approved_deposits.aggregate(Sum('fine_amount'))['fine_amount__sum'] or 0
        total_shares = approved_deposits.aggregate(Sum('shares_amount'))['shares_amount__sum'] or 0
        total_interest = _loan_interest_total(
            member.loan_requests.filter(status='APPROVED', approved_on__year=selected_year)
        )
        total_weeks = approved_deposits.count()
        account_labels = list(
            approved_deposits.exclude(account__isnull=True)
            .order_by('account__label')
            .values_list('account__label', flat=True)
            .distinct()
        )
        if not account_labels:
            account_labels = list(
                member.savings_accounts.filter(is_active=True)
                .order_by('label')
                .values_list('label', flat=True)
            )

        report_data.append({
            'member': member,
            'savings_account': ', '.join(account_labels) if account_labels else '-',
            'total_amount': total_amount,
            'total_weeks': total_weeks,
            'total_saving': total_saving,
            'total_welfare': total_welfare,
            'total_annual': total_annual,
            'total_membership': total_membership,
            'total_fine': total_fine,
            'total_shares': total_shares,
            'total_interest': total_interest,
        })

    return render(request, 'deposits/treasurer_reports.html', {
        'report_data': report_data,
        'page_obj': members_page,
        'pagination_query': pagination_query(request),
        'selected_year': selected_year,
        'years': years,
        'search_query': search_query,
    })


@login_required
def treasurer_week_options(request):
    if not request.user.is_treasurer():
        return JsonResponse({'error': 'Access denied.'}, status=403)
    member = get_object_or_404(
        MemberProfile, pk=request.GET.get('member'), is_superuser=False
    )
    account_id = request.GET.get('account')
    active_accounts = member.savings_accounts.filter(is_active=True)
    account = active_accounts.filter(pk=account_id).first() if account_id else None
    if not account:
        possible_accounts = list(active_accounts.order_by('id')[:2])
        if len(possible_accounts) == 1:
            account = possible_accounts[0]
    group_settings = GroupSettings.get_active()
    if not group_settings:
        return JsonResponse({'weeks': []})

    today = timezone.localdate()
    _saving_week, saving_weeks = saving_year_weeks(group_settings.week_one_start, today)
    paid_totals = weekly_savings_totals_by_week(member, account, saving_weeks)
    savings_statuses = saving_week_statuses(member, account, saving_weeks, today)
    # Fine obligations may come from a prior saving cycle, so expose all
    # outstanding/paid fine weeks rather than limiting them to current-year
    # savings dates.
    fine_states = fine_week_options(member, account)
    weeks = []
    for index, week in enumerate(saving_weeks, start=1):
        total = paid_totals.get(week, Decimal('0.00'))
        savings_state = savings_statuses.get(week, {})
        status = savings_state.get('status', 'available')
        paid = status in ('paid_on_time', 'paid_late')
        is_future = week > today
        is_current = week <= today < week + timedelta(days=7)
        selectable = not paid
        date_label = f'{week:%A} {week.day} {week:%B %Y}'
        fine_state = fine_states.get(week, {})
        weeks.append({
            'date': week.isoformat(),
            'paid': paid,
            'paid_amount': str(total),
            'savings_status': status,
            'completion_date': savings_state.get('completion_date').isoformat() if savings_state.get('completion_date') else None,
            'deadline': savings_state.get('deadline').isoformat() if savings_state.get('deadline') else None,
            'fine_status': fine_state.get('status', 'none'),
            'fine_outstanding': str(fine_state.get('outstanding', Decimal('0.00'))),
            'fine_ids': fine_state.get('outstanding_ids', []),
            'label': f'Week {index} · {week:%A %d %b}',
            'date_label': date_label,
            'accessible_label': f'Week {index}, {date_label}',
            'week_number': index,
            'month': week.month,
            'month_name': week.strftime('%B'),
            'year': week.year,
            'day': week.day,
            'is_future': is_future,
            'is_current': is_current,
            'is_available': selectable,
            'selectable': selectable,
        })
    fines = []
    for week, state in fine_states.items():
        fines.append({
            'date': week.isoformat(),
            'label': f'Fine · {week:%A %d %b %Y}',
            'accessible_label': f'Fine for {week:%A} {week.day} {week:%B %Y}',
            'status': state['status'],
            'outstanding': str(state['outstanding']),
            'selectable': state['selectable'],
            'fine_ids': state['outstanding_ids'],
            'fine_allocations': [
                {'id': item['id'], 'amount': str(item['amount'])}
                for item in state['fine_allocations']
            ],
        })
    welfare = [
        {
            'date': item['friday'].isoformat(),
            'label': f"Welfare Week {item['number']} · {item['friday']:%d %b %Y}",
            'month': f"{item['friday']:%B %Y}",
            'day': item['friday'].day,
            'status': item['status'],
            'selectable': item['selectable'],
        }
        for item in build_welfare_calendar(member, account, today).get('weeks', [])
    ]
    loans = []
    if account:
        for loan in LoanRequest.objects.filter(
            member=member, account=account, status='APPROVED'
        ).prefetch_related('repayments').order_by('approved_on', 'id'):
            balance = loan.outstanding_balance_as_of(today)
            if balance > 0:
                loans.append({
                    'id': loan.id,
                    'label': f'Loan #{loan.id} · Balance UGX {balance:,.0f}',
                    'balance': str(balance),
                })
    return JsonResponse({'weeks': weeks, 'fines': fines, 'welfare': welfare, 'loans': loans})



def download_member_report(request, member_id, format):
    selected_year = parse_report_year(request.GET.get('year'))
    member = get_object_or_404(MemberProfile, id=member_id)
    deposits = member.deposits.filter(
        status='APPROVED',
        payment_week__year=selected_year,
    ).select_related('account').order_by('payment_week')

    if format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{member.username}_report_{selected_year}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        cell_style = styles['BodyText'].clone('report-cell')
        cell_style.fontSize = 7
        cell_style.leading = 8
        header_style = styles['BodyText'].clone('report-header')
        header_style.fontSize = 7
        header_style.leading = 8
        header_style.textColor = colors.white
        header_style.alignment = 1
        elements = []

        # Title
        full_name = member.get_full_name() or member.username
        title = Paragraph(f"<b>Contribution Report for {full_name} - {selected_year}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Table header
        headers = [
            '#', 'Week', 'Savings Account', 'Total', 'Saving', 'Welfare', 'Annual',
            'Membership', 'Fine', 'Shares', 'Date Submitted', 'Payment Date',
            'Payment Time', 'Remarks',
        ]
        data = [[_pdf_text(header, header_style) for header in headers]]

        # Table rows
        for i, deposit in enumerate(deposits, start=1):
            data.append([_pdf_text(value, cell_style) for value in [
                i,
                deposit.payment_week.strftime('%Y-%m-%d'),
                deposit.account.label if deposit.account else '-',
                f"{deposit.amount:,.0f}",
                f"{deposit.saving_amount:,.0f}",
                f"{deposit.welfare_amount:,.0f}",
                f"{deposit.annual_subscription_amount:,.0f}",
                f"{deposit.membership_amount:,.0f}",
                f"{deposit.fine_amount:,.0f}",
                f"{deposit.shares_amount:,.0f}",
                deposit.date_submitted.strftime('%Y-%m-%d'),
                deposit.payment_date.strftime('%Y-%m-%d'),
                deposit.payment_time.strftime('%H:%M'),
                deposit.remarks or '-',
            ]])

        # Table styling
        table = Table(
            data,
            repeatRows=1,
            colWidths=[18, 52, 64, 48, 42, 42, 42, 52, 38, 42, 58, 54, 48, 120],
        )
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    elif format == 'excel':
        

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Member Contributions"

        full_name = member.get_full_name() or member.username
        ws.merge_cells('A1:N1')
        ws['A1'] = f"Contribution Report for {full_name} - {selected_year}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Optional: Add group logo
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            img.height = 60
            img.width = 60
            ws.add_image(img, 'F1')

        # Define headers
        headers = ['#', 'Week', 'Savings Account', 'Total (UGX)', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine', 'Shares', 'Date Submitted', 'Payment Date', 'Payment Time', 'Remarks']
        ws.append(headers)

        # Header styling
        header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Fill data rows
        total_amount = 0
        for i, deposit in enumerate(deposits, start=1):
            row = [
                i,
                deposit.payment_week.strftime('%Y-%m-%d'),
                deposit.account.label if deposit.account else '-',
                float(deposit.amount),
                float(deposit.saving_amount),
                float(deposit.welfare_amount),
                float(deposit.annual_subscription_amount),
                float(deposit.membership_amount),
                float(deposit.fine_amount),
                float(deposit.shares_amount),
                deposit.date_submitted.strftime('%Y-%m-%d'),
                deposit.payment_date.strftime('%Y-%m-%d'),
                deposit.payment_time.strftime('%H:%M'),
                deposit.remarks or '-'
            ]
            ws.append(row)
            total_amount += float(deposit.amount)

        last_data_row = 2 + len(deposits)

        # Total row
        total_label_cell = ws.cell(row=last_data_row + 1, column=2, value="TOTAL")
        total_label_cell.font = Font(bold=True)
        total_label_cell.alignment = Alignment(horizontal='right')

        ws.cell(row=last_data_row + 1, column=3, value=total_amount).font = Font(bold=True)

        # Footer with generated timestamp
        ws.merge_cells(start_row=last_data_row + 3, start_column=1, end_row=last_data_row + 3, end_column=14)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        footer_cell = ws.cell(row=last_data_row + 3, column=1)
        footer_cell.value = f"Generated on: {timestamp}"
        footer_cell.font = Font(italic=True, size=10)
        footer_cell.alignment = Alignment(horizontal='right')

        # Auto column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 36)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center' if cell.row == 2 else 'left',
                    vertical='top',
                    wrap_text=True,
                )

        # Download response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{member.username}_report_{selected_year}.xlsx"'
        wb.save(response)
        return response


def download_all_reports(request, format):
    selected_year = parse_report_year(request.GET.get('year'))
    search_query = (request.GET.get('q') or '').strip()
    members = MemberProfile.objects.filter(is_superuser=False)
    if search_query:
        members = members.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(savings_accounts__label__icontains=search_query)
        ).distinct()
    members = alphabetical_members(members)

    if format == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        cell_style = styles['BodyText'].clone('all-report-cell')
        cell_style.fontSize = 7
        cell_style.leading = 8
        header_style = styles['BodyText'].clone('all-report-header')
        header_style.fontSize = 7
        header_style.leading = 8
        header_style.textColor = colors.white
        header_style.alignment = 1
        elements = [
            Paragraph(f"Group Contribution Report (All Members) - {selected_year}", styles['Heading1']),
            Paragraph(f"Generated on: {now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']),
            Spacer(1, 12),
        ]

        for member in members:
            deposits = member.deposits.filter(
                status='APPROVED',
                payment_week__year=selected_year,
            ).select_related('account').order_by('payment_week')
            if not deposits.exists():
                continue

            elements.append(Paragraph(f"Member: {member.get_full_name() or member.username}", styles['Heading3']))
            headers = ['Week', 'Savings Account', 'Total', 'Saving', 'Welfare', 'Annual', 'Membership', 'Fine', 'Shares', 'Payment Date', 'Payment Time']
            data = [[_pdf_text(header, header_style) for header in headers]]
            total_amount = 0

            for dep in deposits:
                data.append([_pdf_text(value, cell_style) for value in [
                    dep.payment_week.strftime('%Y-%m-%d'),
                    dep.account.label if dep.account else '-',
                    f"{dep.amount:,.0f}",
                    f"{dep.saving_amount:,.0f}",
                    f"{dep.welfare_amount:,.0f}",
                    f"{dep.annual_subscription_amount:,.0f}",
                    f"{dep.membership_amount:,.0f}",
                    f"{dep.fine_amount:,.0f}",
                    f"{dep.shares_amount:,.0f}",
                    dep.payment_date.strftime('%Y-%m-%d'),
                    dep.payment_time.strftime('%H:%M'),
                ]])
                total_amount += dep.amount

            data.append([_pdf_text(value, cell_style) for value in ['TOTAL', '', f"{total_amount:,.0f}", '', '', '', '', '', '', '', '']])
            table = Table(data, repeatRows=1, colWidths=[58, 64, 54, 48, 48, 48, 54, 42, 46, 62, 58])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 24))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="all_member_reports_{selected_year}.pdf"'
        return response

    if format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Contributions {selected_year}"

        ws.merge_cells('A1:L1')
        ws['A1'] = f"Group Contribution Report (All Members) - {selected_year}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append(["Generated on:", now().strftime('%Y-%m-%d %H:%M')])
        ws.append([])
        ws.append(["Member", "Week", "Savings Account", "Total (UGX)", "Saving", "Welfare", "Annual", "Membership", "Fine", "Shares", 'Payment Date', 'Payment Time'])

        for member in members:
            deposits = member.deposits.filter(
                status='APPROVED',
                payment_week__year=selected_year,
            ).select_related('account').order_by('payment_week')
            total_amount = 0

            for dep in deposits:
                ws.append([
                    member.get_full_name() or member.username,
                    dep.payment_week.strftime('%Y-%m-%d'),
                    dep.account.label if dep.account else '-',
                    float(dep.amount),
                    float(dep.saving_amount),
                    float(dep.welfare_amount),
                    float(dep.annual_subscription_amount),
                    float(dep.membership_amount),
                    float(dep.fine_amount),
                    float(dep.shares_amount),
                    dep.payment_date.strftime('%Y-%m-%d'),
                    dep.payment_time.strftime('%H:%M'),
                ])
                total_amount += float(dep.amount)

            if deposits.exists():
                ws.append([
                    f"TOTAL for {member.get_full_name() or member.username}", "", "", total_amount,
                    "", "", "", "", "", "", "", "",
                ])
                ws.append([])

        for i, column_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = min(max_length + 4, 36)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='center' if cell.row == 4 else 'left',
                    vertical='top',
                    wrap_text=True,
                )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="all_member_reports_{selected_year}.xlsx"'
        return response

    return HttpResponse("Invalid format", status=400)
    

def _can_view_current_week_status(user):
    return (
        user.is_treasurer()
        or user.is_mobilizer()
        or user.is_chairman()
        or user.is_vice_chairman()
        or user.is_overseer()
    )


def _current_week_settings_redirect(request):
    group_settings = GroupSettings.get_active()
    if group_settings:
        return group_settings, None
    if request.user.is_superuser or request.user.is_treasurer() or request.user.is_chairman():
        messages.error(request, "Open the saving cycle in Group Settings before checking current payments.")
        return None, redirect('group_settings')
    messages.error(request, "The saving cycle has not been opened yet. Please contact the Treasurer.")
    return None, redirect('member_dashboard')


def _status_entry(member, account, has_paid):
    return {
        'member': member,
        'member_name': member.get_full_name() or member.username,
        'account': account,
        'account_label': account.label if account else '-',
        'has_paid': has_paid,
        'status_label': 'Paid' if has_paid else 'Not Paid',
    }


def _current_week_payment_status_data(request, create_fines=False):
    group_settings = GroupSettings.get_active()
    saving_week = current_saving_week(group_settings.week_one_start, timezone.localdate())
    current_week_start = saving_week.week_start
    fines_can_be_created = create_fines and missed_saving_fines_can_be_created(current_week_start)
    grace_ends_on = current_week_start + timedelta(days=2)

    paid_entries = []
    unpaid_entries = []

    search_query = (request.GET.get('q') or '').strip()
    members = MemberProfile.objects.filter(is_superuser=False)
    if search_query:
        members = members.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(username__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(phone_number__icontains=search_query)
            | Q(savings_accounts__label__icontains=search_query)
        ).distinct()
    members = alphabetical_members(members)
    for member in members:
        member_accounts = list(SavingsAccount.objects.filter(owner=member, is_active=True).order_by('label'))
        accounts_to_check = member_accounts or [None]

        for account in accounts_to_check:
            deposit_filter = {
                'status': 'APPROVED',
                'payment_week': current_week_start,
            }
            if account:
                deposit_filter['account'] = account
            else:
                deposit_filter['account__isnull'] = True

            paid_total, completion = completion_for_week(
                member,
                account,
                current_week_start,
                statuses=('PENDING', 'APPROVED'),
            )
            has_paid = paid_total >= MIN_WEEKLY_SAVINGS
            entry = _status_entry(member, account, has_paid)
            if has_paid:
                paid_entries.append(entry)
                continue

            unpaid_entries.append(entry)
            if fines_can_be_created and not (
                completion and completion <= week_deadline(current_week_start)
            ):
                account_note = f" for account {account.label}" if account else ""
                Fine.objects.get_or_create(
                    member=member,
                    account=account,
                    fine_type='MISSED_WEEKLY_SAVING',
                    reference_week=current_week_start,
                    defaults={
                        'reason': f'Failed to save{account_note} for week closing {current_week_start}',
                        'amount': LATE_FINE_AMOUNT,
                        'issued_by': request.user,
                    }
                )

    return {
        'current_week': current_week_start,
        'current_week_number': saving_week.week_number,
        'current_saving_year': saving_week.saving_year,
        'paid_entries': paid_entries,
        'unpaid_entries': unpaid_entries,
        'all_entries': paid_entries + unpaid_entries,
        'fines_can_be_created': fines_can_be_created,
        'grace_ends_on': grace_ends_on,
        'search_query': search_query,
    }


def _current_week_status_filename(data, extension):
    scope = data.get('export_scope')
    scope_part = f'_{scope}_accounts' if scope in {'paid', 'unpaid'} else ''
    return (
        f"current_week_payment_status{scope_part}_week_"
        f"{data['current_week_number']}_{data['current_week'].strftime('%Y-%m-%d')}.{extension}"
    )


def _export_current_week_status_excel(data):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_current_week_status_filename(data, "xlsx")}"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Current Payments'
    ws['A1'] = f"Payment Status for Week {data['current_week_number']} Closing {data['current_week'].strftime('%Y-%m-%d')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')

    headers = ['#', 'Member', 'Account', 'Status', 'Week Closing']
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_number, entry in enumerate(data['all_entries'], start=4):
        ws.cell(row=row_number, column=1, value=row_number - 3)
        ws.cell(row=row_number, column=2, value=entry['member_name'])
        ws.cell(row=row_number, column=3, value=entry['account_label'])
        ws.cell(row=row_number, column=4, value=entry['status_label'])
        ws.cell(row=row_number, column=5, value=data['current_week'].strftime('%Y-%m-%d'))

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 42)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(response)
    return response


def _export_current_week_status_pdf(data):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_current_week_status_filename(data, "pdf")}"'

    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    scope = data.get('export_scope')
    if scope == 'paid':
        entries = data['paid_entries']
        title = f"Paid Accounts for Week {data['current_week_number']}"
    elif scope == 'unpaid':
        entries = data['unpaid_entries']
        title = f"Unpaid Accounts for Week {data['current_week_number']}"
    else:
        entries = data['all_entries']
        title = f"Payment Status for Week {data['current_week_number']}"
    elements = [
        Paragraph(title, styles['Title']),
        Paragraph(f"Week Closing: {data['current_week'].strftime('%A, %d %b %Y')}", styles['Normal']),
        Paragraph(
            (
                f"Accounts: {len(entries)}"
                if scope in {'paid', 'unpaid'}
                else f"Paid: {len(data['paid_entries'])} | Unpaid: {len(data['unpaid_entries'])}"
            ),
            styles['Normal'],
        ),
        Spacer(1, 12),
    ]

    table_rows = [['#', 'Member', 'Account', 'Status']]
    for index, entry in enumerate(entries, start=1):
        table_rows.append([index, entry['member_name'], entry['account_label'], entry['status_label']])
    if len(table_rows) == 1:
        table_rows.append(['-', 'No members found', '-', '-'])

    table = Table(table_rows, repeatRows=1, colWidths=[35, 210, 120, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def export_current_week_payment_status(request, format):
    if not _can_view_current_week_status(request.user):
        messages.error(request, "Access denied.")
        return redirect('member_dashboard')

    _group_settings, redirect_response = _current_week_settings_redirect(request)
    if redirect_response:
        return redirect_response

    data = _current_week_payment_status_data(request, create_fines=False)
    scope = (request.GET.get('scope') or '').strip().lower()
    if scope:
        if scope not in {'paid', 'unpaid'}:
            return HttpResponse('Invalid payment status scope', status=400)
        if format == 'pdf':
            data['export_scope'] = scope
    if format == 'excel':
        return _export_current_week_status_excel(data)
    if format == 'pdf':
        return _export_current_week_status_pdf(data)
    return HttpResponse("Invalid format", status=400)


@login_required
def current_week_payment_status(request):
    if not _can_view_current_week_status(request.user):
        messages.error(request, "Access denied.")
        return redirect('member_dashboard')

    _group_settings, redirect_response = _current_week_settings_redirect(request)
    if redirect_response:
        return redirect_response

    context = _current_week_payment_status_data(request, create_fines=True)
    return render(request, 'deposits/current_week_status.html', context)
